#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook
from openpyxl import workbook
from datetime import datetime
import os
import sys
import shutil
import pandas as pd
import numpy as np
import re
import sys
import csv
import openpyxl
import decimal
from os import listdir
from os.path import isfile, join
#######################################################################
#
# CSV extraction file for Ruby Project  
# Author: David Kleppang
# Revision: D
# Date: 08/19/2019
#
# Application will extract information from multiple CSV files for Ruby 
# project and copy into one combined file suitable for importing into 
# SQL server through an import procedure
#
# Rev A: Initial (05/14/2019)
# Rev B: Corrected bug that overwrote mode_indicator variable as it iterated through file list
# Rev C: Corrected bug that prevented session name coming through in manual files (in config.txt file) 
# Rev D: Updated program to accommodate different software version of excel files to import
#        also allowed importing of unicode in autoexport files (rather than string)
#        also updated device date to ensure it is consistent with European format
#       (Toulouse was first site and date was received first in EU format)
# Rev F: Removed elif from read_input_file function to allow csv read when first two lines 
#        in the csv file have a ',' instead of a 1 or 2.
# Rev G: Added functions to build a list of hash keys from the comdined_list and compared it to the list built
#        from the contents of a folder. The comparison ensures that there are no duplicate entries into the dest file
#
#######################################################################

def input_config_data(config):
    '''Function to read in CONFIG.txt file and develop file configuration data from file'''
    with open(config, 'r') as f:
        lines = f.read().splitlines() 
        for line in lines:
            exec(line,globals()) #execute python strings in config file as python commands
        return [TRIALKEY, REGION_DICT, EQUIPSN_SITE_DICT,EQUIPSN_SITE_DICT_REVERSE, SINK_ROW_INIT, SOURCE_ROW_INIT, DEST_DICT, DEST_DICT_REVERSE, BINSEARCH_UB, MAN_FORM_STRUCT, AUTO_FORM_STRUCT, MANUAL_INTERPS_DICT] 

def extract_rawdata(rawdata_filepath):
    '''Function to add each Raw data source name and filepath to a list for further processing'''
    rawdata_filelist = [] #create empty list
    for filename in os.listdir(rawdata_filepath): #iterate thru each file in raw data file folder
        rawdata_filelist.append(rawdata_filepath + filename) #append to list of raw data files c/w full path information
    return rawdata_filelist


#--------------------------------------------------------------NEW FUNCTIONS TO COMPARE FILEHASH---------------------------------------------------

def lookin_folder(root_dir, r_exfile=None):
    '''Looks in a folder via a path and builds a list with files that have extension xlsm'''
    ex_files =[]
    x = 'combined_results.xlsx'
    for file in os.listdir(root_dir):
        if x in file:
            ex_files.append(file)
            #ex_files = [i for i in ex_files if not i.startswith('~$')] # ----------- removes ~$ issue from list
            r_exfile = [root_dir + i for i in ex_files]
    for i in r_exfile:
        if x in i:
            combine_excel = i
    return combine_excel


def df_combined(excel_file1): #-----------------------reads in combined excel workbok and creates a dataframe for processing 
    dfx = pd.read_excel(excel_file1)
    return dfx



def build_combined_hash(dfx):#------------------------------------creating a list of filehash keys for comparison later
    combined_hash_key = dfx.Filename_Hash_Key[1:]
    com_hash =[]
    for i in combined_hash_key:
        com_hash.append(i)
    return com_hash

########################################################### next two functions ensure that there are matching file hash keys for both pdf and csv files an#################
########################################################### also ensures there isn't a file hash key with the read in list that is already in the destination file#########
def extract_pdf_list(rawdata_filelist, com_hash):
    ''' Function to add pdf filenames to a list for csv validation comparison'''#--Use this function to vet the hash list----
    pdf_list = []
    pdf_not =[]
    for rawfile in rawdata_filelist:
        if rawfile[-3:] == 'pdf':
            pdf_list.append(rawfile[-36:-4]) #extract file hash key from pdf files
    for i in pdf_list:
        if i not in com_hash:
            pdf_not.append(i)
    return pdf_not


def extract_csv_list(rawdata_filelist, com_hash):
    ''' Function to add csv filenames to a list for csv validation comparison'''#--Use this function to vet the hash list----
    csv_list = []
    csv_not =[]
    for rawfile in rawdata_filelist:
        if rawfile[-3:] == 'csv':
            csv_list.append(rawfile[-36:-4]) #extract file hash key from csv files
    for i in csv_list:
        if i not in com_hash:
            csv_not.append(i)
    return csv_not    
#--------------------------------------------------------------------------------------------------------------------------------------

def val_csv_pdf(pdf_list, csv_list):#----------------------Ensures that the two pdf and csv lists are identical-----------------------------------------------------------
    val_list = []
    for csv in csv_list:
        for pdf in pdf_list:
            if pdf in csv:
                val_list.append(csv)
    return val_list


#--------compares two lists to get elements out of the larger list from a smaller list by comparing elemets like eachother-----------------------------------------------
def n_rawdata_filelist(rawdata_filelist, val_list):    
    raw_list = []
    for raw in rawdata_filelist: #-------main list---------
        for h_raw in val_list:#----------small list--------------
            if h_raw in raw:
                raw_list.append(raw)
    return raw_list


def v_csv_list(rawdata_filelist):#-----------------builds raw filepathlist of all csv's that have been checked against raw list------------------------------------------
    val_csv_list=[]
    for val in rawdata_filelist:
        if '('in val:
            continue
        if val[-3:]=='csv':
            val_csv_list.append(val)
    return val_csv_list

###########################################################################################################################################################################
def read_input_file(rawdatafile_list, pdf_list):
    ''' Function to read manual csv files.  The function works by inputting a read-only list of all files to compare in sequential
        form'''
    file_list = []
    for rawfile in rawdatafile_list:
        if '(' in rawfile: #in case there is any duplicate files which usually have a '(n)' in the file name
            continue
        rawfile_hash =  rawfile[-36:-4] #extract file hash from csv files
        if rawfile[-3:] == 'csv' and rawfile_hash in pdf_list: #make sure that there is a corresponding pdf file with each csv file
            with open(rawfile, 'r') as f:
                lines = f.read().splitlines()
                lines = lines[:-1]# fixes out-of-range issue
                for line in lines:
                    try:
                        if line[0] == '1' or line[0] == '2':
                            line = (line + ',' + rawfile_hash).split(',')
                            file_list.append(line)
                    except:
                           print('Error in csv read.')
    return file_list

def read_autoexport_input_file(rawdatafile_list, SOURCE_ROW_INIT):
    ''' Function to read auto xlsx file.  The function works by inputting a read-only list of all files to compare in sequential
        form'''
    autofile_list = []
    SOURCECOLUMN_INITIAL = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    SOURCECOLUMN_A = ['AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']
    SOURCECOLUMN_B = ['BA','BB','BC','BD']
    for rawfile in rawdatafile_list:
        filename_forslash = rawfile.split('/')[-1] #split filepath at forward slash to get filename
        filename_period = filename_forslash[:-5] #split filename at period to strip
        source_row = SOURCE_ROW_INIT
        if '(' in rawfile: #in case there is any duplicate files which usually have a '(n)' in the file name
            continue
        if rawfile[-4:] == 'xlsx' and ('/A_' in str(rawfile).upper() or '/B_' in str(rawfile).upper()): #make sure there is a _A or a _B in the filename
            wb = load_workbook(filename = rawfile)
            autosheet = wb['sheet']
            while str(autosheet['A' + str(source_row)].value) != "None": 
                row_list = []
                sourcecolumn = SOURCECOLUMN_INITIAL if '_report-' in filename_period else SOURCECOLUMN_INITIAL + SOURCECOLUMN_A + SOURCECOLUMN_B
                for column in sourcecolumn:
                    row_list.append(autosheet[column + str(source_row)].value) #Updated to unicode due to presence of new columns for Indiana trial which included these non-ASCII characters
                row_list.append(filename_period) #add filename on to end of each row
                autofile_list.append(row_list)
                source_row += 1
    return autofile_list

def find_insertion_point(wsdest, BINSEARCH_UB, SINK_ROW_INIT):
    ''' Function using binary search to determine insertion point for destination workbook'''
    low = SINK_ROW_INIT
    high = BINSEARCH_UB 
    i=0
    while low <= high:
        mid = (low + high) // 2
        #insertion_point is final mid in this algorithm
        i+=1
        if str(wsdest['A' + str(mid)].value) == "None" and str(wsdest['A' + str(mid-1)].value) != "None":
            return mid
        elif str(wsdest['A' + str(mid)].value) == "None":
            high = mid - 1
        else:
            low = mid + 1
    return None

def add_manual_results(orig_file, wsdest, destfile, insertion_point, DEST_DICT, DEST_DICT_REVERSE, MAN_FORM_STRUCT, SINK_ROW_INIT, TRIALKEY, EQUIPSN_SITE_DICT_REVERSE, MANUAL_INTERPS_DICT): #This is for valid data results
    ''' Function to iterate through each csv file and 
    pull data from relevant fields to insert into destfile'''
    manual_feature_dict = {}
    dest_feature_dict = {}
    for result in orig_file:
        for index, manual_feature in enumerate(MAN_FORM_STRUCT):
            manual_feature_dict[manual_feature] = result[index]
        for dest_feature in DEST_DICT:
            if dest_feature in manual_feature_dict:
                dest_feature_value = DEST_DICT.get(dest_feature, "")
                manual_feature_value = manual_feature_dict.get(dest_feature, "")
                dest_feature_dict[dest_feature_value] =  manual_feature_value
        for column in DEST_DICT_REVERSE:
            if column in dest_feature_dict:
                if column == DEST_DICT.get("IsExcluded", "") and dest_feature_dict.get(column, "") == 'true':
                    wsdest[column + str(insertion_point)] = 1 
                elif column == DEST_DICT.get("IsExcluded", "") and dest_feature_dict.get(column, "") == 'false':
                    wsdest[column + str(insertion_point)] = 0 
                elif column == DEST_DICT.get("DataEntryPersonNumber", "") and dest_feature_dict.get(column, "") == '1':
                    wsdest[column + str(insertion_point)] = 1 
                elif column == DEST_DICT.get("DataEntryPersonNumber", "") and dest_feature_dict.get(column, "") == '2':
                    wsdest[column + str(insertion_point)] = 2 
                elif column == DEST_DICT.get("Well_Positivity", ""):
                    wsdest[DEST_DICT.get("Well_Positivity", "") + str(insertion_point)] = manual_feature_dict.get("Well_Positivity", "") 
                    wsdest[DEST_DICT.get("Well_Result", "") + str(insertion_point)] = MANUAL_INTERPS_DICT.get(manual_feature_dict.get("Well_Positivity", ""), "") 
                else:
                    wsdest[column + str(insertion_point)] = dest_feature_dict.get(column, "")
        wsdest[DEST_DICT.get("RowID", "") + str(insertion_point)] = insertion_point - SINK_ROW_INIT + 1 
        wsdest[DEST_DICT.get("Device_SerialNum", "") + str(insertion_point)] =  EQUIPSN_SITE_DICT_REVERSE.get(wsdest[DEST_DICT.get("SiteCode", "") + str(insertion_point)].value, "")
        wsdest[DEST_DICT.get("TrialKey", "") + str(insertion_point)] = TRIALKEY 
        wsdest[DEST_DICT.get("DateLoaded", "") + str(insertion_point)] = datetime.now() 
        insertion_point += 1
    return insertion_point

def add_auto_results(orig_file, wsdest, destfile, insertion_point, DEST_DICT, DEST_DICT_REVERSE, AUTO_FORM_STRUCT, SINK_ROW_INIT, TRIALKEY, EQUIPSN_SITE_DICT, REGION_DICT): #This is for valid data results
    ''' Function to iterate through each excel file and 
    pull data from relevant fields to insert into destfile'''
    auto_feature_dict = {}
    dest_feature_dict = {}
    for result in orig_file:
        UI_ver = result[1] #select data attributes from AUTO_FORM_STRUCT dict based on UI Version (This is result[1] which is associated with data attributes)
        for index, auto_feature in enumerate(AUTO_FORM_STRUCT[UI_ver]):
            auto_feature_dict[auto_feature] = result[index]
        for dest_feature in DEST_DICT:
            if dest_feature in auto_feature_dict:
                dest_feature_value = DEST_DICT.get(dest_feature, "")
                auto_feature_value = auto_feature_dict.get(dest_feature, "")
                if dest_feature == 'auto_filename' and auto_feature_value.lstrip()[:2] == 'A_':
                    mode_indicator = 'A'
                elif dest_feature == 'auto_filename' and auto_feature_value.lstrip()[:2] == 'B_':
                    mode_indicator = 'B'
                dest_feature_dict[dest_feature_value] =  auto_feature_value
        wsdest[DEST_DICT.get("RowID", "") + str(insertion_point)] = insertion_point - SINK_ROW_INIT + 1 
        wsdest[DEST_DICT.get("TrialKey", "") + str(insertion_point)] = TRIALKEY 
        wsdest[DEST_DICT.get("DataEntryPersonNumber", "") + str(insertion_point)] = 0 
        wsdest[DEST_DICT.get("IsExcluded", "") + str(insertion_point)] = 0 
        wsdest[DEST_DICT.get("DateLoaded", "") + str(insertion_point)] = datetime.now() 
        for column in DEST_DICT_REVERSE:
            if column in dest_feature_dict:
                if column == DEST_DICT.get("Well_Dilution", "") and dest_feature_dict.get(column, "") == '1/40' and mode_indicator == 'A':
                    wsdest[column + str(insertion_point)] = dest_feature_dict.get(column, "")[2:]  
                    wsdest[DEST_DICT.get("Mode", "") + str(insertion_point)] = 'A Prime' 
                elif column == DEST_DICT.get("Well_Dilution", "") and dest_feature_dict.get(column, "") == '1/40' and mode_indicator == 'B':
                    wsdest[column + str(insertion_point)] = dest_feature_dict.get(column, "")[2:]  
                    wsdest[DEST_DICT.get("Mode", "") + str(insertion_point)] = 'B Prime' 
                elif column == DEST_DICT.get("Well_Dilution", "") and mode_indicator == 'A':
                    wsdest[column + str(insertion_point)] = dest_feature_dict.get(column, "")[2:]  
                    wsdest[DEST_DICT.get("Mode", "") + str(insertion_point)] = 'A' 
                elif column == DEST_DICT.get("Well_Dilution", "") and mode_indicator == 'B':
                    wsdest[column + str(insertion_point)] = dest_feature_dict.get(column, "")[2:]  
                    wsdest[DEST_DICT.get("Mode", "") + str(insertion_point)] = 'B' 
                else:
                    wsdest[column + str(insertion_point)] = dest_feature_dict.get(column, "")
        # Get SiteCode based on SerialNumber
        Dev_SN = wsdest[DEST_DICT.get("Device_SerialNum", "") + str(insertion_point)].value
        Dev_SN_spacestrip = str(Dev_SN.replace(" ","")) #in case actual files have space in between numeric and text suffix
        site_code = EQUIPSN_SITE_DICT.get(Dev_SN_spacestrip, "")
        wsdest[DEST_DICT.get("Device_SerialNum", "") + str(insertion_point)] = Dev_SN_spacestrip
        wsdest[DEST_DICT.get("SiteCode", "") + str(insertion_point)] = site_code

        # Snippet of code to change Device Date to European format (to keep consistency for DB import) 
        dev_date = wsdest[DEST_DICT.get("Device_Date", "") + str(insertion_point)].value
        region = REGION_DICT.get(site_code,"") 
        if region != 'EU': ## Convert Device Date to European format to make consistent with SQL database (Toulouse was imported first!)
            wsdest[DEST_DICT.get("Device_Date", "") + str(insertion_point)] = dev_date[3:5] + '-' + dev_date[0:2] + '-' + dev_date[-4:]  #convert MM-DD-YYYY to DD-MM-YYYY

        # Set pattern to false as default
        patt_list = ["Well_Pat_hom", "Well_Pat_sp", "Well_Pat_cen", "Well_Pat_nud", "Well_Pat_nuc", "Well_Pat_cyt"] 
        for patt in patt_list:
            wsdest[DEST_DICT.get(patt, "") + str(insertion_point)] = 'false'
        # change pattern if text present in Well_Pattern field
        if 'homogeneous' in  auto_feature_dict.get('Well_Pattern',"").lower():
            wsdest[DEST_DICT.get("Well_Pat_hom", "") + str(insertion_point)] = 'true' 
        if 'speckled' in  auto_feature_dict.get('Well_Pattern',"").lower():
            wsdest[DEST_DICT.get("Well_Pat_sp", "") + str(insertion_point)] = 'true' 
        if 'centromere' in  auto_feature_dict.get('Well_Pattern',"").lower():
            wsdest[DEST_DICT.get("Well_Pat_cen", "") + str(insertion_point)] = 'true' 
        if 'nuclear dots' in  auto_feature_dict.get('Well_Pattern',"").lower():
            wsdest[DEST_DICT.get("Well_Pat_nud", "") + str(insertion_point)] = 'true' 
        if 'nucleolar' in  auto_feature_dict.get('Well_Pattern',"").lower():
            wsdest[DEST_DICT.get("Well_Pat_nuc", "") + str(insertion_point)] = 'true' 
        if 'cyt' in  auto_feature_dict.get('Well_Pattern',"").lower() or 'ama' in  auto_feature_dict.get('Well_Pattern',"").lower()  :
            wsdest[DEST_DICT.get("Well_Pat_cyt", "") + str(insertion_point)] = 'true' 
        wsdest[DEST_DICT.get("SiteCode", "") + str(insertion_point)] = EQUIPSN_SITE_DICT.get(wsdest[DEST_DICT.get("Device_SerialNum", "") + str(insertion_point)].value, "")
        insertion_point += 1
    return insertion_point

def archive_buffer(buffer_file, dm_arc, type):
    '''Function to archive buffer file in dm_arc location'''
    if type == 'move':
        shutil.move(buffer_file, dm_arc + 'Arc_' + str(format(datetime.now().strftime('%Y%m%d %H%M%S ')))+ '_' + buffer_file.rsplit("/",1)[-1])
    else:
        shutil.copy2(buffer_file, dm_arc + 'Arc_' + str(format(datetime.now().strftime('%Y%m%d %H%M%S ')))+ '_' + buffer_file.rsplit("/",1)[-1])
    return

def main():
    '''Main function'''

    ###################################################   FILE PATHS
    ##Test Files    
    #rawdatapathlist_manual = '/Entry/' # production Filelocation list
    #rawdatapathlist_auto = 'Automatic Export Files/' # production Filelocation list
    #configfile = '//' # production config file
    #destfile = 'sx' # production file
    
    root_dir = 'h Testing Empty/' 
    rawdatapathlist_manual = 'al/' # production Filelocation list
    rawdatapathlist_auto = 'Auto/' # production Filelocation list
    configfile = ''
    destfile = 'xlsx' # production file
    dm_arc = 'C' # archive location
    
    
    ##Production Files below
   
    ####################################################   END OF FILE PATHS

    ###################################################   GET CONFIGURATION DATA
    [TRIALKEY, REGION_DICT, EQUIPSN_SITE_DICT, EQUIPSN_SITE_DICT_REVERSE, SINK_ROW_INIT, SOURCE_ROW_INIT, DEST_DICT, DEST_DICT_REVERSE, BINSEARCH_UB, MAN_FORM_STRUCT, AUTO_FORM_STRUCT, MANUAL_INTERPS_DICT] = input_config_data(configfile) #obtain configuration params
    ###################################################   END OF GET CONFIGURATION DATA

    #### GET SINK (DESTINATION COMBINED REPRO RESULTS) WORKBOOK AS FILE OBJECT 
    wb = load_workbook(filename = destfile)
    wsdest = wb['Data']

    #### GET SOURCE INFO FROM FILE LOCATIONS LIST AND LOAD DATA INTO OBJECT LIST
    manual_files = extract_rawdata(rawdatapathlist_manual)
    auto_files = extract_rawdata(rawdatapathlist_auto)

    #### OBTAIN INSERTION POINT INTO SINK 
    insertion_point = find_insertion_point(wsdest, BINSEARCH_UB, SINK_ROW_INIT)
    if insertion_point == None:
        print('Binary search upper bound exceeded.  Increase upper bound parameter in config file')
        sys.exit()


    #################################################################  BUILD UNIQUE LIST: NO DUPLICATE ENTRIES#######################
    excel_file1= lookin_folder(root_dir)#-----------gets the combine_results xlsx
    dfx = df_combined(excel_file1)#-----------------creates dataframe from combine_results xlsx
    com_hash = build_combined_hash(dfx)#-----------creates list of hashkeys from dataframe
    pdf_list = extract_pdf_list(manual_files, com_hash)#---------creates pdf list from folder
    csv_list  = extract_csv_list(manual_files, com_hash)#-------------creates csv list from folder
    val_list = val_csv_pdf(pdf_list, csv_list)#---------verifies hashkey from com_hash and pdf, csv lists are not in com_hash. 
    new_manual_list = n_rawdata_filelist(manual_files, val_list) #-------------Is the csv and pdf list that ensures no duplicates------
    new_list = v_csv_list(new_manual_list)#------------------------------------Only CSV's ---------------------------------------------
    ###################################################################################################################################
    
    
    #### PROCESS MANUAL LIST 
   # pdf_list = extract_pdf_list(manual_files, com_hash)
    orig_file_args = (new_manual_list, pdf_list)
    manual_orig_file = read_input_file(*orig_file_args)
    print('Processing Manual Form Results...')
    manual_result_args = (manual_orig_file, wsdest, destfile, insertion_point, DEST_DICT, DEST_DICT_REVERSE, MAN_FORM_STRUCT, SINK_ROW_INIT, TRIALKEY, EQUIPSN_SITE_DICT_REVERSE, MANUAL_INTERPS_DICT)
    new_insertion_point = add_manual_results(*manual_result_args)
    print('Saving Manual Form Results...')

    #### PROCESS AUTOEXPORT LIST 
    orig_file_args = (auto_files, SOURCE_ROW_INIT)
    auto_orig_file = read_autoexport_input_file(*orig_file_args)
    print('Processing Automatic Export Results...')
    auto_result_args = (auto_orig_file, wsdest, destfile, new_insertion_point, DEST_DICT, DEST_DICT_REVERSE, AUTO_FORM_STRUCT, SINK_ROW_INIT, TRIALKEY, EQUIPSN_SITE_DICT, REGION_DICT)
    add_auto_results(*auto_result_args)
    print('Saving Automatic Export Results...')

    wb.save(filename = destfile)

    #### ARCHIVE BUFFER FILE _Manual
    for buffer_manual_file in manual_files:
        bufferarchive_manual_args = (buffer_manual_file, dm_arc, 'move')
        archive_buffer(*bufferarchive_manual_args)

    #### ARCHIVE BUFFER FILE _Auto
    for buffer_auto_file in auto_files:
        if '/A_' in str(buffer_auto_file).upper() or '/B_' in str(buffer_auto_file).upper():
            bufferarchive_auto_args = (buffer_auto_file, dm_arc, 'move')
            archive_buffer(*bufferarchive_auto_args)

    #### ARCHIVE COMBINED EXPORT FILE
    bufferarchive_auto_args = (destfile, dm_arc, 'copy')
    archive_buffer(*bufferarchive_auto_args)

if __name__ == '__main__':
    main()


# In[ ]:





# In[ ]:




