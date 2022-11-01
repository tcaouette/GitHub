#!/usr/bin/env python
# coding: utf-8

# In[ ]:





# In[1]:


import pandas as pd
import numpy as np
import re
import sys
import csv
import openpyxl
import decimal
import os
from os import listdir
from os.path import isfile, join
import datetime as dt
import shutil
import uuid
from openpyxl import load_workbook
pd.set_option('mode.chained_assignment', None)


root_dir = r"DM only/"
excel_Peter = r"G"
excel_GenHosp = r"G.xlsx"
excel_BIOrad = r"x"
output_manual = r"Gx"
output_manual_pool =  r"x"



# def archive_excel(dest_dir,root_dir,auto_file,manual_file):#-------archives autofile and manualfile-----instead of deleting the files for history---------
#     exfiles=[]
#     new_files=[]
#     add_date = dt.datetime.now().strftime("%Y%m%d %H%M%S%f")
#     for f in listdir(root_dir):
#         if auto_file in  f:
#             exfiles.append(f)
#             exfiles = [i for i in exfiles if not i.startswith('~$')] # ----------- removes ~$ issue from list
#             new_files = [root_dir + i for i in exfiles]
#         if manual_file in f:
#             exfiles.append(f)
#             exfiles = [i for i in exfiles if not i.startswith('~$')] # ----------- removes ~$ issue from list
#             new_files = [root_dir + i for i in exfiles]
            
#     if len(new_files) > 0:
    
#         for files in new_files:
#             shutil.move(files, dest_dir)
    
#         for file in listdir(dest_dir):
#             if not file.startswith("Archived"):
#                 dst = dest_dir + 'Archived_'+ add_date +'_' +  file 
#                 src =  dest_dir + file
        
#                 os.rename(src, dst)
#     else: 
#         print('No files to be moved to Archive.')
        
#archive_excel(root_dir)

def lookin_folder(root_dir, r_exfile=None):
    '''Looks in a folder via a path and builds a list with files that have extension xlsm'''
    ex_files =[]
    for file in listdir(root_dir):
        if 'Repro'in file and file.endswith('.xlsx'):
            ex_files.append(file)
            ex_files = [i for i in ex_files if not i.startswith('~$')] # ----------- removes ~$ issue from list
            r_exfile = [root_dir + i for i in ex_files]
    return r_exfile


files = lookin_folder(root_dir)



excel_file1,excel_file2,excel_file3 = [files[i] for i in range(len(files))]



# theFile = load_workbook(filename = excel_file1, data_only = True)
# allSheetNames = theFile.sheetnames



######################build more than 8##########################
def build_9_dfs(file_name):
    theFile = load_workbook(filename = file_name, data_only = True)
    allSheetNames = theFile.sheetnames



    data_notpoolnacl_plus =[]
    data_notpoolnacl_proph = []
    data_LISScoombs_proph = []
    data_LISScoombs_plus =[]
    data_COOMBS_igg_proph = []
    data_COOMBS_igg_Plus =[]

    data_LISScoombs_pool = []
    data_COOMBS_igg_pool = []
    data_poolnacl =[]

    ################################### NACL Plus #############################
    for sheet in allSheetNames:
        if 'Pool' in sheet and 'NaCl'in sheet :
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, currentSheet.max_column + 1 ):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_poolnacl.append(data_col)
            dfpool_nacl = pd.DataFrame(data_poolnacl)
    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'NaCl' in sheet and 'Plus' in sheet:
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, 34):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_notpoolnacl_plus.append(data_col)
            df_nacl_plus = pd.DataFrame(data_notpoolnacl_plus)
    ##########################NACL proph####################################
    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'NaCl' in sheet and 'Prophylax' in sheet:
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, 34):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_notpoolnacl_proph.append(data_col)
            df_nacl_proph = pd.DataFrame(data_notpoolnacl_proph)
    #################################LISS Coombs Plus #######################
    # for sheet in allSheetNames:
        if 'Pool' in sheet and 'LISS'in sheet:
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, currentSheet.max_column + 1 ):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_LISScoombs_pool.append(data_col)
            df_Liss_pool = pd.DataFrame(data_LISScoombs_pool)
    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'LISS' in sheet and 'Plus' in sheet:
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, 34):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_LISScoombs_plus.append(data_col)
            df_Liss_plus = pd.DataFrame(data_LISScoombs_plus)
     #####################################LISS Coombs proph#######################
    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'LISS' in sheet and 'Prophyl' in sheet:
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, 34):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_LISScoombs_proph.append(data_col)
            df_Liss_proph = pd.DataFrame(data_LISScoombs_proph)
    #################################Coombs Igg plus ######################################
    # for sheet in allSheetNames:
        if 'Pool' in sheet and 'IgG'in sheet :
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, currentSheet.max_column + 1 ):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_COOMBS_igg_pool.append(data_col)
            df_igg_pool = pd.DataFrame(data_COOMBS_igg_pool)
    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'IgG' in sheet and 'Prophyl' in sheet:
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, 34):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_COOMBS_igg_proph.append(data_col)
            df_igg_proph = pd.DataFrame(data_COOMBS_igg_proph)
    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'IgG' in sheet and 'Plus' in sheet:
            currentSheet = theFile[sheet]
            for row in range(22, 50):
                data_col=[]
                for column in range(3, 34):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_COOMBS_igg_Plus.append(data_col)
            df_igg_plus = pd.DataFrame(data_COOMBS_igg_Plus)
            
    return dfpool_nacl, df_nacl_plus, df_nacl_proph, df_Liss_pool, df_Liss_plus, df_Liss_proph, df_igg_pool, df_igg_proph, df_igg_plus

dfpeter_poolnacl, dfpeter_naclplus, dfpeter_naclproph, dfpeter_lisspool, dfpeter_lissplus, dfpeter_lissproph, dfpeter_iggpool, dfpeter_iggproph, dfpeter_iggplus = build_9_dfs(excel_file1)

dfgenhos_poolnacl, dfgenhos_naclplus, dfgenhos_naclproph, dfgenhos_lisspool, dfgenhos_lissplus, dfgenhos_lissproph, dfgenhos_iggpool, dfgenhos_iggproph, dfgenhos_iggplus = build_9_dfs(excel_file2)

dfbiorad_poolnacl, dfbiorad_naclplus, dfbiorad_naclproph, dfbiorad_lisspool, dfbiorad_lissplus, dfbiorad_lissproph, dfbiorad_iggpool, dfbiorad_iggproph, dfbiorad_iggplus = build_9_dfs(excel_file3)







def get_lotNumber_9(file_path):   
    theFile = load_workbook(filename = file_path, data_only = True)
    allSheetNames = theFile.sheetnames
   
    data_notpoolnacl_plus =[]
    data_notpoolnacl_proph = []
    data_LISScoombs_proph = []
    data_LISScoombs_plus =[]
    data_COOMBS_igg_proph = []
    data_COOMBS_igg_Plus =[]

    data_LISScoombs_pool = []
    data_COOMBS_igg_pool = []
    data_poolnacl =[]
 
    for sheet in allSheetNames:
        if 'Pool' in sheet and 'NaCl'in sheet :
            currentSheet = theFile[sheet]
            for row in range(15, 18):
                data_col=[]
                for column in range(6,7 ):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_poolnacl.append(data_col)

 
        if 'Pool' not in sheet and 'NaCl' in sheet and 'Plus' in sheet:
            currentSheet = theFile[sheet]
            for row in range(15, 16):
                data_col=[]
                for column in range(6,7):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_notpoolnacl_plus.append(data_col)

    ##########################NACL proph####################################
   
        if 'Pool' not in sheet and 'NaCl' in sheet and 'Prophylax' in sheet:
            currentSheet = theFile[sheet]
            for row in range(15, 16):
                data_col=[]
                for column in range(6,7):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_notpoolnacl_proph.append(data_col)

    #################################LISS Coombs Plus #######################
    # for sheet in allSheetNames:
        if 'Pool' in sheet and 'LISS'in sheet:
            currentSheet = theFile[sheet]
            for row in range(15, 18):
                data_col=[]
                for column in range(6,7 ):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_LISScoombs_pool.append(data_col)

    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'LISS' in sheet and 'Plus' in sheet:
            currentSheet = theFile[sheet]
            for row in range(15, 16):
                data_col=[]
                for column in range(6,7):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_LISScoombs_plus.append(data_col)

     #####################################LISS Coombs proph#######################
    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'LISS' in sheet and 'Prophyl' in sheet:
            currentSheet = theFile[sheet]
            for row in range(15, 16):
                data_col=[]
                for column in range(6,7):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_LISScoombs_proph.append(data_col)

    #################################Coombs Igg plus ######################################
    # for sheet in allSheetNames:
        if 'Pool' in sheet and 'IgG'in sheet :
            currentSheet = theFile[sheet]
            for row in range(15, 18):
                data_col=[]
                for column in range(6,7 ):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_COOMBS_igg_pool.append(data_col)

    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'IgG' in sheet and 'Prophyl' in sheet:
            currentSheet = theFile[sheet]
            for row in range(15, 16):
                data_col=[]
                for column in range(6,7):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_COOMBS_igg_proph.append(data_col)

    # for sheet in allSheetNames:
        if 'Pool' not in sheet and 'IgG' in sheet and 'Plus' in sheet:
            currentSheet = theFile[sheet]
            for row in range(15, 16):
                data_col=[]
                for column in range(6,7):
                    cell_value = currentSheet.cell(column = column, row = row).value
                    data_col.append(cell_value)
                data_COOMBS_igg_Plus.append(data_col)
        
            
    return data_poolnacl, data_notpoolnacl_plus, data_notpoolnacl_proph,data_LISScoombs_pool,data_LISScoombs_plus, data_LISScoombs_proph,data_COOMBS_igg_pool,data_COOMBS_igg_proph, data_COOMBS_igg_Plus

lotpeter_poolnacl, lotpeter_naclplus, lotpeter_naclproph, lotpeter_lisspool, lotpeter_lissplus, lotpeter_lissproph, lotpeter_iggpool, lotpeter_iggproph, lotpeter_iggplus = get_lotNumber_9(excel_file1)

lotgenhos_poolnacl, lotgenhos_naclplus, lotgenhos_naclproph, lotgenhos_lisspool, lotgenhos_lissplus, lotgenhos_lissproph, lotgenhos_iggpool, lotgenhos_iggproph, lotgenhos_iggplus = get_lotNumber_9(excel_file2)

lotbiorad_poolnacl, lotbiorad_naclplus, lotbiorad_naclproph, lotbiorad_lisspool, lotbiorad_lissplus, lotbiorad_lissproph, lotbiorad_iggpool, lotbiorad_iggproph, lotbiorad_iggplus = get_lotNumber_9(excel_file3)













def build_nonpool(df, lotnumber):
    df.fillna(value=pd.np.nan, inplace =True)
    df.replace('0,5','0.5', inplace = True)
    df.replace('0.5',0.5, inplace = True)
    df.replace('Run 2', np.nan, inplace = True)
    df.replace('*',np.nan, inplace = True)
    df.replace('Run 1', np.nan, inplace = True)
    df.replace('Identifier',np.nan, inplace =True)
    df.replace('Sample',np.nan, inplace =True)
    df = df.dropna(axis =1, how ='all')
    df = df.dropna()
    df = df.reset_index(drop = True)
    df['index'] = df.index.tolist()
    
    df.loc[df.index <= 9 , 'AM-PM'] = 'AM'
    df.loc[(df.index > 9) & (df.index <= 19) , 'AM-PM'] = 'PM'
    df.loc[(df.index > 19) & (df.index <= 29) , 'AM-PM'] = 'AM'
    df.loc[(df.index > 29) & (df.index <= 39) , 'AM-PM'] = 'PM'
    df.loc[(df.index > 39) & (df.index <= 49) , 'AM-PM'] = 'AM'
    df.loc[df.index > 49 , 'AM-PM'] = 'PM'

    
    df.loc[df.index <= 19 , 'LOT'] = lotnumber[0][0]
    df.loc[(df.index > 19) & (df.index <= 39) , 'LOT'] = lotnumber[1][0]
    df.loc[df.index > 39 , 'LOT'] = lotnumber[2][0]

    df['LOT'] =  df['LOT'].astype(str)
    df['LOT']=  df['LOT'].str.rstrip('.0')
    df.sort_values(by = 2, inplace = True)

    df = df.reset_index(drop = True)

    nopoolnewheaders ={0:'Identifier', 2:'Sample', 3:'1',4:'2',5:'3',6:'4',7:'5',8:'6',11:'Identifier',13:'Sample',14:'1',15:'2',16:'3',17:'4',18:'5',19:'6',22:'Identifier',24:'Sample',25:'1',26:'2',27:'3',28:'4',
                29:'5',30:'6'}

    df.rename(columns = nopoolnewheaders, inplace = True)



    day1 = df.iloc[:,[0,1,2,3,4,5,6,7,24,25,26]]
    day2 = df.iloc[:, [8,9,10,11,12,13,14,15,24,25,26]]
    day3 = df.iloc[:,[16,17,18,19,20,21,22,23,24,25,26]]    



    day3['Day'] = 3
    day2['Day'] = 2
    day1['Day'] = 1

    day1_day2 = pd.concat([day1,day2], sort = False)

    day1_day2_day3 = pd.concat([day1_day2, day3], sort = False)

    day1_day2_day3.sort_values(by =['Sample'], inplace = True)
    day1_day2_day3= day1_day2_day3.reset_index(drop = True)
    return day1_day2_day3





############################## Use this for finding unique identifier #######################

def make_list_man(lot):
    gg = []
    for i in lot:
        for j in i:
            gg.append(j)
    return gg
        
def unique(list1): 
  
    # intilize a null list 
    unique_list = [] 
      
    # traverse for all elements 
    for x in list1: 
        # check if exists in unique_list or not 
        if x not in unique_list: 
            unique_list.append(x) 
    # print list 
    return unique_list








# dfpeter = build_nonpool(dfpeter, lotpeter)
# dfgenhosp = build_nonpool(dfgenhosp,lotgenhosp)
# dfbiorad = build_nonpool(dfbiorad,lotbiorad)

dfpeter_iggplus = build_nonpool(dfpeter_iggplus,lotpeter_iggplus)
dfpeter_iggproph = build_nonpool(dfpeter_iggproph, lotpeter_iggproph)

dfpeter_lissproph = build_nonpool(dfpeter_lissproph, lotpeter_lissproph)
dfpeter_lissplus = build_nonpool(dfpeter_lissplus, lotpeter_lissplus)

dfpeter_naclproph = build_nonpool(dfpeter_naclproph, lotpeter_naclproph)
dfpeter_naclplus = build_nonpool(dfpeter_naclplus, lotpeter_naclplus)

dfgenhos_iggplus = build_nonpool(dfgenhos_iggplus,lotgenhos_iggplus)
dfgenhos_iggproph = build_nonpool(dfgenhos_iggproph, lotgenhos_iggproph)

dfgenhos_lissproph = build_nonpool(dfgenhos_lissproph, lotgenhos_lissproph)
dfgenhos_lissplus = build_nonpool(dfgenhos_lissplus, lotgenhos_lissplus)

dfgenhos_naclproph = build_nonpool(dfgenhos_naclproph, lotgenhos_naclproph)
dfgenhos_naclplus = build_nonpool(dfgenhos_naclplus, lotgenhos_naclplus)

dfbiorad_iggplus = build_nonpool(dfbiorad_iggplus,lotbiorad_iggplus)
dfbiorad_iggproph = build_nonpool(dfbiorad_iggproph, lotbiorad_iggproph)

dfbiorad_lissproph = build_nonpool(dfbiorad_lissproph, lotbiorad_lissproph)
dfbiorad_lissplus = build_nonpool(dfbiorad_lissplus, lotbiorad_lissplus)

dfbiorad_naclproph = build_nonpool(dfbiorad_naclproph, lotbiorad_naclproph)
dfbiorad_naclplus = build_nonpool(dfbiorad_naclplus, lotbiorad_naclplus)



def peter_df(df):
#     df.drop('index', axis = 1, inplace = True)
    df= df.reset_index(drop = True)
    df['Site'] = 'Peterborough'
    df = df[['Site','Day', 'AM-PM', 'LOT', 'Identifier','Sample', '1','2','3','4','5','6']]
    df = df.sort_values(by=['Site', 'Day','AM-PM'])
    return df
dfpeter_iggplus = peter_df(dfpeter_iggplus)
dfpeter_iggproph = peter_df(dfpeter_iggproph)
dfpeter_lissproph = peter_df(dfpeter_lissproph)
dfpeter_lissplus = peter_df(dfpeter_lissplus)
dfpeter_naclproph = peter_df(dfpeter_naclproph)
dfpeter_naclproph = peter_df(dfpeter_naclproph)
dfpeter_naclplus = peter_df(dfpeter_naclplus)



def genhosp_df(df):
#     df.drop('index', axis = 1, inplace = True)
    df= df.reset_index(drop = True)
    df['Site'] = 'Site Alexandra Marine & Gen Hosp'
    df = df[['Site','Day', 'AM-PM', 'LOT', 'Identifier','Sample', '1','2','3','4','5','6']]
    df = df.sort_values(by=['Site', 'Day','AM-PM'])
    return df
dfgenhos_iggplus = genhosp_df(dfgenhos_iggplus)
dfgenhos_iggproph = genhosp_df(dfgenhos_iggproph)
dfgenhos_lissproph = genhosp_df(dfgenhos_lissproph)
dfgenhos_lissplus = genhosp_df(dfgenhos_lissplus)
dfgenhos_naclproph = genhosp_df(dfgenhos_naclproph)
dfgenhos_naclproph = genhosp_df(dfgenhos_naclproph)
dfgenhos_naclplus = genhosp_df(dfgenhos_naclplus)



def biorad_df(df):
#     df.drop('index', axis = 1, inplace = True)
    df= df.reset_index(drop = True)
    df['Site'] = 'Site Bio-Rad'
    df = df[['Site','Day', 'AM-PM', 'LOT', 'Identifier','Sample', '1','2','3','4','5','6']]
    df = df.sort_values(by=['Site', 'Day','AM-PM'])
    return df
# dfbiorad = biorad_df(dfbiorad)
dfbiorad_iggplus = biorad_df(dfbiorad_iggplus)
dfbiorad_iggproph = biorad_df(dfbiorad_iggproph)
dfbiorad_lissproph = biorad_df(dfbiorad_lissproph)
dfbiorad_lissplus = biorad_df(dfbiorad_lissplus)
dfbiorad_naclproph = biorad_df(dfbiorad_naclproph)
dfbiorad_naclproph = biorad_df(dfbiorad_naclproph)
dfbiorad_naclplus = biorad_df(dfbiorad_naclplus)

###########simplify with pandas iterrows#######






def make_plus(df):
    #########################################sample 1
    df_Sam1_lot1 = df[(df.Sample == 1) & (df.LOT == '45670061') ] 
    df_Sam1_lot1 = df_Sam1_lot1.reset_index(drop = True)
    df_Sam1_lot1['index'] =  df_Sam1_lot1.index.tolist()
    df_Sam1_lot1 = df_Sam1_lot1.drop(['LOT'], axis = 1)
    
    df_Sam1_lot2 = df[(df.Sample == 1) & (df.LOT == '45670071') ]
    df_Sam1_lot2 = df_Sam1_lot2.reset_index(drop = True)
    df_Sam1_lot2['index'] = df_Sam1_lot2.index.tolist()
    df_Sam1_lot2 = df_Sam1_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam1_lot3 = df[(df.Sample == 1) & (df.LOT == '45670081') ]
    df_Sam1_lot3 = df_Sam1_lot3.reset_index(drop = True)
    df_Sam1_lot3['index'] = df_Sam1_lot3.index.tolist()
    df_Sam1_lot3= df_Sam1_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam1_lot1,df_Sam1_lot2], axis = 1, sort = False)
    df_Sam1 = pd.concat([df1_df2,  df_Sam1_lot3], axis = 1, sort = False)
    df_Sam1 = df_Sam1.drop(['index'], axis = 1)
    df_Sam1.insert(11, 'blank','')
    df_Sam1.insert(18, 'blank2','')
    #########################################sample 2
    df_Sam2_lot1 = df[(df.Sample == 2) & (df.LOT == '45670061') ] 
    df_Sam2_lot1 = df_Sam2_lot1.reset_index(drop = True)
    df_Sam2_lot1['index'] =  df_Sam2_lot1.index.tolist()
    df_Sam2_lot1 = df_Sam2_lot1.drop(['LOT'], axis = 1)

    df_Sam2_lot2 = df[(df.Sample == 2) & (df.LOT == '45670071') ]
    df_Sam2_lot2 = df_Sam2_lot2.reset_index(drop = True)
    df_Sam2_lot2['index'] = df_Sam2_lot2.index.tolist()
    df_Sam2_lot2 = df_Sam2_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam2_lot3 = df[(df.Sample == 2) & (df.LOT == '45670081') ]
    df_Sam2_lot3 = df_Sam2_lot3.reset_index(drop = True)
    df_Sam2_lot3['index'] = df_Sam2_lot3.index.tolist()
    df_Sam2_lot3= df_Sam2_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam2_lot1,df_Sam2_lot2], axis = 1, sort = False)
    df_Sam2 = pd.concat([df1_df2,  df_Sam2_lot3], axis = 1, sort = False)
    df_Sam2 = df_Sam2.drop(['index'], axis = 1)
    df_Sam2.insert(11, 'blank','')
    df_Sam2.insert(18, 'blank2','')
    ########################################sample 3
    df_Sam3_lot1 = df[(df.Sample == 3) & (df.LOT == '45670061') ] 
    df_Sam3_lot1 = df_Sam3_lot1.reset_index(drop = True)
    df_Sam3_lot1['index'] =  df_Sam3_lot1.index.tolist()
    df_Sam3_lot1 = df_Sam3_lot1.drop(['LOT'], axis = 1)

    df_Sam3_lot2 = df[(df.Sample == 3) & (df.LOT == '45670071') ]
    df_Sam3_lot2 = df_Sam3_lot2.reset_index(drop = True)
    df_Sam3_lot2['index'] = df_Sam3_lot2.index.tolist()
    df_Sam3_lot2 = df_Sam3_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam3_lot3 = df[(df.Sample == 3) & (df.LOT == '45670081') ]
    df_Sam3_lot3 = df_Sam3_lot3.reset_index(drop = True)
    df_Sam3_lot3['index'] = df_Sam3_lot3.index.tolist()
    df_Sam3_lot3= df_Sam3_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam3_lot1,df_Sam3_lot2], axis = 1, sort = False)
    df_Sam3 = pd.concat([df1_df2,  df_Sam3_lot3], axis = 1, sort = False)
    df_Sam3 = df_Sam3.drop(['index'], axis = 1)
    
    df_Sam3.insert(11, 'blank','')
    df_Sam3.insert(18, 'blank2','')
    #################################sample 4
    df_Sam4_lot1 = df[(df.Sample == 4) & (df.LOT == '45670061') ] 
    df_Sam4_lot1 = df_Sam4_lot1.reset_index(drop = True)
    df_Sam4_lot1['index'] =  df_Sam4_lot1.index.tolist()
    df_Sam4_lot1 = df_Sam4_lot1.drop(['LOT'], axis = 1)

    df_Sam4_lot2 = df[(df.Sample == 4) & (df.LOT == '45670071') ]
    df_Sam4_lot2 = df_Sam4_lot2.reset_index(drop = True)
    df_Sam4_lot2['index'] = df_Sam4_lot2.index.tolist()
    df_Sam4_lot2 = df_Sam4_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam4_lot3 = df[(df.Sample == 4) & (df.LOT == '45670081') ]
    df_Sam4_lot3 = df_Sam4_lot3.reset_index(drop = True)
    df_Sam4_lot3['index'] = df_Sam4_lot3.index.tolist()
    df_Sam4_lot3= df_Sam4_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam4_lot1,df_Sam4_lot2], axis = 1, sort = False)
    df_Sam4 = pd.concat([df1_df2,  df_Sam4_lot3], axis = 1, sort = False)
    df_Sam4 = df_Sam4.drop(['index'], axis = 1)
    
    df_Sam4.insert(11, 'blank','')
    df_Sam4.insert(18, 'blank2','')
    #######################################sample 5
    df_Sam5_lot1 = df[(df.Sample == 5) & (df.LOT == '45670061') ] 
    df_Sam5_lot1 = df_Sam5_lot1.reset_index(drop = True)
    df_Sam5_lot1['index'] =  df_Sam5_lot1.index.tolist()
    df_Sam5_lot1 = df_Sam5_lot1.drop(['LOT'], axis = 1)

    df_Sam5_lot2 = df[(df.Sample == 5) & (df.LOT == '45670071') ]
    df_Sam5_lot2 = df_Sam5_lot2.reset_index(drop = True)
    df_Sam5_lot2['index'] = df_Sam5_lot2.index.tolist()
    df_Sam5_lot2 = df_Sam5_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam5_lot3 = df[(df.Sample == 5) & (df.LOT == '45670081') ]
    df_Sam5_lot3 = df_Sam5_lot3.reset_index(drop = True)
    df_Sam5_lot3['index'] = df_Sam5_lot3.index.tolist()
    df_Sam5_lot3= df_Sam5_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam5_lot1,df_Sam5_lot2], axis = 1, sort = False)
    df_Sam5 = pd.concat([df1_df2,  df_Sam5_lot3], axis = 1, sort = False)
    df_Sam5 = df_Sam5.drop(['index'], axis = 1)
    
    df_Sam5.insert(11, 'blank','')
    df_Sam5.insert(18, 'blank2','')
    #################################################sample 6
    df_Sam6_lot1 = df[(df.Sample == 6) & (df.LOT == '45670061') ] 
    df_Sam6_lot1 = df_Sam6_lot1.reset_index(drop = True)
    df_Sam6_lot1['index'] =  df_Sam6_lot1.index.tolist()
    df_Sam6_lot1 = df_Sam6_lot1.drop(['LOT'], axis = 1)

    df_Sam6_lot2 = df[(df.Sample == 6) & (df.LOT == '45670071') ]
    df_Sam6_lot2 = df_Sam6_lot2.reset_index(drop = True)
    df_Sam6_lot2['index'] = df_Sam6_lot2.index.tolist()
    df_Sam6_lot2 = df_Sam6_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam6_lot3 = df[(df.Sample == 6) & (df.LOT == '45670081') ]
    df_Sam6_lot3 = df_Sam6_lot3.reset_index(drop = True)
    df_Sam6_lot3['index'] = df_Sam6_lot3.index.tolist()
    df_Sam6_lot3= df_Sam6_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam6_lot1,df_Sam6_lot2], axis = 1, sort = False)
    df_Sam6 = pd.concat([df1_df2,  df_Sam6_lot3], axis = 1, sort = False)
    df_Sam6 = df_Sam6.drop(['index'], axis = 1)
    
    df_Sam6.insert(11, 'blank','')
    df_Sam6.insert(18, 'blank2','')
    #############################################sample 7
    df_Sam7_lot1 = df[(df.Sample == 7) & (df.LOT == '45670061') ] 
    df_Sam7_lot1 = df_Sam7_lot1.reset_index(drop = True)
    df_Sam7_lot1['index'] =  df_Sam7_lot1.index.tolist()
    df_Sam7_lot1 = df_Sam7_lot1.drop(['LOT'], axis = 1)

    df_Sam7_lot2 = df[(df.Sample == 7) & (df.LOT == '45670071') ]
    df_Sam7_lot2 = df_Sam7_lot2.reset_index(drop = True)
    df_Sam7_lot2['index'] = df_Sam7_lot2.index.tolist()
    df_Sam7_lot2 = df_Sam7_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam7_lot3 = df[(df.Sample == 7) & (df.LOT == '45670081') ]
    df_Sam7_lot3 = df_Sam7_lot3.reset_index(drop = True)
    df_Sam7_lot3['index'] = df_Sam7_lot3.index.tolist()
    df_Sam7_lot3= df_Sam7_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam7_lot1,df_Sam7_lot2], axis = 1, sort = False)
    df_Sam7 = pd.concat([df1_df2,  df_Sam7_lot3], axis = 1, sort = False)
    df_Sam7 = df_Sam7.drop(['index'], axis = 1)
    
    df_Sam7.insert(11, 'blank','')
    df_Sam7.insert(18, 'blank2','')
    ################################################sample 8
    df_Sam8_lot1 = df[(df.Sample == 8) & (df.LOT == '45670061') ] 
    df_Sam8_lot1 = df_Sam8_lot1.reset_index(drop = True)
    df_Sam8_lot1['index'] =  df_Sam8_lot1.index.tolist()
    df_Sam8_lot1 = df_Sam8_lot1.drop(['LOT'], axis = 1)

    df_Sam8_lot2 = df[(df.Sample == 8) & (df.LOT == '45670071') ]
    df_Sam8_lot2 = df_Sam8_lot2.reset_index(drop = True)
    df_Sam8_lot2['index'] = df_Sam8_lot2.index.tolist()
    df_Sam8_lot2 = df_Sam8_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam8_lot3 = df[(df.Sample == 8) & (df.LOT == '45670081') ]
    df_Sam8_lot3 = df_Sam8_lot3.reset_index(drop = True)
    df_Sam8_lot3['index'] = df_Sam8_lot3.index.tolist()
    df_Sam8_lot3= df_Sam8_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam8_lot1,df_Sam8_lot2], axis = 1, sort = False)
    df_Sam8 = pd.concat([df1_df2,  df_Sam8_lot3], axis = 1, sort = False)
    df_Sam8 = df_Sam8.drop(['index'], axis = 1)
    
    df_Sam8.insert(11, 'blank','')
    df_Sam8.insert(18, 'blank2','')
    #############################################sample 9
    df_Sam9_lot1 = df[(df.Sample == 9) & (df.LOT == '45670061') ] 
    df_Sam9_lot1 = df_Sam9_lot1.reset_index(drop = True)
    df_Sam9_lot1['index'] =  df_Sam9_lot1.index.tolist()
    df_Sam9_lot1 = df_Sam9_lot1.drop(['LOT'], axis = 1)

    df_Sam9_lot2 = df[(df.Sample == 9) & (df.LOT == '45670071') ]
    df_Sam9_lot2 = df_Sam9_lot2.reset_index(drop = True)
    df_Sam9_lot2['index'] = df_Sam9_lot2.index.tolist()
    df_Sam9_lot2 = df_Sam9_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam9_lot3 = df[(df.Sample == 9) & (df.LOT == '45670081') ]
    df_Sam9_lot3 = df_Sam9_lot3.reset_index(drop = True)
    df_Sam9_lot3['index'] = df_Sam9_lot3.index.tolist()
    df_Sam9_lot3= df_Sam9_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam9_lot1,df_Sam9_lot2], axis = 1, sort = False)
    df_Sam9 = pd.concat([df1_df2,  df_Sam9_lot3], axis = 1, sort = False)
    df_Sam9 = df_Sam9.drop(['index'], axis = 1)
    
    df_Sam9.insert(11, 'blank','')
    df_Sam9.insert(18, 'blank2','')
    ##############################################sample 10
    df_Sam10_lot1 = df[(df.Sample == 10) & (df.LOT == '45670061') ] 
    df_Sam10_lot1 = df_Sam10_lot1.reset_index(drop = True)
    df_Sam10_lot1['index'] =  df_Sam10_lot1.index.tolist()
    df_Sam10_lot1 = df_Sam10_lot1.drop(['LOT'], axis = 1)

    df_Sam10_lot2 = df[(df.Sample == 10) & (df.LOT == '45670071') ]
    df_Sam10_lot2 = df_Sam10_lot2.reset_index(drop = True)
    df_Sam10_lot2['index'] = df_Sam10_lot2.index.tolist()
    df_Sam10_lot2 = df_Sam10_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam10_lot3 = df[(df.Sample == 10) & (df.LOT == '45670081') ]
    df_Sam10_lot3 = df_Sam10_lot3.reset_index(drop = True)
    df_Sam10_lot3['index'] = df_Sam10_lot3.index.tolist()
    df_Sam10_lot3= df_Sam10_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam10_lot1,df_Sam10_lot2], axis = 1, sort = False)
    df_Sam10 = pd.concat([df1_df2,  df_Sam10_lot3], axis = 1, sort = False)
    df_Sam10 = df_Sam10.drop(['index'], axis = 1)
    
    df_Sam10.insert(11, 'blank','')
    df_Sam10.insert(18, 'blank2','')
    
    return df_Sam1,df_Sam2,df_Sam3,df_Sam4,df_Sam5,df_Sam6,df_Sam7,df_Sam8,df_Sam9,df_Sam10




def make_proph(df):
    #########################################sample 1
    df_Sam1_lot1 = df[(df.Sample == 1) & (df.LOT == '45660082') ] 
    df_Sam1_lot1 = df_Sam1_lot1.reset_index(drop = True)
    df_Sam1_lot1['index'] =  df_Sam1_lot1.index.tolist()
    df_Sam1_lot1 = df_Sam1_lot1.drop(['LOT'], axis = 1)

    df_Sam1_lot2 = df[(df.Sample == 1) & (df.LOT == '45660091') ]
    df_Sam1_lot2 = df_Sam1_lot2.reset_index(drop = True)
    df_Sam1_lot2['index'] = df_Sam1_lot2.index.tolist()
    df_Sam1_lot2 = df_Sam1_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam1_lot3 = df[(df.Sample == 1) & (df.LOT == '45660101') ]
    df_Sam1_lot3 = df_Sam1_lot3.reset_index(drop = True)
    df_Sam1_lot3['index'] = df_Sam1_lot3.index.tolist()
    df_Sam1_lot3= df_Sam1_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam1_lot1,df_Sam1_lot2], axis = 1, sort = False)
    df_Sam1 = pd.concat([df1_df2,  df_Sam1_lot3], axis = 1, sort = False)
    df_Sam1 = df_Sam1.drop(['index'], axis = 1)
    df_Sam1.insert(11, 'blank','')
    df_Sam1.insert(18, 'blank2','')
    #########################################sample 2
    df_Sam2_lot1 = df[(df.Sample == 2) & (df.LOT == '45660082') ] 
    df_Sam2_lot1 = df_Sam2_lot1.reset_index(drop = True)
    df_Sam2_lot1['index'] =  df_Sam2_lot1.index.tolist()
    df_Sam2_lot1 = df_Sam2_lot1.drop(['LOT'], axis = 1)

    df_Sam2_lot2 = df[(df.Sample == 2) & (df.LOT == '45660091') ]
    df_Sam2_lot2 = df_Sam2_lot2.reset_index(drop = True)
    df_Sam2_lot2['index'] = df_Sam2_lot2.index.tolist()
    df_Sam2_lot2 = df_Sam2_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam2_lot3 = df[(df.Sample == 2) & (df.LOT == '45660101') ]
    df_Sam2_lot3 = df_Sam2_lot3.reset_index(drop = True)
    df_Sam2_lot3['index'] = df_Sam2_lot3.index.tolist()
    df_Sam2_lot3= df_Sam2_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam2_lot1,df_Sam2_lot2], axis = 1, sort = False)
    df_Sam2 = pd.concat([df1_df2,  df_Sam2_lot3], axis = 1, sort = False)
    df_Sam2 = df_Sam2.drop(['index'], axis = 1)
    df_Sam2.insert(11, 'blank','')
    df_Sam2.insert(18, 'blank2','')
    ########################################sample 3
    df_Sam3_lot1 = df[(df.Sample == 3) & (df.LOT == '45660082') ] 
    df_Sam3_lot1 = df_Sam3_lot1.reset_index(drop = True)
    df_Sam3_lot1['index'] =  df_Sam3_lot1.index.tolist()
    df_Sam3_lot1 = df_Sam3_lot1.drop(['LOT'], axis = 1)

    df_Sam3_lot2 = df[(df.Sample == 3) & (df.LOT == '45660091') ]
    df_Sam3_lot2 = df_Sam3_lot2.reset_index(drop = True)
    df_Sam3_lot2['index'] = df_Sam3_lot2.index.tolist()
    df_Sam3_lot2 = df_Sam3_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam3_lot3 = df[(df.Sample == 3) & (df.LOT =='45660101') ]
    df_Sam3_lot3 = df_Sam3_lot3.reset_index(drop = True)
    df_Sam3_lot3['index'] = df_Sam3_lot3.index.tolist()
    df_Sam3_lot3= df_Sam3_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam3_lot1,df_Sam3_lot2], axis = 1, sort = False)
    df_Sam3 = pd.concat([df1_df2,  df_Sam3_lot3], axis = 1, sort = False)
    df_Sam3 = df_Sam3.drop(['index'], axis = 1)
    
    df_Sam3.insert(11, 'blank','')
    df_Sam3.insert(18, 'blank2','')
    #################################sample 4
    df_Sam4_lot1 = df[(df.Sample == 4) & (df.LOT == '45660082') ] 
    df_Sam4_lot1 = df_Sam4_lot1.reset_index(drop = True)
    df_Sam4_lot1['index'] =  df_Sam4_lot1.index.tolist()
    df_Sam4_lot1 = df_Sam4_lot1.drop(['LOT'], axis = 1)

    df_Sam4_lot2 = df[(df.Sample == 4) & (df.LOT == '45660091') ]
    df_Sam4_lot2 = df_Sam4_lot2.reset_index(drop = True)
    df_Sam4_lot2['index'] = df_Sam4_lot2.index.tolist()
    df_Sam4_lot2 = df_Sam4_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam4_lot3 = df[(df.Sample == 4) & (df.LOT == '45660101') ]
    df_Sam4_lot3 = df_Sam4_lot3.reset_index(drop = True)
    df_Sam4_lot3['index'] = df_Sam4_lot3.index.tolist()
    df_Sam4_lot3= df_Sam4_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam4_lot1,df_Sam4_lot2], axis = 1, sort = False)
    df_Sam4 = pd.concat([df1_df2,  df_Sam4_lot3], axis = 1, sort = False)
    df_Sam4 = df_Sam4.drop(['index'], axis = 1)
    
    df_Sam4.insert(11, 'blank','')
    df_Sam4.insert(18, 'blank2','')
    #######################################sample 5
    df_Sam5_lot1 = df[(df.Sample == 5) & (df.LOT == '45660082') ] 
    df_Sam5_lot1 = df_Sam5_lot1.reset_index(drop = True)
    df_Sam5_lot1['index'] =  df_Sam5_lot1.index.tolist()
    df_Sam5_lot1 = df_Sam5_lot1.drop(['LOT'], axis = 1)

    df_Sam5_lot2 = df[(df.Sample == 5) & (df.LOT == '45660091') ]
    df_Sam5_lot2 = df_Sam5_lot2.reset_index(drop = True)
    df_Sam5_lot2['index'] = df_Sam5_lot2.index.tolist()
    df_Sam5_lot2 = df_Sam5_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam5_lot3 = df[(df.Sample == 5) & (df.LOT == '45660101') ]
    df_Sam5_lot3 = df_Sam5_lot3.reset_index(drop = True)
    df_Sam5_lot3['index'] = df_Sam5_lot3.index.tolist()
    df_Sam5_lot3= df_Sam5_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam5_lot1,df_Sam5_lot2], axis = 1, sort = False)
    df_Sam5 = pd.concat([df1_df2,  df_Sam5_lot3], axis = 1, sort = False)
    df_Sam5 = df_Sam5.drop(['index'], axis = 1)
    
    df_Sam5.insert(11, 'blank','')
    df_Sam5.insert(18, 'blank2','')
    #################################################sample 6
    df_Sam6_lot1 = df[(df.Sample == 6) & (df.LOT == '45660082') ] 
    df_Sam6_lot1 = df_Sam6_lot1.reset_index(drop = True)
    df_Sam6_lot1['index'] =  df_Sam6_lot1.index.tolist()
    df_Sam6_lot1 = df_Sam6_lot1.drop(['LOT'], axis = 1)

    df_Sam6_lot2 = df[(df.Sample == 6) & (df.LOT == '45660091') ]
    df_Sam6_lot2 = df_Sam6_lot2.reset_index(drop = True)
    df_Sam6_lot2['index'] = df_Sam6_lot2.index.tolist()
    df_Sam6_lot2 = df_Sam6_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam6_lot3 = df[(df.Sample == 6) & (df.LOT == '45660101') ]
    df_Sam6_lot3 = df_Sam6_lot3.reset_index(drop = True)
    df_Sam6_lot3['index'] = df_Sam6_lot3.index.tolist()
    df_Sam6_lot3= df_Sam6_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam6_lot1,df_Sam6_lot2], axis = 1, sort = False)
    df_Sam6 = pd.concat([df1_df2,  df_Sam6_lot3], axis = 1, sort = False)
    df_Sam6 = df_Sam6.drop(['index'], axis = 1)
    
    df_Sam6.insert(11, 'blank','')
    df_Sam6.insert(18, 'blank2','')
    #############################################sample 7
    df_Sam7_lot1 = df[(df.Sample == 7) & (df.LOT == '45660082') ] 
    df_Sam7_lot1 = df_Sam7_lot1.reset_index(drop = True)
    df_Sam7_lot1['index'] =  df_Sam7_lot1.index.tolist()
    df_Sam7_lot1 = df_Sam7_lot1.drop(['LOT'], axis = 1)

    df_Sam7_lot2 = df[(df.Sample == 7) & (df.LOT == '45660091') ]
    df_Sam7_lot2 = df_Sam7_lot2.reset_index(drop = True)
    df_Sam7_lot2['index'] = df_Sam7_lot2.index.tolist()
    df_Sam7_lot2 = df_Sam7_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam7_lot3 = df[(df.Sample == 7) & (df.LOT == '45660101') ]
    df_Sam7_lot3 = df_Sam7_lot3.reset_index(drop = True)
    df_Sam7_lot3['index'] = df_Sam7_lot3.index.tolist()
    df_Sam7_lot3= df_Sam7_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam7_lot1,df_Sam7_lot2], axis = 1, sort = False)
    df_Sam7 = pd.concat([df1_df2,  df_Sam7_lot3], axis = 1, sort = False)
    df_Sam7 = df_Sam7.drop(['index'], axis = 1)
    
    df_Sam7.insert(11, 'blank','')
    df_Sam7.insert(18, 'blank2','')
    ################################################sample 8
    df_Sam8_lot1 = df[(df.Sample == 8) & (df.LOT == '45660082') ] 
    df_Sam8_lot1 = df_Sam8_lot1.reset_index(drop = True)
    df_Sam8_lot1['index'] =  df_Sam8_lot1.index.tolist()
    df_Sam8_lot1 = df_Sam8_lot1.drop(['LOT'], axis = 1)

    df_Sam8_lot2 = df[(df.Sample == 8) & (df.LOT == '45660091') ]
    df_Sam8_lot2 = df_Sam8_lot2.reset_index(drop = True)
    df_Sam8_lot2['index'] = df_Sam8_lot2.index.tolist()
    df_Sam8_lot2 = df_Sam8_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam8_lot3 = df[(df.Sample == 8) & (df.LOT == '45660101') ]
    df_Sam8_lot3 = df_Sam8_lot3.reset_index(drop = True)
    df_Sam8_lot3['index'] = df_Sam8_lot3.index.tolist()
    df_Sam8_lot3= df_Sam8_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam8_lot1,df_Sam8_lot2], axis = 1, sort = False)
    df_Sam8 = pd.concat([df1_df2,  df_Sam8_lot3], axis = 1, sort = False)
    df_Sam8 = df_Sam8.drop(['index'], axis = 1)
    
    df_Sam8.insert(11, 'blank','')
    df_Sam8.insert(18, 'blank2','')
    #############################################sample 9
    df_Sam9_lot1 = df[(df.Sample == 9) & (df.LOT == '45660082') ] 
    df_Sam9_lot1 = df_Sam9_lot1.reset_index(drop = True)
    df_Sam9_lot1['index'] =  df_Sam9_lot1.index.tolist()
    df_Sam9_lot1 = df_Sam9_lot1.drop(['LOT'], axis = 1)

    df_Sam9_lot2 = df[(df.Sample == 9) & (df.LOT == '45660091') ]
    df_Sam9_lot2 = df_Sam9_lot2.reset_index(drop = True)
    df_Sam9_lot2['index'] = df_Sam9_lot2.index.tolist()
    df_Sam9_lot2 = df_Sam9_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam9_lot3 = df[(df.Sample == 9) & (df.LOT == '45660101') ]
    df_Sam9_lot3 = df_Sam9_lot3.reset_index(drop = True)
    df_Sam9_lot3['index'] = df_Sam9_lot3.index.tolist()
    df_Sam9_lot3= df_Sam9_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam9_lot1,df_Sam9_lot2], axis = 1, sort = False)
    df_Sam9 = pd.concat([df1_df2,  df_Sam9_lot3], axis = 1, sort = False)
    df_Sam9 = df_Sam9.drop(['index'], axis = 1)
    
    df_Sam9.insert(11, 'blank','')
    df_Sam9.insert(18, 'blank2','')
    ##############################################sample 10
    df_Sam10_lot1 = df[(df.Sample == 10) & (df.LOT == '45660082') ] 
    df_Sam10_lot1 = df_Sam10_lot1.reset_index(drop = True)
#     for i in df_Sam10_lot1.columns.tolist():
#         if 'index' in i:
    df_Sam10_lot1['index'] =  df_Sam10_lot1.index.tolist()
    df_Sam10_lot1 = df_Sam10_lot1.drop(['LOT'], axis = 1)

    df_Sam10_lot2 = df[(df.Sample == 10) & (df.LOT == '45660091') ]
    df_Sam10_lot2 = df_Sam10_lot2.reset_index(drop = True)
    df_Sam10_lot2['index'] = df_Sam10_lot2.index.tolist()
    df_Sam10_lot2 = df_Sam10_lot2.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)


    df_Sam10_lot3 = df[(df.Sample == 10) & (df.LOT == '45660101') ]
    df_Sam10_lot3 = df_Sam10_lot3.reset_index(drop = True)
    df_Sam10_lot3['index'] = df_Sam10_lot3.index.tolist()
    df_Sam10_lot3= df_Sam10_lot3.drop(['Site', 'Day', 'AM-PM', 'Identifier', 'Sample', 'LOT'], axis = 1)

    df1_df2 = pd.concat([df_Sam10_lot1,df_Sam10_lot2], axis = 1, sort = False)
    df_Sam10 = pd.concat([df1_df2,  df_Sam10_lot3], axis = 1, sort = False)
    df_Sam10 = df_Sam10.drop(['index'], axis = 1)
    
    df_Sam10.insert(11, 'blank','')
    df_Sam10.insert(18, 'blank2','')
    
    return df_Sam1,df_Sam2,df_Sam3,df_Sam4,df_Sam5,df_Sam6,df_Sam7,df_Sam8,df_Sam9,df_Sam10

# def make_plus(dfplus):
dfpeteriggplus_Sam1,dfpeteriggplus_Sam2,dfpeteriggplus_Sam3,dfpeteriggplus_Sam4,dfpeteriggplus_Sam5,dfpeteriggplus_Sam6,dfpeteriggplus_Sam7,dfpeteriggplus_Sam8,dfpeteriggplus_Sam9,dfpeteriggplus_Sam10 = make_plus(dfpeter_iggplus)
dfpeterlissplus_Sam1,dfpeterlissplus_Sam2,dfpeterlissplus_Sam3,dfpeterlissplus_Sam4,dfpeterlissplus_Sam5,dfpeterlissplus_Sam6,dfpeterlissplus_Sam7,dfpeterlissplus_Sam8,dfpeterlissplus_Sam9,dfpeterlissplus_Sam10 = make_plus(dfpeter_lissplus)
dfpeternaclplus_Sam1,dfpeternaclplus_Sam2,dfpeternaclplus_Sam3,dfpeternaclplus_Sam4,dfpeternaclplus_Sam5,dfpeternaclplus_Sam6,dfpeternaclplus_Sam7,dfpeternaclplus_Sam8,dfpeternaclplus_Sam9,dfpeternaclplus_Sam10 = make_plus(dfpeter_naclplus)
########################################################################################################################################################################################################################################################
dfpeteriggproph_Sam1,dfpeteriggproph_Sam2,dfpeteriggproph_Sam3,dfpeteriggproph_Sam4,dfpeteriggproph_Sam5,dfpeteriggproph_Sam6,dfpeteriggproph_Sam7,dfpeteriggproph_Sam8,dfpeteriggproph_Sam9,dfpeteriggproph_Sam10 = make_proph(dfpeter_iggproph)
dfpeterlissproph_Sam1,dfpeterlissproph_Sam2,dfpeterlissproph_Sam3,dfpeterlissproph_Sam4,dfpeterlissproph_Sam5,dfpeterlissproph_Sam6,dfpeterlissproph_Sam7,dfpeterlissproph_Sam8,dfpeterlissproph_Sam9,dfpeterlissproph_Sam10 = make_proph(dfpeter_lissproph)
dfpeternaclproph_Sam1,dfpeternaclproph_Sam2,dfpeternaclproph_Sam3,dfpeternaclproph_Sam4,dfpeternaclproph_Sam5,dfpeternaclproph_Sam6,dfpeternaclproph_Sam7,dfpeternaclproph_Sam8,dfpeternaclproph_Sam9,dfpeternaclproph_Sam10 = make_proph(dfpeter_naclproph)

dfgenhosiggplus_Sam1,dfgenhosiggplus_Sam2,dfgenhosiggplus_Sam3,dfgenhosiggplus_Sam4,dfgenhosiggplus_Sam5,dfgenhosiggplus_Sam6,dfgenhosiggplus_Sam7,dfgenhosiggplus_Sam8,dfgenhosiggplus_Sam9,dfgenhosiggplus_Sam10 = make_plus(dfgenhos_iggplus)
dfgenhoslissplus_Sam1,dfgenhoslissplus_Sam2,dfgenhoslissplus_Sam3,dfgenhoslissplus_Sam4,dfgenhoslissplus_Sam5,dfgenhoslissplus_Sam6,dfgenhoslissplus_Sam7,dfgenhoslissplus_Sam8,dfgenhoslissplus_Sam9,dfgenhoslissplus_Sam10 = make_plus(dfgenhos_lissplus)
dfgenhosnaclplus_Sam1,dfgenhosnaclplus_Sam2,dfgenhosnaclplus_Sam3,dfgenhosnaclplus_Sam4,dfgenhosnaclplus_Sam5,dfgenhosnaclplus_Sam6,dfgenhosnaclplus_Sam7,dfgenhosnaclplus_Sam8,dfgenhosnaclplus_Sam9,dfgenhosnaclplus_Sam10 = make_plus(dfgenhos_naclplus)
########################################################################################################################################################################################################################################################
dfgenhosiggproph_Sam1,dfgenhosiggproph_Sam2,dfgenhosiggproph_Sam3,dfgenhosiggproph_Sam4,dfgenhosiggproph_Sam5,dfgenhosiggproph_Sam6,dfgenhosiggproph_Sam7,dfgenhosiggproph_Sam8,dfgenhosiggproph_Sam9,dfgenhosiggproph_Sam10 = make_proph(dfgenhos_iggproph)
dfgenhoslissproph_Sam1,dfgenhoslissproph_Sam2,dfgenhoslissproph_Sam3,dfgenhoslissproph_Sam4,dfgenhoslissproph_Sam5,dfgenhoslissproph_Sam6,dfgenhoslissproph_Sam7,dfgenhoslissproph_Sam8,dfgenhoslissproph_Sam9,dfgenhoslissproph_Sam10 = make_proph(dfgenhos_lissproph)
dfgenhosnaclproph_Sam1,dfgenhosnaclproph_Sam2,dfgenhosnaclproph_Sam3,dfgenhosnaclproph_Sam4,dfgenhosnaclproph_Sam5,dfgenhosnaclproph_Sam6,dfgenhosnaclproph_Sam7,dfgenhosnaclproph_Sam8,dfgenhosnaclproph_Sam9,dfgenhosnaclproph_Sam10 = make_proph(dfgenhos_naclproph)

dfbioradiggplus_Sam1,dfbioradiggplus_Sam2,dfbioradiggplus_Sam3,dfbioradiggplus_Sam4,dfbioradiggplus_Sam5,dfbioradiggplus_Sam6,dfbioradiggplus_Sam7,dfbioradiggplus_Sam8,dfbioradiggplus_Sam9,dfbioradiggplus_Sam10 = make_plus(dfbiorad_iggplus)
dfbioradlissplus_Sam1,dfbioradlissplus_Sam2,dfbioradlissplus_Sam3,dfbioradlissplus_Sam4,dfbioradlissplus_Sam5,dfbioradlissplus_Sam6,dfbioradlissplus_Sam7,dfbioradlissplus_Sam8,dfbioradlissplus_Sam9,dfbioradlissplus_Sam10 = make_plus(dfbiorad_lissplus)
dfbioradnaclplus_Sam1,dfbioradnaclplus_Sam2,dfbioradnaclplus_Sam3,dfbioradnaclplus_Sam4,dfbioradnaclplus_Sam5,dfbioradnaclplus_Sam6,dfbioradnaclplus_Sam7,dfbioradnaclplus_Sam8,dfbioradnaclplus_Sam9,dfbioradnaclplus_Sam10 = make_plus(dfbiorad_naclplus)
########################################################################################################################################################################################################################################################
dfbioradiggproph_Sam1,dfbioradiggproph_Sam2,dfbioradiggproph_Sam3,dfbioradiggproph_Sam4,dfbioradiggproph_Sam5,dfbioradiggproph_Sam6,dfbioradiggproph_Sam7,dfbioradiggproph_Sam8,dfbioradiggproph_Sam9,dfbioradiggproph_Sam10 = make_proph(dfbiorad_iggproph)
dfbioradlissproph_Sam1,dfbioradlissproph_Sam2,dfbioradlissproph_Sam3,dfbioradlissproph_Sam4,dfbioradlissproph_Sam5,dfbioradlissproph_Sam6,dfbioradlissproph_Sam7,dfbioradlissproph_Sam8,dfbioradlissproph_Sam9,dfbioradlissproph_Sam10 = make_proph(dfbiorad_lissproph)
dfbioradnaclproph_Sam1,dfbioradnaclproph_Sam2,dfbioradnaclproph_Sam3,dfbioradnaclproph_Sam4,dfbioradnaclproph_Sam5,dfbioradnaclproph_Sam6,dfbioradnaclproph_Sam7,dfbioradnaclproph_Sam8,dfbioradnaclproph_Sam9,dfbioradnaclproph_Sam10 = make_proph(dfbiorad_naclproph)













test_dict ={}

def Ident_3(df1,df2,df3):
    two_combined = pd.concat([df1,df2], sort = False)
    three_combined = pd.concat([two_combined,df3], sort = False)
    three_combined = three_combined.reset_index(drop = True)
    return three_combined



##############original was Sam1 df1 df2 df3   etc.##################

Sam1_iggplus = Ident_3(dfpeteriggplus_Sam1,dfgenhosiggplus_Sam1,dfbioradiggplus_Sam1)

Sam1_iggproph = Ident_3(dfpeteriggproph_Sam1,dfgenhosiggproph_Sam1,dfbioradiggproph_Sam1)

Sam1_lissplus = Ident_3(dfpeterlissplus_Sam1,dfgenhoslissplus_Sam1,dfbioradlissplus_Sam1)

Sam1_lissproph = Ident_3(dfpeterlissproph_Sam1,dfgenhoslissproph_Sam1,dfbioradlissproph_Sam1)

Sam1_naclplus = Ident_3(dfpeternaclplus_Sam1,dfgenhosnaclplus_Sam1,dfbioradnaclplus_Sam1)

Sam1_naclproph = Ident_3(dfpeternaclproph_Sam1,dfgenhosnaclproph_Sam1,dfbioradnaclproph_Sam1)
#############################################################################################

########################################################################################3




Sam2_iggplus = Ident_3(dfpeteriggplus_Sam2,dfgenhosiggplus_Sam2,dfbioradiggplus_Sam2)

Sam2_iggproph = Ident_3(dfpeteriggproph_Sam2,dfgenhosiggproph_Sam2,dfbioradiggproph_Sam2)

Sam2_lissplus = Ident_3(dfpeterlissplus_Sam2,dfgenhoslissplus_Sam2,dfbioradlissplus_Sam2)

Sam2_lissproph = Ident_3(dfpeterlissproph_Sam2,dfgenhoslissproph_Sam2,dfbioradlissproph_Sam2)

Sam2_naclplus = Ident_3(dfpeternaclplus_Sam2,dfgenhosnaclplus_Sam2,dfbioradnaclplus_Sam2)

Sam2_naclproph = Ident_3(dfpeternaclproph_Sam2,dfgenhosnaclproph_Sam2,dfbioradnaclproph_Sam2)

Sam3_iggplus = Ident_3(dfpeteriggplus_Sam3,dfgenhosiggplus_Sam3,dfbioradiggplus_Sam3)

Sam3_iggproph = Ident_3(dfpeteriggproph_Sam3,dfgenhosiggproph_Sam3,dfbioradiggproph_Sam3)

Sam3_lissplus = Ident_3(dfpeterlissplus_Sam3,dfgenhoslissplus_Sam3,dfbioradlissplus_Sam3)

Sam3_lissproph = Ident_3(dfpeterlissproph_Sam3,dfgenhoslissproph_Sam3,dfbioradlissproph_Sam3)

Sam3_naclplus = Ident_3(dfpeternaclplus_Sam3,dfgenhosnaclplus_Sam3,dfbioradnaclplus_Sam3)

Sam3_naclproph = Ident_3(dfpeternaclproph_Sam3,dfgenhosnaclproph_Sam3,dfbioradnaclproph_Sam3)

Sam4_iggplus = Ident_3(dfpeteriggplus_Sam4,dfgenhosiggplus_Sam4,dfbioradiggplus_Sam4)

Sam4_iggproph = Ident_3(dfpeteriggproph_Sam4,dfgenhosiggproph_Sam4,dfbioradiggproph_Sam4)

Sam4_lissplus = Ident_3(dfpeterlissplus_Sam4,dfgenhoslissplus_Sam4,dfbioradlissplus_Sam4)

Sam4_lissproph = Ident_3(dfpeterlissproph_Sam4,dfgenhoslissproph_Sam4,dfbioradlissproph_Sam4)

Sam4_naclplus = Ident_3(dfpeternaclplus_Sam4,dfgenhosnaclplus_Sam4,dfbioradnaclplus_Sam4)

Sam4_naclproph = Ident_3(dfpeternaclproph_Sam4,dfgenhosnaclproph_Sam4,dfbioradnaclproph_Sam4)

Sam5_iggplus = Ident_3(dfpeteriggplus_Sam5,dfgenhosiggplus_Sam5,dfbioradiggplus_Sam5)

Sam5_iggproph = Ident_3(dfpeteriggproph_Sam5,dfgenhosiggproph_Sam5,dfbioradiggproph_Sam5)

Sam5_lissplus = Ident_3(dfpeterlissplus_Sam5,dfgenhoslissplus_Sam5,dfbioradlissplus_Sam5)

Sam5_lissproph = Ident_3(dfpeterlissproph_Sam5,dfgenhoslissproph_Sam5,dfbioradlissproph_Sam5)

Sam5_naclplus = Ident_3(dfpeternaclplus_Sam5,dfgenhosnaclplus_Sam5,dfbioradnaclplus_Sam5)

Sam5_naclproph = Ident_3(dfpeternaclproph_Sam5,dfgenhosnaclproph_Sam5,dfbioradnaclproph_Sam5)

Sam6_iggplus = Ident_3(dfpeteriggplus_Sam6,dfgenhosiggplus_Sam6,dfbioradiggplus_Sam6)

Sam6_iggproph = Ident_3(dfpeteriggproph_Sam6,dfgenhosiggproph_Sam6,dfbioradiggproph_Sam6)

Sam6_lissplus = Ident_3(dfpeterlissplus_Sam6,dfgenhoslissplus_Sam6,dfbioradlissplus_Sam6)

Sam6_lissproph = Ident_3(dfpeterlissproph_Sam6,dfgenhoslissproph_Sam6,dfbioradlissproph_Sam6)

Sam6_naclplus = Ident_3(dfpeternaclplus_Sam6,dfgenhosnaclplus_Sam6,dfbioradnaclplus_Sam6)

Sam6_naclproph = Ident_3(dfpeternaclproph_Sam6,dfgenhosnaclproph_Sam6,dfbioradnaclproph_Sam6)

Sam7_iggplus = Ident_3(dfpeteriggplus_Sam7,dfgenhosiggplus_Sam7,dfbioradiggplus_Sam7)

Sam7_iggproph = Ident_3(dfpeteriggproph_Sam7,dfgenhosiggproph_Sam7,dfbioradiggproph_Sam7)

Sam7_lissplus = Ident_3(dfpeterlissplus_Sam7,dfgenhoslissplus_Sam7,dfbioradlissplus_Sam7)

Sam7_lissproph = Ident_3(dfpeterlissproph_Sam7,dfgenhoslissproph_Sam7,dfbioradlissproph_Sam7)

Sam7_naclplus = Ident_3(dfpeternaclplus_Sam7,dfgenhosnaclplus_Sam7,dfbioradnaclplus_Sam7)

Sam7_naclproph = Ident_3(dfpeternaclproph_Sam7,dfgenhosnaclproph_Sam7,dfbioradnaclproph_Sam7)

Sam8_iggplus = Ident_3(dfpeteriggplus_Sam8,dfgenhosiggplus_Sam8,dfbioradiggplus_Sam8)

Sam8_iggproph = Ident_3(dfpeteriggproph_Sam8,dfgenhosiggproph_Sam8,dfbioradiggproph_Sam8)

Sam8_lissplus = Ident_3(dfpeterlissplus_Sam8,dfgenhoslissplus_Sam8,dfbioradlissplus_Sam8)

Sam8_lissproph = Ident_3(dfpeterlissproph_Sam8,dfgenhoslissproph_Sam8,dfbioradlissproph_Sam8)

Sam8_naclplus = Ident_3(dfpeternaclplus_Sam8,dfgenhosnaclplus_Sam8,dfbioradnaclplus_Sam8)

Sam8_naclproph = Ident_3(dfpeternaclproph_Sam8,dfgenhosnaclproph_Sam8,dfbioradnaclproph_Sam8)

Sam9_iggplus = Ident_3(dfpeteriggplus_Sam9,dfgenhosiggplus_Sam9,dfbioradiggplus_Sam9)

Sam9_iggproph = Ident_3(dfpeteriggproph_Sam9,dfgenhosiggproph_Sam9,dfbioradiggproph_Sam9)

Sam9_lissplus = Ident_3(dfpeterlissplus_Sam9,dfgenhoslissplus_Sam9,dfbioradlissplus_Sam9)

Sam9_lissproph = Ident_3(dfpeterlissproph_Sam9,dfgenhoslissproph_Sam9,dfbioradlissproph_Sam9)

Sam9_naclplus = Ident_3(dfpeternaclplus_Sam9,dfgenhosnaclplus_Sam9,dfbioradnaclplus_Sam9)

Sam9_naclproph = Ident_3(dfpeternaclproph_Sam9,dfgenhosnaclproph_Sam9,dfbioradnaclproph_Sam9)

Sam10_iggplus = Ident_3(dfpeteriggplus_Sam10,dfgenhosiggplus_Sam10,dfbioradiggplus_Sam10)

Sam10_iggproph = Ident_3(dfpeteriggproph_Sam10,dfgenhosiggproph_Sam10,dfbioradiggproph_Sam10)

Sam10_lissplus = Ident_3(dfpeterlissplus_Sam10,dfgenhoslissplus_Sam10,dfbioradlissplus_Sam10)

Sam10_lissproph = Ident_3(dfpeterlissproph_Sam10,dfgenhoslissproph_Sam10,dfbioradlissproph_Sam10)

Sam10_naclplus = Ident_3(dfpeternaclplus_Sam10,dfgenhosnaclplus_Sam10,dfbioradnaclplus_Sam10)

Sam10_naclproph = Ident_3(dfpeternaclproph_Sam10,dfgenhosnaclproph_Sam10,dfbioradnaclproph_Sam10)

output_manual = r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_IAT_Coombs_IgG_Plus_split_v5.xlsx"







mm_writer = pd.ExcelWriter(output_manual, engine = 'xlsxwriter')
Sam1_iggplus.to_excel(mm_writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_iggplus.to_excel(mm_writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_iggplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618675293')



workbook  = mm_writer.book
worksheet = mm_writer.sheets
worksheet1 = mm_writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet10 = mm_writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet9 = mm_writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet8 = mm_writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet7 = mm_writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet6 = mm_writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet5 = mm_writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet4 = mm_writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet3 = mm_writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'IAT - Coombs IgG Plus ')
worksheet2 = mm_writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'IAT - Coombs IgG Plus ')


worksheet1.write_string(24,5,'Reaction')
worksheet1.write_string(25,5,'Highest')
worksheet1.write_string(26,5,'Lowest')
worksheet1.write_string(27,5,'Difference')
worksheet1.write_string(28,5,'4')
worksheet1.write_string(29,5,'3')
worksheet1.write_string(30,5,'2')
worksheet1.write_string(31,5,'1')
worksheet1.write_string(32,5,'0.5')
worksheet1.write_string(33,5,'0')

worksheet1.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet1.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet1.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet2.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet2.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet2.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet3.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet3.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet3.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet4.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet4.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet4.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet5.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet5.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet5.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet6.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet6.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet6.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet7.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet7.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet7.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet8.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet8.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet8.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet9.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet9.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet9.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet10.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet10.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet10.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','H','I','J','K','L','N','O','P','Q','R','S','U','V','W','X','Y','Z']
# for y in range(1,11):

for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet1.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-21), '=W141617235292!'+j+str(i))
        if i == 29:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            

worksheet2.write_string(24,5,'Reaction')
worksheet2.write_string(25,5,'Highest')
worksheet2.write_string(26,5,'Lowest')
worksheet2.write_string(27,5,'Difference')
worksheet2.write_string(28,5,'4')
worksheet2.write_string(29,5,'3')
worksheet2.write_string(30,5,'2')
worksheet2.write_string(31,5,'1')
worksheet2.write_string(32,5,'0.5')
worksheet2.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet2.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-20), '=W141617555914!'+j+str(i))
        if i == 29:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet3.write_string(24,5,'Reaction')
worksheet3.write_string(25,5,'Highest')
worksheet3.write_string(26,5,'Lowest')
worksheet3.write_string(27,5,'Difference')
worksheet3.write_string(28,5,'4')
worksheet3.write_string(29,5,'3')
worksheet3.write_string(30,5,'2')
worksheet3.write_string(31,5,'1')
worksheet3.write_string(32,5,'0.5')
worksheet3.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet3.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-19), '=W141618750091!'+j+str(i))
        if i == 29:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet4.write_string(24,5,'Reaction')
worksheet4.write_string(25,5,'Highest')
worksheet4.write_string(26,5,'Lowest')
worksheet4.write_string(27,5,'Difference')
worksheet4.write_string(28,5,'4')
worksheet4.write_string(29,5,'3')
worksheet4.write_string(30,5,'2')
worksheet4.write_string(31,5,'1')
worksheet4.write_string(32,5,'0.5')
worksheet4.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet4.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-18), '=W141618591839!'+j+str(i))
        if i == 29:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet5.write_string(24,5,'Reaction')
worksheet5.write_string(25,5,'Highest')
worksheet5.write_string(26,5,'Lowest')
worksheet5.write_string(27,5,'Difference')
worksheet5.write_string(28,5,'4')
worksheet5.write_string(29,5,'3')
worksheet5.write_string(30,5,'2')
worksheet5.write_string(31,5,'1')
worksheet5.write_string(32,5,'0.5')
worksheet5.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet5.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-17), '=W141618133512!'+j+str(i))
        if i == 29:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(24,5,'Reaction')
worksheet6.write_string(25,5,'Highest')
worksheet6.write_string(26,5,'Lowest')
worksheet6.write_string(27,5,'Difference')
worksheet6.write_string(28,5,'4')
worksheet6.write_string(29,5,'3')
worksheet6.write_string(30,5,'2')
worksheet6.write_string(31,5,'1')
worksheet6.write_string(32,5,'0.5')
worksheet6.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet6.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-16), '=W141618154440!'+j+str(i))
        if i == 29:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet7.write_string(24,5,'Reaction')
worksheet7.write_string(25,5,'Highest')
worksheet7.write_string(26,5,'Lowest')
worksheet7.write_string(27,5,'Difference')
worksheet7.write_string(28,5,'4')
worksheet7.write_string(29,5,'3')
worksheet7.write_string(30,5,'2')
worksheet7.write_string(31,5,'1')
worksheet7.write_string(32,5,'0.5')
worksheet7.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet7.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-15), '=W137518119128!'+j+str(i))
        if i == 29:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet8.write_string(24,5,'Reaction')
worksheet8.write_string(25,5,'Highest')
worksheet8.write_string(26,5,'Lowest')
worksheet8.write_string(27,5,'Difference')
worksheet8.write_string(28,5,'4')
worksheet8.write_string(29,5,'3')
worksheet8.write_string(30,5,'2')
worksheet8.write_string(31,5,'1')
worksheet8.write_string(32,5,'0.5')
worksheet8.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet8.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-14), '=W141619730500!'+j+str(i))
        if i == 29:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet9.write_string(24,5,'Reaction')
worksheet9.write_string(25,5,'Highest')
worksheet9.write_string(26,5,'Lowest')
worksheet9.write_string(27,5,'Difference')
worksheet9.write_string(28,5,'4')
worksheet9.write_string(29,5,'3')
worksheet9.write_string(30,5,'2')
worksheet9.write_string(31,5,'1')
worksheet9.write_string(32,5,'0.5')
worksheet9.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet9.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-13), '=W141618534656!'+j+str(i))
        if i == 29:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(24,5,'Reaction')
worksheet10.write_string(25,5,'Highest')
worksheet10.write_string(26,5,'Lowest')
worksheet10.write_string(27,5,'Difference')
worksheet10.write_string(28,5,'4')
worksheet10.write_string(29,5,'3')
worksheet10.write_string(30,5,'2')
worksheet10.write_string(31,5,'1')
worksheet10.write_string(32,5,'0.5')
worksheet10.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet10.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-12), '=W141618675293!'+j+str(i))
        if i == 29:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet11.write_string('K5','Panel Members')
worksheet11.write_string('R5','Panel Members')
worksheet11.write_string('Y5','Panel Members')
worksheet11.write_string('G6','1')
worksheet11.write_string('H6','2')
worksheet11.write_string('I6','3')
worksheet11.write_string('J6','4')
worksheet11.write_string('K6','5')
worksheet11.write_string('L6','6')
worksheet11.write_string('N6','1')
worksheet11.write_string('O6','2')
worksheet11.write_string('P6','3')
worksheet11.write_string('Q6','4')
worksheet11.write_string('R6','5')
worksheet11.write_string('S6','6')
worksheet11.write_string('U6','1')
worksheet11.write_string('V6','2')
worksheet11.write_string('W6','3')
worksheet11.write_string('X6','4')
worksheet11.write_string('Y6','5')
worksheet11.write_string('Z6','6')
worksheet11.write_string('E7','W141617235292')
worksheet11.write_string('E8','W141617555914')
worksheet11.write_string('E9','W141618750091')
worksheet11.write_string('E10','W141618591839')
worksheet11.write_string('E11','W141618133512')
worksheet11.write_string('E12','W141618154440')
worksheet11.write_string('E13','W137518119128')
worksheet11.write_string('E14','W141619730500')
worksheet11.write_string('E15','W141618534656')
worksheet11.write_string('E16','W141618675293')

worksheet11.write_string('F17','Difference High - Low')
worksheet11.write_string('F18', 'Frequency')
worksheet11.write_string('F19','4')
worksheet11.write_string('F20','3')
worksheet11.write_string('F21','2')
worksheet11.write_string('F22','1')
worksheet11.write_string('F23','0.5')
worksheet11.write_string('F24','0')
worksheet11.write_string('M17','Difference High - Low')
worksheet11.write_string('M18', 'Frequency')
worksheet11.write_string('M19','4')
worksheet11.write_string('M20','3')
worksheet11.write_string('M21','2')
worksheet11.write_string('M22','1')
worksheet11.write_string('M23','0.5')
worksheet11.write_string('M24','0')
worksheet11.write_string('T17','Difference High - Low')
worksheet11.write_string('T18', 'Frequency')
worksheet11.write_string('T19','4')
worksheet11.write_string('T20','3')
worksheet11.write_string('T21','2')
worksheet11.write_string('T22','1')
worksheet11.write_string('T23','0.5')
worksheet11.write_string('T24','0')
worksheet11.write('H19','=IF((COUNTIF(G7:L16,"=4")/60)=0,"",COUNTIF(G7:L16,"=4")/60)',percent_format)
worksheet11.write('H20','=IF((COUNTIF(G7:L16,"=3")/60)=0,"",COUNTIF(G7:L16,"=3")/60)',percent_format)
worksheet11.write('H21','=IF((COUNTIF(G7:L16,"=2")/60)=0,"",COUNTIF(G7:L16,"=2")/60)',percent_format)
worksheet11.write('H22','=IF((COUNTIF(G7:L16,"=1")/60)=0,"",COUNTIF(G7:L16,"=1")/60)',percent_format)
worksheet11.write('H23','=IF((COUNTIF(G7:L16,"=0.5")/60)=0,"",COUNTIF(G7:L16,"=0.5")/60)',percent_format)
worksheet11.write('H24','=IF((COUNTIF(G7:L16,"=0")/60)=0,"",COUNTIF(G7:L16,"=0")/60)',percent_format)

worksheet11.write('O19','=IF((COUNTIF(N7:S16,"=4")/60)=0,"",COUNTIF(N7:S16,"=4")/60)',percent_format)
worksheet11.write('O20','=IF((COUNTIF(N7:S16,"=3")/60)=0,"",COUNTIF(N7:S16,"=3")/60)',percent_format)
worksheet11.write('O21','=IF((COUNTIF(N7:S16,"=2")/60)=0,"",COUNTIF(N7:S16,"=2")/60)',percent_format)
worksheet11.write('O22','=IF((COUNTIF(N7:S16,"=1")/60)=0,"",COUNTIF(N7:S16,"=1")/60)',percent_format)
worksheet11.write('O23','=IF((COUNTIF(N7:S16,"=0.5")/60)=0,"",COUNTIF(N7:S16,"=0.5")/60)',percent_format)
worksheet11.write('O24','=IF((COUNTIF(N7:S16,"=0")/60)=0,"",COUNTIF(N7:S16,"=0")/60)',percent_format)

worksheet11.write('V19','=IF((COUNTIF(U7:Z16,"=4")/60)=0,"",COUNTIF(U7:Z16,"=4")/60)',percent_format)
worksheet11.write('V20','=IF((COUNTIF(U7:Z16,"=3")/60)=0,"",COUNTIF(U7:Z16,"=3")/60)',percent_format)
worksheet11.write('V21','=IF((COUNTIF(U7:Z16,"=2")/60)=0,"",COUNTIF(U7:Z16,"=2")/60)',percent_format)
worksheet11.write('V22','=IF((COUNTIF(U7:Z16,"=1")/60)=0,"",COUNTIF(U7:Z16,"=1")/60)',percent_format)
worksheet11.write('V23','=IF((COUNTIF(U7:Z16,"=0.5")/60)=0,"",COUNTIF(U7:Z16,"=0.5")/60)',percent_format)
worksheet11.write('V24','=IF((COUNTIF(U7:Z16,"=0")/60)=0,"",COUNTIF(U7:Z16,"=0")/60)',percent_format)




            
worksheet1.set_column('A:G', 15)
worksheet2.set_column('A:G', 15)
worksheet3.set_column('A:G', 15)
worksheet4.set_column('A:G', 15)
worksheet5.set_column('A:G', 15)
worksheet6.set_column('A:G', 15)
worksheet7.set_column('A:G', 15)
worksheet8.set_column('A:G', 15)
worksheet9.set_column('A:G', 15)
worksheet10.set_column('A:G', 15)

mm_writer.save()
    






output_manual = r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_IAT_Coombs_IgG_Prophyl_split_v5.xlsx"

mm_writer = pd.ExcelWriter(output_manual, engine = 'xlsxwriter')
Sam1_iggproph.to_excel(mm_writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_iggproph.to_excel(mm_writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_iggproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618675293')


workbook  = mm_writer.book
worksheet1 = mm_writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 
worksheet2 = mm_writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'IAT - Coombs IgG + Prophyl ') 
worksheet3 = mm_writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 
worksheet4 = mm_writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'IAT - Coombs IgG + Prophyl ') 
worksheet5 = mm_writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 
worksheet6 = mm_writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 
worksheet7 = mm_writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 
worksheet8 = mm_writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 
worksheet9 = mm_writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 
worksheet10 = mm_writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'IAT - Coombs IgG + Prophyl') 

worksheet1.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet1.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet1.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet2.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet2.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet2.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet3.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet3.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet3.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet4.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet4.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet4.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet5.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet5.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet5.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet6.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet6.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet6.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet7.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet7.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet7.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet8.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet8.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet8.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet9.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet9.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet9.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet10.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet10.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet10.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))


worksheet1.write_string(24,5,'Reaction')
worksheet1.write_string(25,5,'Highest')
worksheet1.write_string(26,5,'Lowest')
worksheet1.write_string(27,5,'Difference')
worksheet1.write_string(28,5,'4')
worksheet1.write_string(29,5,'3')
worksheet1.write_string(30,5,'2')
worksheet1.write_string(31,5,'1')
worksheet1.write_string(32,5,'0.5')
worksheet1.write_string(33,5,'0')




worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','H','I','J','K','L','N','O','P','Q','R','S','U','V','W','X','Y','Z']
# for y in range(1,11):

for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet1.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-21), '=W141617235292!'+j+str(i))
        if i == 29:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            

worksheet2.write_string(24,5,'Reaction')
worksheet2.write_string(25,5,'Highest')
worksheet2.write_string(26,5,'Lowest')
worksheet2.write_string(27,5,'Difference')
worksheet2.write_string(28,5,'4')
worksheet2.write_string(29,5,'3')
worksheet2.write_string(30,5,'2')
worksheet2.write_string(31,5,'1')
worksheet2.write_string(32,5,'0.5')
worksheet2.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet2.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-20), '=W141617555914!'+j+str(i))
        if i == 29:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet3.write_string(24,5,'Reaction')
worksheet3.write_string(25,5,'Highest')
worksheet3.write_string(26,5,'Lowest')
worksheet3.write_string(27,5,'Difference')
worksheet3.write_string(28,5,'4')
worksheet3.write_string(29,5,'3')
worksheet3.write_string(30,5,'2')
worksheet3.write_string(31,5,'1')
worksheet3.write_string(32,5,'0.5')
worksheet3.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet3.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-19), '=W141618750091!'+j+str(i))
        if i == 29:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet4.write_string(24,5,'Reaction')
worksheet4.write_string(25,5,'Highest')
worksheet4.write_string(26,5,'Lowest')
worksheet4.write_string(27,5,'Difference')
worksheet4.write_string(28,5,'4')
worksheet4.write_string(29,5,'3')
worksheet4.write_string(30,5,'2')
worksheet4.write_string(31,5,'1')
worksheet4.write_string(32,5,'0.5')
worksheet4.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet4.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-18), '=W141618591839!'+j+str(i))
        if i == 29:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet5.write_string(24,5,'Reaction')
worksheet5.write_string(25,5,'Highest')
worksheet5.write_string(26,5,'Lowest')
worksheet5.write_string(27,5,'Difference')
worksheet5.write_string(28,5,'4')
worksheet5.write_string(29,5,'3')
worksheet5.write_string(30,5,'2')
worksheet5.write_string(31,5,'1')
worksheet5.write_string(32,5,'0.5')
worksheet5.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet5.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-17), '=W141618133512!'+j+str(i))
        if i == 29:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(24,5,'Reaction')
worksheet6.write_string(25,5,'Highest')
worksheet6.write_string(26,5,'Lowest')
worksheet6.write_string(27,5,'Difference')
worksheet6.write_string(28,5,'4')
worksheet6.write_string(29,5,'3')
worksheet6.write_string(30,5,'2')
worksheet6.write_string(31,5,'1')
worksheet6.write_string(32,5,'0.5')
worksheet6.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet6.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-16), '=W141618154440!'+j+str(i))
        if i == 29:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet7.write_string(24,5,'Reaction')
worksheet7.write_string(25,5,'Highest')
worksheet7.write_string(26,5,'Lowest')
worksheet7.write_string(27,5,'Difference')
worksheet7.write_string(28,5,'4')
worksheet7.write_string(29,5,'3')
worksheet7.write_string(30,5,'2')
worksheet7.write_string(31,5,'1')
worksheet7.write_string(32,5,'0.5')
worksheet7.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet7.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-15), '=W137518119128!'+j+str(i))
        if i == 29:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet8.write_string(24,5,'Reaction')
worksheet8.write_string(25,5,'Highest')
worksheet8.write_string(26,5,'Lowest')
worksheet8.write_string(27,5,'Difference')
worksheet8.write_string(28,5,'4')
worksheet8.write_string(29,5,'3')
worksheet8.write_string(30,5,'2')
worksheet8.write_string(31,5,'1')
worksheet8.write_string(32,5,'0.5')
worksheet8.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet8.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-14), '=W141619730500!'+j+str(i))
        if i == 29:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet9.write_string(24,5,'Reaction')
worksheet9.write_string(25,5,'Highest')
worksheet9.write_string(26,5,'Lowest')
worksheet9.write_string(27,5,'Difference')
worksheet9.write_string(28,5,'4')
worksheet9.write_string(29,5,'3')
worksheet9.write_string(30,5,'2')
worksheet9.write_string(31,5,'1')
worksheet9.write_string(32,5,'0.5')
worksheet9.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet9.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-13), '=W141618534656!'+j+str(i))
        if i == 29:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(24,5,'Reaction')
worksheet10.write_string(25,5,'Highest')
worksheet10.write_string(26,5,'Lowest')
worksheet10.write_string(27,5,'Difference')
worksheet10.write_string(28,5,'4')
worksheet10.write_string(29,5,'3')
worksheet10.write_string(30,5,'2')
worksheet10.write_string(31,5,'1')
worksheet10.write_string(32,5,'0.5')
worksheet10.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet10.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-12), '=W141618675293!'+j+str(i))
        if i == 29:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet11.write_string('K5','Panel Members')
worksheet11.write_string('R5','Panel Members')
worksheet11.write_string('Y5','Panel Members')
worksheet11.write_string('G6','1')
worksheet11.write_string('H6','2')
worksheet11.write_string('I6','3')
worksheet11.write_string('J6','4')
worksheet11.write_string('K6','5')
worksheet11.write_string('L6','6')
worksheet11.write_string('N6','1')
worksheet11.write_string('O6','2')
worksheet11.write_string('P6','3')
worksheet11.write_string('Q6','4')
worksheet11.write_string('R6','5')
worksheet11.write_string('S6','6')
worksheet11.write_string('U6','1')
worksheet11.write_string('V6','2')
worksheet11.write_string('W6','3')
worksheet11.write_string('X6','4')
worksheet11.write_string('Y6','5')
worksheet11.write_string('Z6','6')
worksheet11.write_string('E7','W141617235292')
worksheet11.write_string('E8','W141617555914')
worksheet11.write_string('E9','W141618750091')
worksheet11.write_string('E10','W141618591839')
worksheet11.write_string('E11','W141618133512')
worksheet11.write_string('E12','W141618154440')
worksheet11.write_string('E13','W137518119128')
worksheet11.write_string('E14','W141619730500')
worksheet11.write_string('E15','W141618534656')
worksheet11.write_string('E16','W141618675293')

worksheet11.write_string('F17','Difference High - Low')
worksheet11.write_string('F18', 'Frequency')
worksheet11.write_string('F19','4')
worksheet11.write_string('F20','3')
worksheet11.write_string('F21','2')
worksheet11.write_string('F22','1')
worksheet11.write_string('F23','0.5')
worksheet11.write_string('F24','0')
worksheet11.write_string('M17','Difference High - Low')
worksheet11.write_string('M18', 'Frequency')
worksheet11.write_string('M19','4')
worksheet11.write_string('M20','3')
worksheet11.write_string('M21','2')
worksheet11.write_string('M22','1')
worksheet11.write_string('M23','0.5')
worksheet11.write_string('M24','0')
worksheet11.write_string('T17','Difference High - Low')
worksheet11.write_string('T18', 'Frequency')
worksheet11.write_string('T19','4')
worksheet11.write_string('T20','3')
worksheet11.write_string('T21','2')
worksheet11.write_string('T22','1')
worksheet11.write_string('T23','0.5')
worksheet11.write_string('T24','0')
worksheet11.write('H19','=IF((COUNTIF(G7:L16,"=4")/60)=0,"",COUNTIF(G7:L16,"=4")/60)',percent_format)
worksheet11.write('H20','=IF((COUNTIF(G7:L16,"=3")/60)=0,"",COUNTIF(G7:L16,"=3")/60)',percent_format)
worksheet11.write('H21','=IF((COUNTIF(G7:L16,"=2")/60)=0,"",COUNTIF(G7:L16,"=2")/60)',percent_format)
worksheet11.write('H22','=IF((COUNTIF(G7:L16,"=1")/60)=0,"",COUNTIF(G7:L16,"=1")/60)',percent_format)
worksheet11.write('H23','=IF((COUNTIF(G7:L16,"=0.5")/60)=0,"",COUNTIF(G7:L16,"=0.5")/60)',percent_format)
worksheet11.write('H24','=IF((COUNTIF(G7:L16,"=0")/60)=0,"",COUNTIF(G7:L16,"=0")/60)',percent_format)

worksheet11.write('O19','=IF((COUNTIF(N7:S16,"=4")/60)=0,"",COUNTIF(N7:S16,"=4")/60)',percent_format)
worksheet11.write('O20','=IF((COUNTIF(N7:S16,"=3")/60)=0,"",COUNTIF(N7:S16,"=3")/60)',percent_format)
worksheet11.write('O21','=IF((COUNTIF(N7:S16,"=2")/60)=0,"",COUNTIF(N7:S16,"=2")/60)',percent_format)
worksheet11.write('O22','=IF((COUNTIF(N7:S16,"=1")/60)=0,"",COUNTIF(N7:S16,"=1")/60)',percent_format)
worksheet11.write('O23','=IF((COUNTIF(N7:S16,"=0.5")/60)=0,"",COUNTIF(N7:S16,"=0.5")/60)',percent_format)
worksheet11.write('O24','=IF((COUNTIF(N7:S16,"=0")/60)=0,"",COUNTIF(N7:S16,"=0")/60)',percent_format)

worksheet11.write('V19','=IF((COUNTIF(U7:Z16,"=4")/60)=0,"",COUNTIF(U7:Z16,"=4")/60)',percent_format)
worksheet11.write('V20','=IF((COUNTIF(U7:Z16,"=3")/60)=0,"",COUNTIF(U7:Z16,"=3")/60)',percent_format)
worksheet11.write('V21','=IF((COUNTIF(U7:Z16,"=2")/60)=0,"",COUNTIF(U7:Z16,"=2")/60)',percent_format)
worksheet11.write('V22','=IF((COUNTIF(U7:Z16,"=1")/60)=0,"",COUNTIF(U7:Z16,"=1")/60)',percent_format)
worksheet11.write('V23','=IF((COUNTIF(U7:Z16,"=0.5")/60)=0,"",COUNTIF(U7:Z16,"=0.5")/60)',percent_format)
worksheet11.write('V24','=IF((COUNTIF(U7:Z16,"=0")/60)=0,"",COUNTIF(U7:Z16,"=0")/60)',percent_format)




            
worksheet1.set_column('A:G', 15)
worksheet2.set_column('A:G', 15)
worksheet3.set_column('A:G', 15)
worksheet4.set_column('A:G', 15)
worksheet5.set_column('A:G', 15)
worksheet6.set_column('A:G', 15)
worksheet7.set_column('A:G', 15)
worksheet8.set_column('A:G', 15)
worksheet9.set_column('A:G', 15)
worksheet10.set_column('A:G', 15)

mm_writer.save()
    
    

output_manual = r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_IAT_LISS_Coombs_Plus_split_v5.xlsx"

mm_writer = pd.ExcelWriter(output_manual, engine = 'xlsxwriter')
Sam1_lissplus.to_excel(mm_writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_lissplus.to_excel(mm_writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_lissplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618675293')

workbook  = mm_writer.book
worksheet1 = mm_writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet2 = mm_writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet3 = mm_writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet4 = mm_writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'IAT - LISS Coombs Plus ') 
worksheet5 = mm_writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet6 = mm_writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet7 = mm_writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet8 = mm_writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet9 = mm_writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'IAT - LISS Coombs Plus') 
worksheet10 = mm_writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'IAT - LISS Coombs Plus') 

worksheet1.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet1.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet1.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet2.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet2.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet2.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet3.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet3.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet3.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet4.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet4.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet4.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet5.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet5.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet5.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet6.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet6.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet6.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet7.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet7.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet7.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet8.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet8.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet8.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet9.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet9.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet9.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet10.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet10.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet10.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet1.write_string(24,5,'Reaction')
worksheet1.write_string(25,5,'Highest')
worksheet1.write_string(26,5,'Lowest')
worksheet1.write_string(27,5,'Difference')
worksheet1.write_string(28,5,'4')
worksheet1.write_string(29,5,'3')
worksheet1.write_string(30,5,'2')
worksheet1.write_string(31,5,'1')
worksheet1.write_string(32,5,'0.5')
worksheet1.write_string(33,5,'0')



worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','H','I','J','K','L','N','O','P','Q','R','S','U','V','W','X','Y','Z']
# for y in range(1,11):

for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet1.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-21), '=W141617235292!'+j+str(i))
        if i == 29:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            

worksheet2.write_string(24,5,'Reaction')
worksheet2.write_string(25,5,'Highest')
worksheet2.write_string(26,5,'Lowest')
worksheet2.write_string(27,5,'Difference')
worksheet2.write_string(28,5,'4')
worksheet2.write_string(29,5,'3')
worksheet2.write_string(30,5,'2')
worksheet2.write_string(31,5,'1')
worksheet2.write_string(32,5,'0.5')
worksheet2.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet2.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-20), '=W141617555914!'+j+str(i))
        if i == 29:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet3.write_string(24,5,'Reaction')
worksheet3.write_string(25,5,'Highest')
worksheet3.write_string(26,5,'Lowest')
worksheet3.write_string(27,5,'Difference')
worksheet3.write_string(28,5,'4')
worksheet3.write_string(29,5,'3')
worksheet3.write_string(30,5,'2')
worksheet3.write_string(31,5,'1')
worksheet3.write_string(32,5,'0.5')
worksheet3.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet3.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-19), '=W141618750091!'+j+str(i))
        if i == 29:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet4.write_string(24,5,'Reaction')
worksheet4.write_string(25,5,'Highest')
worksheet4.write_string(26,5,'Lowest')
worksheet4.write_string(27,5,'Difference')
worksheet4.write_string(28,5,'4')
worksheet4.write_string(29,5,'3')
worksheet4.write_string(30,5,'2')
worksheet4.write_string(31,5,'1')
worksheet4.write_string(32,5,'0.5')
worksheet4.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet4.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-18), '=W141618591839!'+j+str(i))
        if i == 29:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet5.write_string(24,5,'Reaction')
worksheet5.write_string(25,5,'Highest')
worksheet5.write_string(26,5,'Lowest')
worksheet5.write_string(27,5,'Difference')
worksheet5.write_string(28,5,'4')
worksheet5.write_string(29,5,'3')
worksheet5.write_string(30,5,'2')
worksheet5.write_string(31,5,'1')
worksheet5.write_string(32,5,'0.5')
worksheet5.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet5.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-17), '=W141618133512!'+j+str(i))
        if i == 29:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(24,5,'Reaction')
worksheet6.write_string(25,5,'Highest')
worksheet6.write_string(26,5,'Lowest')
worksheet6.write_string(27,5,'Difference')
worksheet6.write_string(28,5,'4')
worksheet6.write_string(29,5,'3')
worksheet6.write_string(30,5,'2')
worksheet6.write_string(31,5,'1')
worksheet6.write_string(32,5,'0.5')
worksheet6.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet6.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-16), '=W141618154440!'+j+str(i))
        if i == 29:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet7.write_string(24,5,'Reaction')
worksheet7.write_string(25,5,'Highest')
worksheet7.write_string(26,5,'Lowest')
worksheet7.write_string(27,5,'Difference')
worksheet7.write_string(28,5,'4')
worksheet7.write_string(29,5,'3')
worksheet7.write_string(30,5,'2')
worksheet7.write_string(31,5,'1')
worksheet7.write_string(32,5,'0.5')
worksheet7.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet7.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-15), '=W137518119128!'+j+str(i))
        if i == 29:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet8.write_string(24,5,'Reaction')
worksheet8.write_string(25,5,'Highest')
worksheet8.write_string(26,5,'Lowest')
worksheet8.write_string(27,5,'Difference')
worksheet8.write_string(28,5,'4')
worksheet8.write_string(29,5,'3')
worksheet8.write_string(30,5,'2')
worksheet8.write_string(31,5,'1')
worksheet8.write_string(32,5,'0.5')
worksheet8.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet8.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-14), '=W141619730500!'+j+str(i))
        if i == 29:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet9.write_string(24,5,'Reaction')
worksheet9.write_string(25,5,'Highest')
worksheet9.write_string(26,5,'Lowest')
worksheet9.write_string(27,5,'Difference')
worksheet9.write_string(28,5,'4')
worksheet9.write_string(29,5,'3')
worksheet9.write_string(30,5,'2')
worksheet9.write_string(31,5,'1')
worksheet9.write_string(32,5,'0.5')
worksheet9.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet9.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-13), '=W141618534656!'+j+str(i))
        if i == 29:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(24,5,'Reaction')
worksheet10.write_string(25,5,'Highest')
worksheet10.write_string(26,5,'Lowest')
worksheet10.write_string(27,5,'Difference')
worksheet10.write_string(28,5,'4')
worksheet10.write_string(29,5,'3')
worksheet10.write_string(30,5,'2')
worksheet10.write_string(31,5,'1')
worksheet10.write_string(32,5,'0.5')
worksheet10.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet10.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-12), '=W141618675293!'+j+str(i))
        if i == 29:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet11.write_string('K5','Panel Members')
worksheet11.write_string('R5','Panel Members')
worksheet11.write_string('Y5','Panel Members')
worksheet11.write_string('G6','1')
worksheet11.write_string('H6','2')
worksheet11.write_string('I6','3')
worksheet11.write_string('J6','4')
worksheet11.write_string('K6','5')
worksheet11.write_string('L6','6')
worksheet11.write_string('N6','1')
worksheet11.write_string('O6','2')
worksheet11.write_string('P6','3')
worksheet11.write_string('Q6','4')
worksheet11.write_string('R6','5')
worksheet11.write_string('S6','6')
worksheet11.write_string('U6','1')
worksheet11.write_string('V6','2')
worksheet11.write_string('W6','3')
worksheet11.write_string('X6','4')
worksheet11.write_string('Y6','5')
worksheet11.write_string('Z6','6')
worksheet11.write_string('E7','W141617235292')
worksheet11.write_string('E8','W141617555914')
worksheet11.write_string('E9','W141618750091')
worksheet11.write_string('E10','W141618591839')
worksheet11.write_string('E11','W141618133512')
worksheet11.write_string('E12','W141618154440')
worksheet11.write_string('E13','W137518119128')
worksheet11.write_string('E14','W141619730500')
worksheet11.write_string('E15','W141618534656')
worksheet11.write_string('E16','W141618675293')

worksheet11.write_string('F17','Difference High - Low')
worksheet11.write_string('F18', 'Frequency')
worksheet11.write_string('F19','4')
worksheet11.write_string('F20','3')
worksheet11.write_string('F21','2')
worksheet11.write_string('F22','1')
worksheet11.write_string('F23','0.5')
worksheet11.write_string('F24','0')
worksheet11.write_string('M17','Difference High - Low')
worksheet11.write_string('M18', 'Frequency')
worksheet11.write_string('M19','4')
worksheet11.write_string('M20','3')
worksheet11.write_string('M21','2')
worksheet11.write_string('M22','1')
worksheet11.write_string('M23','0.5')
worksheet11.write_string('M24','0')
worksheet11.write_string('T17','Difference High - Low')
worksheet11.write_string('T18', 'Frequency')
worksheet11.write_string('T19','4')
worksheet11.write_string('T20','3')
worksheet11.write_string('T21','2')
worksheet11.write_string('T22','1')
worksheet11.write_string('T23','0.5')
worksheet11.write_string('T24','0')
worksheet11.write('H19','=IF((COUNTIF(G7:L16,"=4")/60)=0,"",COUNTIF(G7:L16,"=4")/60)',percent_format)
worksheet11.write('H20','=IF((COUNTIF(G7:L16,"=3")/60)=0,"",COUNTIF(G7:L16,"=3")/60)',percent_format)
worksheet11.write('H21','=IF((COUNTIF(G7:L16,"=2")/60)=0,"",COUNTIF(G7:L16,"=2")/60)',percent_format)
worksheet11.write('H22','=IF((COUNTIF(G7:L16,"=1")/60)=0,"",COUNTIF(G7:L16,"=1")/60)',percent_format)
worksheet11.write('H23','=IF((COUNTIF(G7:L16,"=0.5")/60)=0,"",COUNTIF(G7:L16,"=0.5")/60)',percent_format)
worksheet11.write('H24','=IF((COUNTIF(G7:L16,"=0")/60)=0,"",COUNTIF(G7:L16,"=0")/60)',percent_format)

worksheet11.write('O19','=IF((COUNTIF(N7:S16,"=4")/60)=0,"",COUNTIF(N7:S16,"=4")/60)',percent_format)
worksheet11.write('O20','=IF((COUNTIF(N7:S16,"=3")/60)=0,"",COUNTIF(N7:S16,"=3")/60)',percent_format)
worksheet11.write('O21','=IF((COUNTIF(N7:S16,"=2")/60)=0,"",COUNTIF(N7:S16,"=2")/60)',percent_format)
worksheet11.write('O22','=IF((COUNTIF(N7:S16,"=1")/60)=0,"",COUNTIF(N7:S16,"=1")/60)',percent_format)
worksheet11.write('O23','=IF((COUNTIF(N7:S16,"=0.5")/60)=0,"",COUNTIF(N7:S16,"=0.5")/60)',percent_format)
worksheet11.write('O24','=IF((COUNTIF(N7:S16,"=0")/60)=0,"",COUNTIF(N7:S16,"=0")/60)',percent_format)

worksheet11.write('V19','=IF((COUNTIF(U7:Z16,"=4")/60)=0,"",COUNTIF(U7:Z16,"=4")/60)',percent_format)
worksheet11.write('V20','=IF((COUNTIF(U7:Z16,"=3")/60)=0,"",COUNTIF(U7:Z16,"=3")/60)',percent_format)
worksheet11.write('V21','=IF((COUNTIF(U7:Z16,"=2")/60)=0,"",COUNTIF(U7:Z16,"=2")/60)',percent_format)
worksheet11.write('V22','=IF((COUNTIF(U7:Z16,"=1")/60)=0,"",COUNTIF(U7:Z16,"=1")/60)',percent_format)
worksheet11.write('V23','=IF((COUNTIF(U7:Z16,"=0.5")/60)=0,"",COUNTIF(U7:Z16,"=0.5")/60)',percent_format)
worksheet11.write('V24','=IF((COUNTIF(U7:Z16,"=0")/60)=0,"",COUNTIF(U7:Z16,"=0")/60)',percent_format)




            
worksheet1.set_column('A:G', 15)
worksheet2.set_column('A:G', 15)
worksheet3.set_column('A:G', 15)
worksheet4.set_column('A:G', 15)
worksheet5.set_column('A:G', 15)
worksheet6.set_column('A:G', 15)
worksheet7.set_column('A:G', 15)
worksheet8.set_column('A:G', 15)
worksheet9.set_column('A:G', 15)
worksheet10.set_column('A:G', 15)

mm_writer.save()
    


output_manual = r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_IAT_LISS_Coombs_Prophyl_split_v5.xlsx"

mm_writer = pd.ExcelWriter(output_manual, engine = 'xlsxwriter')
Sam1_lissproph.to_excel(mm_writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_lissproph.to_excel(mm_writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_lissproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618675293')

workbook  = mm_writer.book
worksheet1 = mm_writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet2 = mm_writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet3 = mm_writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet4 = mm_writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet5 = mm_writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet6 = mm_writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet7 = mm_writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet8 = mm_writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet9 = mm_writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 
worksheet10 = mm_writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'IAT - LISS Coombs + Prophyl') 



worksheet1.write_string(24,5,'Reaction')
worksheet1.write_string(25,5,'Highest')
worksheet1.write_string(26,5,'Lowest')
worksheet1.write_string(27,5,'Difference')
worksheet1.write_string(28,5,'4')
worksheet1.write_string(29,5,'3')
worksheet1.write_string(30,5,'2')
worksheet1.write_string(31,5,'1')
worksheet1.write_string(32,5,'0.5')
worksheet1.write_string(33,5,'0')

worksheet1.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet1.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet1.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet2.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet2.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet2.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet3.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet3.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet3.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet4.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet4.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet4.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet5.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet5.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet5.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet6.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet6.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet6.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet7.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet7.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet7.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet8.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet8.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet8.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet9.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet9.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet9.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet10.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet10.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet10.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','H','I','J','K','L','N','O','P','Q','R','S','U','V','W','X','Y','Z']
# for y in range(1,11):

for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet1.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-21), '=W141617235292!'+j+str(i))
        if i == 29:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            

worksheet2.write_string(24,5,'Reaction')
worksheet2.write_string(25,5,'Highest')
worksheet2.write_string(26,5,'Lowest')
worksheet2.write_string(27,5,'Difference')
worksheet2.write_string(28,5,'4')
worksheet2.write_string(29,5,'3')
worksheet2.write_string(30,5,'2')
worksheet2.write_string(31,5,'1')
worksheet2.write_string(32,5,'0.5')
worksheet2.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet2.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-20), '=W141617555914!'+j+str(i))
        if i == 29:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet3.write_string(24,5,'Reaction')
worksheet3.write_string(25,5,'Highest')
worksheet3.write_string(26,5,'Lowest')
worksheet3.write_string(27,5,'Difference')
worksheet3.write_string(28,5,'4')
worksheet3.write_string(29,5,'3')
worksheet3.write_string(30,5,'2')
worksheet3.write_string(31,5,'1')
worksheet3.write_string(32,5,'0.5')
worksheet3.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet3.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-19), '=W141618750091!'+j+str(i))
        if i == 29:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet4.write_string(24,5,'Reaction')
worksheet4.write_string(25,5,'Highest')
worksheet4.write_string(26,5,'Lowest')
worksheet4.write_string(27,5,'Difference')
worksheet4.write_string(28,5,'4')
worksheet4.write_string(29,5,'3')
worksheet4.write_string(30,5,'2')
worksheet4.write_string(31,5,'1')
worksheet4.write_string(32,5,'0.5')
worksheet4.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet4.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-18), '=W141618591839!'+j+str(i))
        if i == 29:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet5.write_string(24,5,'Reaction')
worksheet5.write_string(25,5,'Highest')
worksheet5.write_string(26,5,'Lowest')
worksheet5.write_string(27,5,'Difference')
worksheet5.write_string(28,5,'4')
worksheet5.write_string(29,5,'3')
worksheet5.write_string(30,5,'2')
worksheet5.write_string(31,5,'1')
worksheet5.write_string(32,5,'0.5')
worksheet5.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet5.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-17), '=W141618133512!'+j+str(i))
        if i == 29:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(24,5,'Reaction')
worksheet6.write_string(25,5,'Highest')
worksheet6.write_string(26,5,'Lowest')
worksheet6.write_string(27,5,'Difference')
worksheet6.write_string(28,5,'4')
worksheet6.write_string(29,5,'3')
worksheet6.write_string(30,5,'2')
worksheet6.write_string(31,5,'1')
worksheet6.write_string(32,5,'0.5')
worksheet6.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet6.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-16), '=W141618154440!'+j+str(i))
        if i == 29:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet7.write_string(24,5,'Reaction')
worksheet7.write_string(25,5,'Highest')
worksheet7.write_string(26,5,'Lowest')
worksheet7.write_string(27,5,'Difference')
worksheet7.write_string(28,5,'4')
worksheet7.write_string(29,5,'3')
worksheet7.write_string(30,5,'2')
worksheet7.write_string(31,5,'1')
worksheet7.write_string(32,5,'0.5')
worksheet7.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet7.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-15), '=W137518119128!'+j+str(i))
        if i == 29:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet8.write_string(24,5,'Reaction')
worksheet8.write_string(25,5,'Highest')
worksheet8.write_string(26,5,'Lowest')
worksheet8.write_string(27,5,'Difference')
worksheet8.write_string(28,5,'4')
worksheet8.write_string(29,5,'3')
worksheet8.write_string(30,5,'2')
worksheet8.write_string(31,5,'1')
worksheet8.write_string(32,5,'0.5')
worksheet8.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet8.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-14), '=W141619730500!'+j+str(i))
        if i == 29:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet9.write_string(24,5,'Reaction')
worksheet9.write_string(25,5,'Highest')
worksheet9.write_string(26,5,'Lowest')
worksheet9.write_string(27,5,'Difference')
worksheet9.write_string(28,5,'4')
worksheet9.write_string(29,5,'3')
worksheet9.write_string(30,5,'2')
worksheet9.write_string(31,5,'1')
worksheet9.write_string(32,5,'0.5')
worksheet9.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet9.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-13), '=W141618534656!'+j+str(i))
        if i == 29:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(24,5,'Reaction')
worksheet10.write_string(25,5,'Highest')
worksheet10.write_string(26,5,'Lowest')
worksheet10.write_string(27,5,'Difference')
worksheet10.write_string(28,5,'4')
worksheet10.write_string(29,5,'3')
worksheet10.write_string(30,5,'2')
worksheet10.write_string(31,5,'1')
worksheet10.write_string(32,5,'0.5')
worksheet10.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet10.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-12), '=W141618675293!'+j+str(i))
        if i == 29:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet11.write_string('K5','Panel Members')
worksheet11.write_string('R5','Panel Members')
worksheet11.write_string('Y5','Panel Members')
worksheet11.write_string('G6','1')
worksheet11.write_string('H6','2')
worksheet11.write_string('I6','3')
worksheet11.write_string('J6','4')
worksheet11.write_string('K6','5')
worksheet11.write_string('L6','6')
worksheet11.write_string('N6','1')
worksheet11.write_string('O6','2')
worksheet11.write_string('P6','3')
worksheet11.write_string('Q6','4')
worksheet11.write_string('R6','5')
worksheet11.write_string('S6','6')
worksheet11.write_string('U6','1')
worksheet11.write_string('V6','2')
worksheet11.write_string('W6','3')
worksheet11.write_string('X6','4')
worksheet11.write_string('Y6','5')
worksheet11.write_string('Z6','6')
worksheet11.write_string('E7','W141617235292')
worksheet11.write_string('E8','W141617555914')
worksheet11.write_string('E9','W141618750091')
worksheet11.write_string('E10','W141618591839')
worksheet11.write_string('E11','W141618133512')
worksheet11.write_string('E12','W141618154440')
worksheet11.write_string('E13','W137518119128')
worksheet11.write_string('E14','W141619730500')
worksheet11.write_string('E15','W141618534656')
worksheet11.write_string('E16','W141618675293')

worksheet11.write_string('F17','Difference High - Low')
worksheet11.write_string('F18', 'Frequency')
worksheet11.write_string('F19','4')
worksheet11.write_string('F20','3')
worksheet11.write_string('F21','2')
worksheet11.write_string('F22','1')
worksheet11.write_string('F23','0.5')
worksheet11.write_string('F24','0')
worksheet11.write_string('M17','Difference High - Low')
worksheet11.write_string('M18', 'Frequency')
worksheet11.write_string('M19','4')
worksheet11.write_string('M20','3')
worksheet11.write_string('M21','2')
worksheet11.write_string('M22','1')
worksheet11.write_string('M23','0.5')
worksheet11.write_string('M24','0')
worksheet11.write_string('T17','Difference High - Low')
worksheet11.write_string('T18', 'Frequency')
worksheet11.write_string('T19','4')
worksheet11.write_string('T20','3')
worksheet11.write_string('T21','2')
worksheet11.write_string('T22','1')
worksheet11.write_string('T23','0.5')
worksheet11.write_string('T24','0')
worksheet11.write('H19','=IF((COUNTIF(G7:L16,"=4")/60)=0,"",COUNTIF(G7:L16,"=4")/60)',percent_format)
worksheet11.write('H20','=IF((COUNTIF(G7:L16,"=3")/60)=0,"",COUNTIF(G7:L16,"=3")/60)',percent_format)
worksheet11.write('H21','=IF((COUNTIF(G7:L16,"=2")/60)=0,"",COUNTIF(G7:L16,"=2")/60)',percent_format)
worksheet11.write('H22','=IF((COUNTIF(G7:L16,"=1")/60)=0,"",COUNTIF(G7:L16,"=1")/60)',percent_format)
worksheet11.write('H23','=IF((COUNTIF(G7:L16,"=0.5")/60)=0,"",COUNTIF(G7:L16,"=0.5")/60)',percent_format)
worksheet11.write('H24','=IF((COUNTIF(G7:L16,"=0")/60)=0,"",COUNTIF(G7:L16,"=0")/60)',percent_format)

worksheet11.write('O19','=IF((COUNTIF(N7:S16,"=4")/60)=0,"",COUNTIF(N7:S16,"=4")/60)',percent_format)
worksheet11.write('O20','=IF((COUNTIF(N7:S16,"=3")/60)=0,"",COUNTIF(N7:S16,"=3")/60)',percent_format)
worksheet11.write('O21','=IF((COUNTIF(N7:S16,"=2")/60)=0,"",COUNTIF(N7:S16,"=2")/60)',percent_format)
worksheet11.write('O22','=IF((COUNTIF(N7:S16,"=1")/60)=0,"",COUNTIF(N7:S16,"=1")/60)',percent_format)
worksheet11.write('O23','=IF((COUNTIF(N7:S16,"=0.5")/60)=0,"",COUNTIF(N7:S16,"=0.5")/60)',percent_format)
worksheet11.write('O24','=IF((COUNTIF(N7:S16,"=0")/60)=0,"",COUNTIF(N7:S16,"=0")/60)',percent_format)

worksheet11.write('V19','=IF((COUNTIF(U7:Z16,"=4")/60)=0,"",COUNTIF(U7:Z16,"=4")/60)',percent_format)
worksheet11.write('V20','=IF((COUNTIF(U7:Z16,"=3")/60)=0,"",COUNTIF(U7:Z16,"=3")/60)',percent_format)
worksheet11.write('V21','=IF((COUNTIF(U7:Z16,"=2")/60)=0,"",COUNTIF(U7:Z16,"=2")/60)',percent_format)
worksheet11.write('V22','=IF((COUNTIF(U7:Z16,"=1")/60)=0,"",COUNTIF(U7:Z16,"=1")/60)',percent_format)
worksheet11.write('V23','=IF((COUNTIF(U7:Z16,"=0.5")/60)=0,"",COUNTIF(U7:Z16,"=0.5")/60)',percent_format)
worksheet11.write('V24','=IF((COUNTIF(U7:Z16,"=0")/60)=0,"",COUNTIF(U7:Z16,"=0")/60)',percent_format)




            
worksheet1.set_column('A:G', 15)
worksheet2.set_column('A:G', 15)
worksheet3.set_column('A:G', 15)
worksheet4.set_column('A:G', 15)
worksheet5.set_column('A:G', 15)
worksheet6.set_column('A:G', 15)
worksheet7.set_column('A:G', 15)
worksheet8.set_column('A:G', 15)
worksheet9.set_column('A:G', 15)
worksheet10.set_column('A:G', 15)

mm_writer.save()
    

output_manual = r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_NaCl_Plus_split_v5.xlsx"

mm_writer = pd.ExcelWriter(output_manual, engine = 'xlsxwriter')
Sam1_naclplus.to_excel(mm_writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_naclplus.to_excel(mm_writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_naclplus.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618675293')

workbook  = mm_writer.book
worksheet1 = mm_writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'NaCl - Plus ') 
worksheet2 = mm_writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'NaCl - Plus ') 
worksheet3 = mm_writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'NaCl - Plus ') 
worksheet4 = mm_writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'NaCl - Plus ') 
worksheet5 = mm_writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'NaCl - Plus ') 
worksheet6 = mm_writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'NaCl - Plus ') 
worksheet7 = mm_writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'NaCl - Plus ') 
worksheet8 = mm_writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'NaCl - Plus ') 
worksheet9 = mm_writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'NaCl - Plus ') 
worksheet10 = mm_writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'NaCl - Plus ') 

worksheet1.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet1.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet1.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet2.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet2.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet2.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet3.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet3.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet3.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet4.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet4.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet4.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet5.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet5.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet5.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet6.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet6.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet6.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet7.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet7.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet7.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet8.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet8.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet8.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet9.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet9.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet9.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))

worksheet10.write_string('J5', 'Lot: ' + str(lotpeter_iggplus[0][0]))
worksheet10.write_string('Q5','Lot: ' + str(lotpeter_iggplus[1][0]))
worksheet10.write_string('X5','Lot: ' + str(lotpeter_iggplus[2][0]))




worksheet1.write_string(24,5,'Reaction')
worksheet1.write_string(25,5,'Highest')
worksheet1.write_string(26,5,'Lowest')
worksheet1.write_string(27,5,'Difference')
worksheet1.write_string(28,5,'4')
worksheet1.write_string(29,5,'3')
worksheet1.write_string(30,5,'2')
worksheet1.write_string(31,5,'1')
worksheet1.write_string(32,5,'0.5')
worksheet1.write_string(33,5,'0')



worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','H','I','J','K','L','N','O','P','Q','R','S','U','V','W','X','Y','Z']
# for y in range(1,11):

for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet1.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-21), '=W141617235292!'+j+str(i))
        if i == 29:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            

worksheet2.write_string(24,5,'Reaction')
worksheet2.write_string(25,5,'Highest')
worksheet2.write_string(26,5,'Lowest')
worksheet2.write_string(27,5,'Difference')
worksheet2.write_string(28,5,'4')
worksheet2.write_string(29,5,'3')
worksheet2.write_string(30,5,'2')
worksheet2.write_string(31,5,'1')
worksheet2.write_string(32,5,'0.5')
worksheet2.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet2.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-20), '=W141617555914!'+j+str(i))
        if i == 29:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet3.write_string(24,5,'Reaction')
worksheet3.write_string(25,5,'Highest')
worksheet3.write_string(26,5,'Lowest')
worksheet3.write_string(27,5,'Difference')
worksheet3.write_string(28,5,'4')
worksheet3.write_string(29,5,'3')
worksheet3.write_string(30,5,'2')
worksheet3.write_string(31,5,'1')
worksheet3.write_string(32,5,'0.5')
worksheet3.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet3.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-19), '=W141618750091!'+j+str(i))
        if i == 29:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet4.write_string(24,5,'Reaction')
worksheet4.write_string(25,5,'Highest')
worksheet4.write_string(26,5,'Lowest')
worksheet4.write_string(27,5,'Difference')
worksheet4.write_string(28,5,'4')
worksheet4.write_string(29,5,'3')
worksheet4.write_string(30,5,'2')
worksheet4.write_string(31,5,'1')
worksheet4.write_string(32,5,'0.5')
worksheet4.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet4.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-18), '=W141618591839!'+j+str(i))
        if i == 29:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet5.write_string(24,5,'Reaction')
worksheet5.write_string(25,5,'Highest')
worksheet5.write_string(26,5,'Lowest')
worksheet5.write_string(27,5,'Difference')
worksheet5.write_string(28,5,'4')
worksheet5.write_string(29,5,'3')
worksheet5.write_string(30,5,'2')
worksheet5.write_string(31,5,'1')
worksheet5.write_string(32,5,'0.5')
worksheet5.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet5.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-17), '=W141618133512!'+j+str(i))
        if i == 29:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(24,5,'Reaction')
worksheet6.write_string(25,5,'Highest')
worksheet6.write_string(26,5,'Lowest')
worksheet6.write_string(27,5,'Difference')
worksheet6.write_string(28,5,'4')
worksheet6.write_string(29,5,'3')
worksheet6.write_string(30,5,'2')
worksheet6.write_string(31,5,'1')
worksheet6.write_string(32,5,'0.5')
worksheet6.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet6.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-16), '=W141618154440!'+j+str(i))
        if i == 29:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet7.write_string(24,5,'Reaction')
worksheet7.write_string(25,5,'Highest')
worksheet7.write_string(26,5,'Lowest')
worksheet7.write_string(27,5,'Difference')
worksheet7.write_string(28,5,'4')
worksheet7.write_string(29,5,'3')
worksheet7.write_string(30,5,'2')
worksheet7.write_string(31,5,'1')
worksheet7.write_string(32,5,'0.5')
worksheet7.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet7.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-15), '=W137518119128!'+j+str(i))
        if i == 29:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet8.write_string(24,5,'Reaction')
worksheet8.write_string(25,5,'Highest')
worksheet8.write_string(26,5,'Lowest')
worksheet8.write_string(27,5,'Difference')
worksheet8.write_string(28,5,'4')
worksheet8.write_string(29,5,'3')
worksheet8.write_string(30,5,'2')
worksheet8.write_string(31,5,'1')
worksheet8.write_string(32,5,'0.5')
worksheet8.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet8.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-14), '=W141619730500!'+j+str(i))
        if i == 29:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet9.write_string(24,5,'Reaction')
worksheet9.write_string(25,5,'Highest')
worksheet9.write_string(26,5,'Lowest')
worksheet9.write_string(27,5,'Difference')
worksheet9.write_string(28,5,'4')
worksheet9.write_string(29,5,'3')
worksheet9.write_string(30,5,'2')
worksheet9.write_string(31,5,'1')
worksheet9.write_string(32,5,'0.5')
worksheet9.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet9.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-13), '=W141618534656!'+j+str(i))
        if i == 29:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(24,5,'Reaction')
worksheet10.write_string(25,5,'Highest')
worksheet10.write_string(26,5,'Lowest')
worksheet10.write_string(27,5,'Difference')
worksheet10.write_string(28,5,'4')
worksheet10.write_string(29,5,'3')
worksheet10.write_string(30,5,'2')
worksheet10.write_string(31,5,'1')
worksheet10.write_string(32,5,'0.5')
worksheet10.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet10.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-12), '=W141618675293!'+j+str(i))
        if i == 29:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet11.write_string('K5','Panel Members')
worksheet11.write_string('R5','Panel Members')
worksheet11.write_string('Y5','Panel Members')
worksheet11.write_string('G6','1')
worksheet11.write_string('H6','2')
worksheet11.write_string('I6','3')
worksheet11.write_string('J6','4')
worksheet11.write_string('K6','5')
worksheet11.write_string('L6','6')
worksheet11.write_string('N6','1')
worksheet11.write_string('O6','2')
worksheet11.write_string('P6','3')
worksheet11.write_string('Q6','4')
worksheet11.write_string('R6','5')
worksheet11.write_string('S6','6')
worksheet11.write_string('U6','1')
worksheet11.write_string('V6','2')
worksheet11.write_string('W6','3')
worksheet11.write_string('X6','4')
worksheet11.write_string('Y6','5')
worksheet11.write_string('Z6','6')
worksheet11.write_string('E7','W141617235292')
worksheet11.write_string('E8','W141617555914')
worksheet11.write_string('E9','W141618750091')
worksheet11.write_string('E10','W141618591839')
worksheet11.write_string('E11','W141618133512')
worksheet11.write_string('E12','W141618154440')
worksheet11.write_string('E13','W137518119128')
worksheet11.write_string('E14','W141619730500')
worksheet11.write_string('E15','W141618534656')
worksheet11.write_string('E16','W141618675293')

worksheet11.write_string('F17','Difference High - Low')
worksheet11.write_string('F18', 'Frequency')
worksheet11.write_string('F19','4')
worksheet11.write_string('F20','3')
worksheet11.write_string('F21','2')
worksheet11.write_string('F22','1')
worksheet11.write_string('F23','0.5')
worksheet11.write_string('F24','0')
worksheet11.write_string('M17','Difference High - Low')
worksheet11.write_string('M18', 'Frequency')
worksheet11.write_string('M19','4')
worksheet11.write_string('M20','3')
worksheet11.write_string('M21','2')
worksheet11.write_string('M22','1')
worksheet11.write_string('M23','0.5')
worksheet11.write_string('M24','0')
worksheet11.write_string('T17','Difference High - Low')
worksheet11.write_string('T18', 'Frequency')
worksheet11.write_string('T19','4')
worksheet11.write_string('T20','3')
worksheet11.write_string('T21','2')
worksheet11.write_string('T22','1')
worksheet11.write_string('T23','0.5')
worksheet11.write_string('T24','0')
worksheet11.write('H19','=IF((COUNTIF(G7:L16,"=4")/60)=0,"",COUNTIF(G7:L16,"=4")/60)',percent_format)
worksheet11.write('H20','=IF((COUNTIF(G7:L16,"=3")/60)=0,"",COUNTIF(G7:L16,"=3")/60)',percent_format)
worksheet11.write('H21','=IF((COUNTIF(G7:L16,"=2")/60)=0,"",COUNTIF(G7:L16,"=2")/60)',percent_format)
worksheet11.write('H22','=IF((COUNTIF(G7:L16,"=1")/60)=0,"",COUNTIF(G7:L16,"=1")/60)',percent_format)
worksheet11.write('H23','=IF((COUNTIF(G7:L16,"=0.5")/60)=0,"",COUNTIF(G7:L16,"=0.5")/60)',percent_format)
worksheet11.write('H24','=IF((COUNTIF(G7:L16,"=0")/60)=0,"",COUNTIF(G7:L16,"=0")/60)',percent_format)

worksheet11.write('O19','=IF((COUNTIF(N7:S16,"=4")/60)=0,"",COUNTIF(N7:S16,"=4")/60)',percent_format)
worksheet11.write('O20','=IF((COUNTIF(N7:S16,"=3")/60)=0,"",COUNTIF(N7:S16,"=3")/60)',percent_format)
worksheet11.write('O21','=IF((COUNTIF(N7:S16,"=2")/60)=0,"",COUNTIF(N7:S16,"=2")/60)',percent_format)
worksheet11.write('O22','=IF((COUNTIF(N7:S16,"=1")/60)=0,"",COUNTIF(N7:S16,"=1")/60)',percent_format)
worksheet11.write('O23','=IF((COUNTIF(N7:S16,"=0.5")/60)=0,"",COUNTIF(N7:S16,"=0.5")/60)',percent_format)
worksheet11.write('O24','=IF((COUNTIF(N7:S16,"=0")/60)=0,"",COUNTIF(N7:S16,"=0")/60)',percent_format)

worksheet11.write('V19','=IF((COUNTIF(U7:Z16,"=4")/60)=0,"",COUNTIF(U7:Z16,"=4")/60)',percent_format)
worksheet11.write('V20','=IF((COUNTIF(U7:Z16,"=3")/60)=0,"",COUNTIF(U7:Z16,"=3")/60)',percent_format)
worksheet11.write('V21','=IF((COUNTIF(U7:Z16,"=2")/60)=0,"",COUNTIF(U7:Z16,"=2")/60)',percent_format)
worksheet11.write('V22','=IF((COUNTIF(U7:Z16,"=1")/60)=0,"",COUNTIF(U7:Z16,"=1")/60)',percent_format)
worksheet11.write('V23','=IF((COUNTIF(U7:Z16,"=0.5")/60)=0,"",COUNTIF(U7:Z16,"=0.5")/60)',percent_format)
worksheet11.write('V24','=IF((COUNTIF(U7:Z16,"=0")/60)=0,"",COUNTIF(U7:Z16,"=0")/60)',percent_format)




            
worksheet1.set_column('A:G', 15)
worksheet2.set_column('A:G', 15)
worksheet3.set_column('A:G', 15)
worksheet4.set_column('A:G', 15)
worksheet5.set_column('A:G', 15)
worksheet6.set_column('A:G', 15)
worksheet7.set_column('A:G', 15)
worksheet8.set_column('A:G', 15)
worksheet9.set_column('A:G', 15)
worksheet10.set_column('A:G', 15)

mm_writer.save()
    

output_manual = r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_NaCl_Prophylax_split_v5.xlsx"

mm_writer = pd.ExcelWriter(output_manual, engine = 'xlsxwriter')
Sam1_naclproph.to_excel(mm_writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_naclproph.to_excel(mm_writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_naclproph.to_excel(mm_writer, startcol =0, startrow =5,sheet_name='W141618675293')

workbook  = mm_writer.book
worksheet1 = mm_writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'NaCl + Prophylax') 
worksheet2 = mm_writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'NaCl + Prophylax') 
worksheet3 = mm_writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'NaCl + Prophylax') 
worksheet4 = mm_writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'NaCl + Prophylax') 
worksheet5 = mm_writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'NaCl + Prophylax') 
worksheet6 = mm_writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'NaCl + Prophylax') 
worksheet7 = mm_writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'NaCl + Prophylax') 
worksheet8 = mm_writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'NaCl + Prophylax') 
worksheet9 = mm_writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'NaCl + Prophylax') 
worksheet10 = mm_writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'NaCl + Prophylax') 



worksheet1.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet1.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet1.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet2.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet2.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet2.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet3.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet3.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet3.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet4.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet4.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet4.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet5.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet5.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet5.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet6.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet6.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet6.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet7.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet7.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet7.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet8.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet8.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet8.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet9.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet9.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet9.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet10.write_string('J5', 'Lot: ' + str(lotpeter_lissproph[0][0]))
worksheet10.write_string('Q5','Lot: ' + str(lotpeter_lissproph[1][0]))
worksheet10.write_string('X5','Lot: ' + str(lotpeter_lissproph[2][0]))

worksheet1.write_string(24,5,'Reaction')
worksheet1.write_string(25,5,'Highest')
worksheet1.write_string(26,5,'Lowest')
worksheet1.write_string(27,5,'Difference')
worksheet1.write_string(28,5,'4')
worksheet1.write_string(29,5,'3')
worksheet1.write_string(30,5,'2')
worksheet1.write_string(31,5,'1')
worksheet1.write_string(32,5,'0.5')
worksheet1.write_string(33,5,'0')

worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','H','I','J','K','L','N','O','P','Q','R','S','U','V','W','X','Y','Z']
# for y in range(1,11):

for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet1.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-21), '=W141617235292!'+j+str(i))
        if i == 29:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            

worksheet2.write_string(24,5,'Reaction')
worksheet2.write_string(25,5,'Highest')
worksheet2.write_string(26,5,'Lowest')
worksheet2.write_string(27,5,'Difference')
worksheet2.write_string(28,5,'4')
worksheet2.write_string(29,5,'3')
worksheet2.write_string(30,5,'2')
worksheet2.write_string(31,5,'1')
worksheet2.write_string(32,5,'0.5')
worksheet2.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet2.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-20), '=W141617555914!'+j+str(i))
        if i == 29:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet3.write_string(24,5,'Reaction')
worksheet3.write_string(25,5,'Highest')
worksheet3.write_string(26,5,'Lowest')
worksheet3.write_string(27,5,'Difference')
worksheet3.write_string(28,5,'4')
worksheet3.write_string(29,5,'3')
worksheet3.write_string(30,5,'2')
worksheet3.write_string(31,5,'1')
worksheet3.write_string(32,5,'0.5')
worksheet3.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet3.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-19), '=W141618750091!'+j+str(i))
        if i == 29:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet4.write_string(24,5,'Reaction')
worksheet4.write_string(25,5,'Highest')
worksheet4.write_string(26,5,'Lowest')
worksheet4.write_string(27,5,'Difference')
worksheet4.write_string(28,5,'4')
worksheet4.write_string(29,5,'3')
worksheet4.write_string(30,5,'2')
worksheet4.write_string(31,5,'1')
worksheet4.write_string(32,5,'0.5')
worksheet4.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet4.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-18), '=W141618591839!'+j+str(i))
        if i == 29:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet5.write_string(24,5,'Reaction')
worksheet5.write_string(25,5,'Highest')
worksheet5.write_string(26,5,'Lowest')
worksheet5.write_string(27,5,'Difference')
worksheet5.write_string(28,5,'4')
worksheet5.write_string(29,5,'3')
worksheet5.write_string(30,5,'2')
worksheet5.write_string(31,5,'1')
worksheet5.write_string(32,5,'0.5')
worksheet5.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet5.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-17), '=W141618133512!'+j+str(i))
        if i == 29:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(24,5,'Reaction')
worksheet6.write_string(25,5,'Highest')
worksheet6.write_string(26,5,'Lowest')
worksheet6.write_string(27,5,'Difference')
worksheet6.write_string(28,5,'4')
worksheet6.write_string(29,5,'3')
worksheet6.write_string(30,5,'2')
worksheet6.write_string(31,5,'1')
worksheet6.write_string(32,5,'0.5')
worksheet6.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet6.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-16), '=W141618154440!'+j+str(i))
        if i == 29:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet7.write_string(24,5,'Reaction')
worksheet7.write_string(25,5,'Highest')
worksheet7.write_string(26,5,'Lowest')
worksheet7.write_string(27,5,'Difference')
worksheet7.write_string(28,5,'4')
worksheet7.write_string(29,5,'3')
worksheet7.write_string(30,5,'2')
worksheet7.write_string(31,5,'1')
worksheet7.write_string(32,5,'0.5')
worksheet7.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet7.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-15), '=W137518119128!'+j+str(i))
        if i == 29:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet8.write_string(24,5,'Reaction')
worksheet8.write_string(25,5,'Highest')
worksheet8.write_string(26,5,'Lowest')
worksheet8.write_string(27,5,'Difference')
worksheet8.write_string(28,5,'4')
worksheet8.write_string(29,5,'3')
worksheet8.write_string(30,5,'2')
worksheet8.write_string(31,5,'1')
worksheet8.write_string(32,5,'0.5')
worksheet8.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet8.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-14), '=W141619730500!'+j+str(i))
        if i == 29:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet9.write_string(24,5,'Reaction')
worksheet9.write_string(25,5,'Highest')
worksheet9.write_string(26,5,'Lowest')
worksheet9.write_string(27,5,'Difference')
worksheet9.write_string(28,5,'4')
worksheet9.write_string(29,5,'3')
worksheet9.write_string(30,5,'2')
worksheet9.write_string(31,5,'1')
worksheet9.write_string(32,5,'0.5')
worksheet9.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet9.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-13), '=W141618534656!'+j+str(i))
        if i == 29:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(24,5,'Reaction')
worksheet10.write_string(25,5,'Highest')
worksheet10.write_string(26,5,'Lowest')
worksheet10.write_string(27,5,'Difference')
worksheet10.write_string(28,5,'4')
worksheet10.write_string(29,5,'3')
worksheet10.write_string(30,5,'2')
worksheet10.write_string(31,5,'1')
worksheet10.write_string(32,5,'0.5')
worksheet10.write_string(33,5,'0')
for i in range(26,35):
    for j in alpha:
        print(j +str(i))
        if i ==26:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==27:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==28:
            worksheet10.write(j+str(i), '='+j+str(26) +'-'+j+str(27))
            worksheet11.write(j+str(i-12), '=W141618675293!'+j+str(i))
        if i == 29:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 30:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 31:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet11.write_string('K5','Panel Members')
worksheet11.write_string('R5','Panel Members')
worksheet11.write_string('Y5','Panel Members')
worksheet11.write_string('G6','1')
worksheet11.write_string('H6','2')
worksheet11.write_string('I6','3')
worksheet11.write_string('J6','4')
worksheet11.write_string('K6','5')
worksheet11.write_string('L6','6')
worksheet11.write_string('N6','1')
worksheet11.write_string('O6','2')
worksheet11.write_string('P6','3')
worksheet11.write_string('Q6','4')
worksheet11.write_string('R6','5')
worksheet11.write_string('S6','6')
worksheet11.write_string('U6','1')
worksheet11.write_string('V6','2')
worksheet11.write_string('W6','3')
worksheet11.write_string('X6','4')
worksheet11.write_string('Y6','5')
worksheet11.write_string('Z6','6')
worksheet11.write_string('E7','W141617235292')
worksheet11.write_string('E8','W141617555914')
worksheet11.write_string('E9','W141618750091')
worksheet11.write_string('E10','W141618591839')
worksheet11.write_string('E11','W141618133512')
worksheet11.write_string('E12','W141618154440')
worksheet11.write_string('E13','W137518119128')
worksheet11.write_string('E14','W141619730500')
worksheet11.write_string('E15','W141618534656')
worksheet11.write_string('E16','W141618675293')

worksheet11.write_string('F17','Difference High - Low')
worksheet11.write_string('F18', 'Frequency')
worksheet11.write_string('F19','4')
worksheet11.write_string('F20','3')
worksheet11.write_string('F21','2')
worksheet11.write_string('F22','1')
worksheet11.write_string('F23','0.5')
worksheet11.write_string('F24','0')
worksheet11.write_string('M17','Difference High - Low')
worksheet11.write_string('M18', 'Frequency')
worksheet11.write_string('M19','4')
worksheet11.write_string('M20','3')
worksheet11.write_string('M21','2')
worksheet11.write_string('M22','1')
worksheet11.write_string('M23','0.5')
worksheet11.write_string('M24','0')
worksheet11.write_string('T17','Difference High - Low')
worksheet11.write_string('T18', 'Frequency')
worksheet11.write_string('T19','4')
worksheet11.write_string('T20','3')
worksheet11.write_string('T21','2')
worksheet11.write_string('T22','1')
worksheet11.write_string('T23','0.5')
worksheet11.write_string('T24','0')
worksheet11.write('H19','=IF((COUNTIF(G7:L16,"=4")/60)=0,"",COUNTIF(G7:L16,"=4")/60)',percent_format)
worksheet11.write('H20','=IF((COUNTIF(G7:L16,"=3")/60)=0,"",COUNTIF(G7:L16,"=3")/60)',percent_format)
worksheet11.write('H21','=IF((COUNTIF(G7:L16,"=2")/60)=0,"",COUNTIF(G7:L16,"=2")/60)',percent_format)
worksheet11.write('H22','=IF((COUNTIF(G7:L16,"=1")/60)=0,"",COUNTIF(G7:L16,"=1")/60)',percent_format)
worksheet11.write('H23','=IF((COUNTIF(G7:L16,"=0.5")/60)=0,"",COUNTIF(G7:L16,"=0.5")/60)',percent_format)
worksheet11.write('H24','=IF((COUNTIF(G7:L16,"=0")/60)=0,"",COUNTIF(G7:L16,"=0")/60)',percent_format)

worksheet11.write('O19','=IF((COUNTIF(N7:S16,"=4")/60)=0,"",COUNTIF(N7:S16,"=4")/60)',percent_format)
worksheet11.write('O20','=IF((COUNTIF(N7:S16,"=3")/60)=0,"",COUNTIF(N7:S16,"=3")/60)',percent_format)
worksheet11.write('O21','=IF((COUNTIF(N7:S16,"=2")/60)=0,"",COUNTIF(N7:S16,"=2")/60)',percent_format)
worksheet11.write('O22','=IF((COUNTIF(N7:S16,"=1")/60)=0,"",COUNTIF(N7:S16,"=1")/60)',percent_format)
worksheet11.write('O23','=IF((COUNTIF(N7:S16,"=0.5")/60)=0,"",COUNTIF(N7:S16,"=0.5")/60)',percent_format)
worksheet11.write('O24','=IF((COUNTIF(N7:S16,"=0")/60)=0,"",COUNTIF(N7:S16,"=0")/60)',percent_format)

worksheet11.write('V19','=IF((COUNTIF(U7:Z16,"=4")/60)=0,"",COUNTIF(U7:Z16,"=4")/60)',percent_format)
worksheet11.write('V20','=IF((COUNTIF(U7:Z16,"=3")/60)=0,"",COUNTIF(U7:Z16,"=3")/60)',percent_format)
worksheet11.write('V21','=IF((COUNTIF(U7:Z16,"=2")/60)=0,"",COUNTIF(U7:Z16,"=2")/60)',percent_format)
worksheet11.write('V22','=IF((COUNTIF(U7:Z16,"=1")/60)=0,"",COUNTIF(U7:Z16,"=1")/60)',percent_format)
worksheet11.write('V23','=IF((COUNTIF(U7:Z16,"=0.5")/60)=0,"",COUNTIF(U7:Z16,"=0.5")/60)',percent_format)
worksheet11.write('V24','=IF((COUNTIF(U7:Z16,"=0")/60)=0,"",COUNTIF(U7:Z16,"=0")/60)',percent_format)




            
worksheet1.set_column('A:G', 15)
worksheet2.set_column('A:G', 15)
worksheet3.set_column('A:G', 15)
worksheet4.set_column('A:G', 15)
worksheet5.set_column('A:G', 15)
worksheet6.set_column('A:G', 15)
worksheet7.set_column('A:G', 15)
worksheet8.set_column('A:G', 15)
worksheet9.set_column('A:G', 15)
worksheet10.set_column('A:G', 15)

mm_writer.save()
    



def build_pool(dfpool, lotnumber):
    dfpool.fillna(value=pd.np.nan, inplace =True)
    dfpool.replace('0,5','0.5', inplace = True)
    dfpool.replace('0.5',0.5, inplace = True)
    
    dfpool.replace('Run 2', np.nan, inplace = True)
    dfpool.replace('*',np.nan, inplace = True)
    dfpool.replace('x',np.nan, inplace = True)
    dfpool.replace('Run 1', np.nan, inplace = True)
    dfpool.replace('Identifier',np.nan, inplace =True)
    dfpool.replace('Sample',np.nan, inplace =True)
    dfpool = dfpool.dropna(axis =1, how ='all')
    dfpool = dfpool.dropna()
    dfpool = dfpool.reset_index(drop = True)
    dfpool['index'] = dfpool.index.tolist()


    ###################removing columns and converting None and Run 2 and Identifier to NaN to help with deletion of rows####################

    ###############sets POOL Dataframes AM PM#######################################
    dfpool.loc[dfpool.index <= 9 , 'AM-PM'] = 'AM'
    dfpool.loc[dfpool.index > 9  , 'AM-PM'] = 'PM'
  
    ############# reset index after cleaning #################

    # df1pool['Site'] = 'Peterborough'

    dfpool.sort_values(by = 2, inplace = True)

    dfpool = dfpool.reset_index(drop = True)

    newheaders ={0:'Identifier', 2:'Sample', 3:'Lot1: '+str(lotnumber[0][0]),4:'Lot2: '+str(lotnumber[1][0]),5:'Lot3: '+str(lotnumber[2][0]),8:'Identifier',10:'Sample',
                 11:'Lot1: '+str(lotnumber[0][0]),12:'Lot2: '+str(lotnumber[1][0]),13:'Lot3: '+str(lotnumber[2][0]),16:'Identifier',18:'Sample',
                 19:'Lot1: '+str(lotnumber[0][0]),20:'Lot2: '+str(lotnumber[1][0]),21:'Lot3: '+str(lotnumber[2][0])}

    dfpool.rename(columns = newheaders, inplace = True)

    day1 = dfpool.iloc[:,[0,1,2,3,4,15,16]]
    day2 = dfpool.iloc[:, [5,6,7,8,9,15,16]]
    day3 = dfpool.iloc[:,[10,11,12,13,14,15,16]]    

    day3['Day'] = 3
    day2['Day'] = 2
    day1['Day'] = 1
    day1_day2 = pd.concat([day1,day2], sort = False)

    day1_day2_day3 = pd.concat([day1_day2, day3], sort = False)

    day1_day2_day3.sort_values(by =['Sample'], inplace = True)

    return day1_day2_day3

#############build the three pools the concat all############# Then order by Identifier and seperate out into own DF that will go into workbook for strength differences


dfpeter_poolnacl = build_pool(dfpeter_poolnacl, lotpeter_poolnacl)

dfpeter_lisspool= build_pool(dfpeter_lisspool, lotpeter_lisspool)

dfpeter_iggpool = build_pool(dfpeter_iggpool, lotpeter_iggpool)

dfgenhos_poolnacl = build_pool(dfgenhos_poolnacl, lotgenhos_poolnacl)

dfgenhos_lisspool= build_pool(dfgenhos_lisspool, lotgenhos_lisspool)

dfgenhos_iggpool = build_pool(dfgenhos_iggpool, lotgenhos_iggpool)

dfbiorad_poolnacl = build_pool(dfbiorad_poolnacl, lotbiorad_poolnacl)

dfbiorad_lisspool= build_pool(dfbiorad_lisspool, lotbiorad_lisspool)

dfbiorad_iggpool = build_pool(dfbiorad_iggpool, lotbiorad_iggpool)






def peter_dfpool(dfpool):
#     df.drop('index', axis = 1, inplace = True)
    dfpool= dfpool.reset_index(drop = True)
    dfpool['Site'] = 'Peterborough'
    dfpool = dfpool[['Site','Day', 'AM-PM', 'Identifier','Sample', 'Lot1: 6070341','Lot2: 6070342','Lot3: 6070361']]
    dfpool = dfpool.sort_values(by=['Site', 'Day'])
    return dfpool
dfpeter_poolnacl = peter_dfpool(dfpeter_poolnacl)
dfpeter_lisspool = peter_dfpool(dfpeter_lisspool)
dfpeter_iggpool = peter_dfpool(dfpeter_iggpool)

def genhosp_dfpool(dfpool):
#     df.drop('index', axis = 1, inplace = True)
    dfpool= dfpool.reset_index(drop = True)
    dfpool['Site'] = 'Site Alexandra Marine & Gen Hosp'
    dfpool = dfpool[['Site','Day', 'AM-PM', 'Identifier','Sample', 'Lot1: 6070341','Lot2: 6070342','Lot3: 6070361']]
    dfpool = dfpool.sort_values(by=['Site', 'Day'])
    return dfpool
dfgenhos_poolnacl = genhosp_dfpool(dfgenhos_poolnacl)
dfgenhos_lisspool = genhosp_dfpool(dfgenhos_lisspool)
dfgenhos_iggpool = genhosp_dfpool(dfgenhos_iggpool)

def biorad_dfpool(dfpool):
#     df.drop('index', axis = 1, inplace = True)
    dfpool= dfpool.reset_index(drop = True)
    dfpool['Site'] = 'Site Bio-Rad'
    dfpool = dfpool[['Site','Day', 'AM-PM', 'Identifier','Sample', 'Lot1: 6070341','Lot2: 6070342','Lot3: 6070361']]
    dfpool = dfpool.sort_values(by=['Site', 'Day'])
    return dfpool
dfbiorad_poolnacl = biorad_dfpool(dfbiorad_poolnacl)
dfbiorad_lisspool = biorad_dfpool(dfbiorad_lisspool)
dfbiorad_iggpool = biorad_dfpool(dfbiorad_iggpool)




def make_pool(df):
    df_Sam1 = df[df.Sample == 1.0]
    df_Sam1.insert(6, 'blank','')
    df_Sam1.insert(8, 'blank2','')
    
    df_Sam2 = df[df.Sample == 2.0]
    df_Sam2.insert(6, 'blank','')
    df_Sam2.insert(8, 'blank2','')
    
    df_Sam3 = df[df.Sample == 3.0]
    df_Sam3.insert(6, 'blank','')
    df_Sam3.insert(8, 'blank2','')
    
    df_Sam4 = df[df.Sample == 4.0]
    df_Sam4.insert(6, 'blank','')
    df_Sam4.insert(8, 'blank2','')
    
    df_Sam5 = df[df.Sample == 5.0]
    df_Sam5.insert(6, 'blank','')
    df_Sam5.insert(8, 'blank2','')
    
    df_Sam6 = df[df.Sample == 6.0]
    df_Sam6.insert(6, 'blank','')
    df_Sam6.insert(8, 'blank2','')
    
    df_Sam7 = df[df.Sample == 7.0]
    df_Sam7.insert(6, 'blank','')
    df_Sam7.insert(8, 'blank2','')
    
    df_Sam8 = df[df.Sample == 8.0]
    df_Sam8.insert(6, 'blank','')
    df_Sam8.insert(8, 'blank2','')
    
    df_Sam9 = df[df.Sample == 9.0]
    df_Sam9.insert(6, 'blank','')
    df_Sam9.insert(8, 'blank2','')
    
    df_Sam10 = df[df.Sample == 10.0]
    df_Sam10.insert(6, 'blank','')
    df_Sam10.insert(8, 'blank2','')
    return df_Sam1,df_Sam2,df_Sam3,df_Sam4,df_Sam5,df_Sam6,df_Sam7,df_Sam8,df_Sam9,df_Sam10





dfpeter_poolnacl_Sam1,dfpeter_poolnacl_Sam2,dfpeter_poolnacl_Sam3,dfpeter_poolnacl_Sam4,dfpeter_poolnacl_Sam5,dfpeter_poolnacl_Sam6,dfpeter_poolnacl_Sam7,dfpeter_poolnacl_Sam8,dfpeter_poolnacl_Sam9,dfpeter_poolnacl_Sam10 = make_pool(dfpeter_poolnacl)

dfpeter_lisspool_Sam1,dfpeter_lisspool_Sam2,dfpeter_lisspool_Sam3,dfpeter_lisspool_Sam4,dfpeter_lisspool_Sam5,dfpeter_lisspool_Sam6,dfpeter_lisspool_Sam7,dfpeter_lisspool_Sam8,dfpeter_lisspool_Sam9,dfpeter_lisspool_Sam10 = make_pool(dfpeter_lisspool)

dfpeter_iggpool_Sam1,dfpeter_iggpool_Sam2,dfpeter_iggpool_Sam3,dfpeter_iggpool_Sam4,dfpeter_iggpool_Sam5,dfpeter_iggpool_Sam6,dfpeter_iggpool_Sam7,dfpeter_iggpool_Sam8,dfpeter_iggpool_Sam9,dfpeter_iggpool_Sam10 = make_pool(dfpeter_iggpool)



dfgenhos_poolnacl_Sam1,dfgenhos_poolnacl_Sam2,dfgenhos_poolnacl_Sam3,dfgenhos_poolnacl_Sam4,dfgenhos_poolnacl_Sam5,dfgenhos_poolnacl_Sam6,dfgenhos_poolnacl_Sam7,dfgenhos_poolnacl_Sam8,dfgenhos_poolnacl_Sam9,dfgenhos_poolnacl_Sam10 = make_pool(dfgenhos_poolnacl)

dfgenhos_lisspool_Sam1,dfgenhos_lisspool_Sam2,dfgenhos_lisspool_Sam3,dfgenhos_lisspool_Sam4,dfgenhos_lisspool_Sam5,dfgenhos_lisspool_Sam6,dfgenhos_lisspool_Sam7,dfgenhos_lisspool_Sam8,dfgenhos_lisspool_Sam9,dfgenhos_lisspool_Sam10 = make_pool(dfgenhos_lisspool)

dfgenhos_iggpool_Sam1,dfgenhos_iggpool_Sam2,dfgenhos_iggpool_Sam3,dfgenhos_iggpool_Sam4,dfgenhos_iggpool_Sam5,dfgenhos_iggpool_Sam6,dfgenhos_iggpool_Sam7,dfgenhos_iggpool_Sam8,dfgenhos_iggpool_Sam9,dfgenhos_iggpool_Sam10 = make_pool(dfgenhos_iggpool)

dfbiorad_poolnacl_Sam1,dfbiorad_poolnacl_Sam2,dfbiorad_poolnacl_Sam3,dfbiorad_poolnacl_Sam4,dfbiorad_poolnacl_Sam5,dfbiorad_poolnacl_Sam6,dfbiorad_poolnacl_Sam7,dfbiorad_poolnacl_Sam8,dfbiorad_poolnacl_Sam9,dfbiorad_poolnacl_Sam10 = make_pool(dfbiorad_poolnacl)

dfbiorad_lisspool_Sam1,dfbiorad_lisspool_Sam2,dfbiorad_lisspool_Sam3,dfbiorad_lisspool_Sam4,dfbiorad_lisspool_Sam5,dfbiorad_lisspool_Sam6,dfbiorad_lisspool_Sam7,dfbiorad_lisspool_Sam8,dfbiorad_lisspool_Sam9,dfbiorad_lisspool_Sam10 = make_pool(dfbiorad_lisspool)

dfbiorad_iggpool_Sam1,dfbiorad_iggpool_Sam2,dfbiorad_iggpool_Sam3,dfbiorad_iggpool_Sam4,dfbiorad_iggpool_Sam5,dfbiorad_iggpool_Sam6,dfbiorad_iggpool_Sam7,dfbiorad_iggpool_Sam8,dfbiorad_iggpool_Sam9,dfbiorad_iggpool_Sam10 = make_pool(dfbiorad_iggpool)

def Ident_3(df1,df2,df3):
    two_combined = pd.concat([df1,df2], sort = False)
    three_combined = pd.concat([two_combined,df3], sort = False)
    three_combined = three_combined.reset_index(drop = True)
    return three_combined

Sam1_poolnacl = Ident_3(dfpeter_poolnacl_Sam1, dfgenhos_poolnacl_Sam1, dfbiorad_poolnacl_Sam1)

Sam1_lisspool = Ident_3(dfpeter_lisspool_Sam1, dfgenhos_lisspool_Sam1, dfbiorad_lisspool_Sam1)

Sam1_iggpool = Ident_3(dfpeter_iggpool_Sam1,dfgenhos_iggpool_Sam1,dfbiorad_iggpool_Sam1)


Sam2_poolnacl = Ident_3(dfpeter_poolnacl_Sam2, dfgenhos_poolnacl_Sam2, dfbiorad_poolnacl_Sam2)

Sam2_lisspool = Ident_3(dfpeter_lisspool_Sam2, dfgenhos_lisspool_Sam2, dfbiorad_lisspool_Sam2)

Sam2_iggpool = Ident_3(dfpeter_iggpool_Sam2,dfgenhos_iggpool_Sam2,dfbiorad_iggpool_Sam2)

Sam3_poolnacl = Ident_3(dfpeter_poolnacl_Sam3, dfgenhos_poolnacl_Sam3, dfbiorad_poolnacl_Sam3)

Sam3_lisspool = Ident_3(dfpeter_lisspool_Sam3, dfgenhos_lisspool_Sam3, dfbiorad_lisspool_Sam3)

Sam3_iggpool = Ident_3(dfpeter_iggpool_Sam3,dfgenhos_iggpool_Sam3,dfbiorad_iggpool_Sam3)

Sam4_poolnacl = Ident_3(dfpeter_poolnacl_Sam4, dfgenhos_poolnacl_Sam4, dfbiorad_poolnacl_Sam4)

Sam4_lisspool = Ident_3(dfpeter_lisspool_Sam4, dfgenhos_lisspool_Sam4, dfbiorad_lisspool_Sam4)

Sam4_iggpool = Ident_3(dfpeter_iggpool_Sam4,dfgenhos_iggpool_Sam4,dfbiorad_iggpool_Sam4)

Sam5_poolnacl = Ident_3(dfpeter_poolnacl_Sam5, dfgenhos_poolnacl_Sam5, dfbiorad_poolnacl_Sam5)

Sam5_lisspool = Ident_3(dfpeter_lisspool_Sam5, dfgenhos_lisspool_Sam5, dfbiorad_lisspool_Sam5)

Sam5_iggpool = Ident_3(dfpeter_iggpool_Sam5,dfgenhos_iggpool_Sam5,dfbiorad_iggpool_Sam5)

Sam6_poolnacl = Ident_3(dfpeter_poolnacl_Sam6, dfgenhos_poolnacl_Sam6, dfbiorad_poolnacl_Sam6)

Sam6_lisspool = Ident_3(dfpeter_lisspool_Sam6, dfgenhos_lisspool_Sam6, dfbiorad_lisspool_Sam6)

Sam6_iggpool = Ident_3(dfpeter_iggpool_Sam6,dfgenhos_iggpool_Sam6,dfbiorad_iggpool_Sam6)

Sam7_poolnacl = Ident_3(dfpeter_poolnacl_Sam7, dfgenhos_poolnacl_Sam7, dfbiorad_poolnacl_Sam7)

Sam7_lisspool = Ident_3(dfpeter_lisspool_Sam7, dfgenhos_lisspool_Sam7, dfbiorad_lisspool_Sam7)

Sam7_iggpool = Ident_3(dfpeter_iggpool_Sam7,dfgenhos_iggpool_Sam7,dfbiorad_iggpool_Sam7)

Sam8_poolnacl = Ident_3(dfpeter_poolnacl_Sam8, dfgenhos_poolnacl_Sam8, dfbiorad_poolnacl_Sam8)

Sam8_lisspool = Ident_3(dfpeter_lisspool_Sam8, dfgenhos_lisspool_Sam8, dfbiorad_lisspool_Sam8)

Sam8_iggpool = Ident_3(dfpeter_iggpool_Sam8,dfgenhos_iggpool_Sam8,dfbiorad_iggpool_Sam8)

Sam9_poolnacl = Ident_3(dfpeter_poolnacl_Sam9, dfgenhos_poolnacl_Sam9, dfbiorad_poolnacl_Sam9)

Sam9_lisspool = Ident_3(dfpeter_lisspool_Sam9, dfgenhos_lisspool_Sam9, dfbiorad_lisspool_Sam9)

Sam9_iggpool = Ident_3(dfpeter_iggpool_Sam9,dfgenhos_iggpool_Sam9,dfbiorad_iggpool_Sam9)

Sam10_poolnacl = Ident_3(dfpeter_poolnacl_Sam10, dfgenhos_poolnacl_Sam10, dfbiorad_poolnacl_Sam10)

Sam10_lisspool = Ident_3(dfpeter_lisspool_Sam10, dfgenhos_lisspool_Sam10, dfbiorad_lisspool_Sam10)

Sam10_iggpool = Ident_3(dfpeter_iggpool_Sam10,dfgenhos_iggpool_Sam10,dfbiorad_iggpool_Sam10)





output_manual_pool =  r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_NaCl_Pool_split_v5.xlsx"
writer = pd.ExcelWriter(output_manual_pool, engine = 'xlsxwriter')
Sam1_poolnacl.to_excel(writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_poolnacl.to_excel(writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_poolnacl.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618675293')


workbook  = writer.book
worksheet1 = writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'NaCl + Pool') 
worksheet2 = writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'NaCl + Pool') 
worksheet3 = writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'NaCl + Pool') 
worksheet4 = writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'NaCl + Pool') 
worksheet5 = writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'NaCl + Pool') 
worksheet6 = writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'NaCl + Pool') 
worksheet7 = writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'NaCl + Pool') 
worksheet8 = writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'NaCl + Pool') 
worksheet9 = writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'NaCl + Pool') 
worksheet10 = writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'NaCl + Pool') 

worksheet1.write_string(26,5,'Reaction')
worksheet1.write_string(27,5,'Highest')
worksheet1.write_string(28,5,'Lowest')
worksheet1.write_string(29,5,'Difference')
worksheet1.write_string(30,5,'4')
worksheet1.write_string(31,5,'3')
worksheet1.write_string(32,5,'2')
worksheet1.write_string(33,5,'1')
worksheet1.write_string(34,5,'0.5')
worksheet1.write_string(35,5,'0')



worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','I','K']
# for y in range(1,11):

for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet1.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-25), '=W141617235292!'+j+str(i))
        if i == 31:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            


worksheet2.write_string(26,5,'Reaction')
worksheet2.write_string(27,5,'Highest')
worksheet2.write_string(28,5,'Lowest')
worksheet2.write_string(29,5,'Difference')
worksheet2.write_string(30,5,'4')
worksheet2.write_string(31,5,'3')
worksheet2.write_string(32,5,'2')
worksheet2.write_string(33,5,'1')
worksheet2.write_string(34,5,'0.5')
worksheet2.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet2.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-24), '=W141617555914!'+j+str(i))
        if i == 31:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet3.write_string(26,5,'Reaction')
worksheet3.write_string(27,5,'Highest')
worksheet3.write_string(28,5,'Lowest')
worksheet3.write_string(29,5,'Difference')
worksheet3.write_string(30,5,'4')
worksheet3.write_string(31,5,'3')
worksheet3.write_string(32,5,'2')
worksheet3.write_string(33,5,'1')
worksheet3.write_string(34,5,'0.5')
worksheet3.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet3.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-23), '=W141618750091!'+j+str(i))
        if i == 31:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet4.write_string(26,5,'Reaction')
worksheet4.write_string(27,5,'Highest')
worksheet4.write_string(28,5,'Lowest')
worksheet4.write_string(29,5,'Difference')
worksheet4.write_string(30,5,'4')
worksheet4.write_string(31,5,'3')
worksheet4.write_string(32,5,'2')
worksheet4.write_string(33,5,'1')
worksheet4.write_string(34,5,'0.5')
worksheet4.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet4.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-22), '=W141618591839!'+j+str(i))
        if i == 31:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet5.write_string(26,5,'Reaction')
worksheet5.write_string(27,5,'Highest')
worksheet5.write_string(28,5,'Lowest')
worksheet5.write_string(29,5,'Difference')
worksheet5.write_string(30,5,'4')
worksheet5.write_string(31,5,'3')
worksheet5.write_string(32,5,'2')
worksheet5.write_string(33,5,'1')
worksheet5.write_string(34,5,'0.5')
worksheet5.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet5.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-21), '=W141618133512!'+j+str(i))
        if i == 31:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(26,5,'Reaction')
worksheet6.write_string(27,5,'Highest')
worksheet6.write_string(28,5,'Lowest')
worksheet6.write_string(29,5,'Difference')
worksheet6.write_string(30,5,'4')
worksheet6.write_string(31,5,'3')
worksheet6.write_string(32,5,'2')
worksheet6.write_string(33,5,'1')
worksheet6.write_string(34,5,'0.5')
worksheet6.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet6.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-20), '=W141618154440!'+j+str(i))
        if i == 31:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)





worksheet7.write_string(26,5,'Reaction')
worksheet7.write_string(27,5,'Highest')
worksheet7.write_string(28,5,'Lowest')
worksheet7.write_string(29,5,'Difference')
worksheet7.write_string(30,5,'4')
worksheet7.write_string(31,5,'3')
worksheet7.write_string(32,5,'2')
worksheet7.write_string(33,5,'1')
worksheet7.write_string(34,5,'0.5')
worksheet7.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet7.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-19), '=W137518119128!'+j+str(i))
        if i == 31:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet8.write_string(26,5,'Reaction')
worksheet8.write_string(27,5,'Highest')
worksheet8.write_string(28,5,'Lowest')
worksheet8.write_string(29,5,'Difference')
worksheet8.write_string(30,5,'4')
worksheet8.write_string(31,5,'3')
worksheet8.write_string(32,5,'2')
worksheet8.write_string(33,5,'1')
worksheet8.write_string(34,5,'0.5')
worksheet8.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet8.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-18), '=W141619730500!'+j+str(i))
        if i == 31:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet9.write_string(26,5,'Reaction')
worksheet9.write_string(27,5,'Highest')
worksheet9.write_string(28,5,'Lowest')
worksheet9.write_string(29,5,'Difference')
worksheet9.write_string(30,5,'4')
worksheet9.write_string(31,5,'3')
worksheet9.write_string(32,5,'2')
worksheet9.write_string(33,5,'1')
worksheet9.write_string(34,5,'0.5')
worksheet9.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet9.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-17), '=W141618534656!'+j+str(i))
        if i == 31:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(26,5,'Reaction')
worksheet10.write_string(27,5,'Highest')
worksheet10.write_string(28,5,'Lowest')
worksheet10.write_string(29,5,'Difference')
worksheet10.write_string(30,5,'4')
worksheet10.write_string(31,5,'3')
worksheet10.write_string(32,5,'2')
worksheet10.write_string(33,5,'1')
worksheet10.write_string(34,5,'0.5')
worksheet10.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet10.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-16), '=W141618675293!'+j+str(i))
        if i == 31:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)

worksheet11.write_string('G3','Lot 1: 6070341')
worksheet11.write_string('I3','Lot 2: 6070342')
worksheet11.write_string('K3','Lot 3: 6070361')

worksheet11.write_string('F5','W141617235292')
worksheet11.write_string('F6','W141617555914')
worksheet11.write_string('F7','W141618750091')
worksheet11.write_string('F8','W141618591839')
worksheet11.write_string('F9','W141618133512')
worksheet11.write_string('F10','W141618154440')
worksheet11.write_string('F11','W137518119128')
worksheet11.write_string('F12','W141619730500')
worksheet11.write_string('F13','W141618534656')
worksheet11.write_string('F14','W141618675293')




worksheet11.write_string('F16','Difference High - Low')
worksheet11.write_string('G16', 'Frequency')
worksheet11.write_string('F17','4')
worksheet11.write_string('F18','3')
worksheet11.write_string('F19','2')
worksheet11.write_string('F20','1')
worksheet11.write_string('F21','0.5')
worksheet11.write_string('F22','0')
worksheet11.write('G17','=IF((COUNTIF(G5:G14,"=4")/10)=0,"",COUNTIF(G5:G14,"=4")/10)',percent_format)
worksheet11.write('G18','=IF((COUNTIF(G5:G14,"=3")/10)=0,"",COUNTIF(G5:G14,"=3")/10)',percent_format)
worksheet11.write('G19','=IF((COUNTIF(G5:G14,"=2")/10)=0,"",COUNTIF(G5:G14,"=2")/10)',percent_format)
worksheet11.write('G20','=IF((COUNTIF(G5:G14,"=1")/10)=0,"",COUNTIF(G5:G14,"=1")/10)',percent_format)
worksheet11.write('G21','=IF((COUNTIF(G5:G14,"=0.5")/10)=0,"",COUNTIF(G5:G14,"=0.5")/10)',percent_format)
worksheet11.write('G22','=IF((COUNTIF(G5:G14,"=0")/10)=0,"",COUNTIF(G5:G14,"=0")/10)',percent_format)


worksheet11.write_string('I16', 'Frequency')
worksheet11.write_string('I17','4')
worksheet11.write_string('I18','3')
worksheet11.write_string('I19','2')
worksheet11.write_string('I20','1')
worksheet11.write_string('I21','0.5')
worksheet11.write_string('I22','0')
worksheet11.write('I17','=IF((COUNTIF(I5:I14,"=4")/10)=0,"",COUNTIF(I5:I14,"=4")/10)',percent_format)
worksheet11.write('I18','=IF((COUNTIF(I5:I14,"=3")/10)=0,"",COUNTIF(I5:I14,"=3")/10)',percent_format)
worksheet11.write('I19','=IF((COUNTIF(I5:I14,"=2")/10)=0,"",COUNTIF(I5:I14,"=2")/10)',percent_format)
worksheet11.write('I20','=IF((COUNTIF(I5:I14,"=1")/10)=0,"",COUNTIF(I5:I14,"=1")/10)',percent_format)
worksheet11.write('I21','=IF((COUNTIF(I5:I14,"=0.5")/10)=0,"",COUNTIF(I5:I14,"=0.5")/10)',percent_format)
worksheet11.write('I22','=IF((COUNTIF(I5:I14,"=0")/10)=0,"",COUNTIF(I5:I14,"=0")/10)',percent_format)


worksheet11.write_string('K16', 'Frequency')
worksheet11.write_string('K17','4')
worksheet11.write_string('K18','3')
worksheet11.write_string('K19','2')
worksheet11.write_string('K20','1')
worksheet11.write_string('K21','0.5')
worksheet11.write_string('K22','0')
worksheet11.write('K17','=IF((COUNTIF(K5:K14,"=4")/10)=0,"",COUNTIF(K5:K14,"=4")/10)',percent_format)
worksheet11.write('K18','=IF((COUNTIF(K5:K14,"=3")/10)=0,"",COUNTIF(K5:K14,"=3")/10)',percent_format)
worksheet11.write('K19','=IF((COUNTIF(K5:K14,"=2")/10)=0,"",COUNTIF(K5:K14,"=2")/10)',percent_format)
worksheet11.write('K20','=IF((COUNTIF(K5:K14,"=1")/10)=0,"",COUNTIF(K5:K14,"=1")/10)',percent_format)
worksheet11.write('K21','=IF((COUNTIF(K5:K14,"=0.5")/10)=0,"",COUNTIF(K5:K14,"=0.5")/10)',percent_format)
worksheet11.write('K22','=IF((COUNTIF(K5:K14,"=0")/10)=0,"",COUNTIF(K5:K14,"=0")/10)',percent_format)
   
worksheet1.set_column('A:L', 15)
worksheet2.set_column('A:L', 15)
worksheet3.set_column('A:L', 15)
worksheet4.set_column('A:L', 15)
worksheet5.set_column('A:L', 15)
worksheet6.set_column('A:L', 15)
worksheet7.set_column('A:L', 15)
worksheet8.set_column('A:L', 15)
worksheet9.set_column('A:L', 15)
worksheet10.set_column('A:L', 15)
worksheet11.set_column('E:F', 20)
writer.save()

output_manual_pool =  r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_Liss_Pool_split_v5.xlsx"


writer = pd.ExcelWriter(output_manual_pool, engine = 'xlsxwriter')
Sam1_lisspool.to_excel(writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_lisspool.to_excel(writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_lisspool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618675293')


workbook  = writer.book
worksheet1 = writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet2 = writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet3 = writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet4 = writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet5 = writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet6 = writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet7 = writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet8 = writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet9 = writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'IAT - LISS Coombs + Pool') 
worksheet10 = writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'IAT - LISS Coombs + Pool') 

worksheet1.write_string(26,5,'Reaction')
worksheet1.write_string(27,5,'Highest')
worksheet1.write_string(28,5,'Lowest')
worksheet1.write_string(29,5,'Difference')
worksheet1.write_string(30,5,'4')
worksheet1.write_string(31,5,'3')
worksheet1.write_string(32,5,'2')
worksheet1.write_string(33,5,'1')
worksheet1.write_string(34,5,'0.5')
worksheet1.write_string(35,5,'0')



worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','I','K']
# for y in range(1,11):

for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet1.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-25), '=W141617235292!'+j+str(i))
        if i == 31:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            


worksheet2.write_string(26,5,'Reaction')
worksheet2.write_string(27,5,'Highest')
worksheet2.write_string(28,5,'Lowest')
worksheet2.write_string(29,5,'Difference')
worksheet2.write_string(30,5,'4')
worksheet2.write_string(31,5,'3')
worksheet2.write_string(32,5,'2')
worksheet2.write_string(33,5,'1')
worksheet2.write_string(34,5,'0.5')
worksheet2.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet2.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-24), '=W141617555914!'+j+str(i))
        if i == 31:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet3.write_string(26,5,'Reaction')
worksheet3.write_string(27,5,'Highest')
worksheet3.write_string(28,5,'Lowest')
worksheet3.write_string(29,5,'Difference')
worksheet3.write_string(30,5,'4')
worksheet3.write_string(31,5,'3')
worksheet3.write_string(32,5,'2')
worksheet3.write_string(33,5,'1')
worksheet3.write_string(34,5,'0.5')
worksheet3.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet3.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-23), '=W141618750091!'+j+str(i))
        if i == 31:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet4.write_string(26,5,'Reaction')
worksheet4.write_string(27,5,'Highest')
worksheet4.write_string(28,5,'Lowest')
worksheet4.write_string(29,5,'Difference')
worksheet4.write_string(30,5,'4')
worksheet4.write_string(31,5,'3')
worksheet4.write_string(32,5,'2')
worksheet4.write_string(33,5,'1')
worksheet4.write_string(34,5,'0.5')
worksheet4.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet4.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-22), '=W141618591839!'+j+str(i))
        if i == 31:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet5.write_string(26,5,'Reaction')
worksheet5.write_string(27,5,'Highest')
worksheet5.write_string(28,5,'Lowest')
worksheet5.write_string(29,5,'Difference')
worksheet5.write_string(30,5,'4')
worksheet5.write_string(31,5,'3')
worksheet5.write_string(32,5,'2')
worksheet5.write_string(33,5,'1')
worksheet5.write_string(34,5,'0.5')
worksheet5.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet5.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-21), '=W141618133512!'+j+str(i))
        if i == 31:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(26,5,'Reaction')
worksheet6.write_string(27,5,'Highest')
worksheet6.write_string(28,5,'Lowest')
worksheet6.write_string(29,5,'Difference')
worksheet6.write_string(30,5,'4')
worksheet6.write_string(31,5,'3')
worksheet6.write_string(32,5,'2')
worksheet6.write_string(33,5,'1')
worksheet6.write_string(34,5,'0.5')
worksheet6.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet6.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-20), '=W141618154440!'+j+str(i))
        if i == 31:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)





worksheet7.write_string(26,5,'Reaction')
worksheet7.write_string(27,5,'Highest')
worksheet7.write_string(28,5,'Lowest')
worksheet7.write_string(29,5,'Difference')
worksheet7.write_string(30,5,'4')
worksheet7.write_string(31,5,'3')
worksheet7.write_string(32,5,'2')
worksheet7.write_string(33,5,'1')
worksheet7.write_string(34,5,'0.5')
worksheet7.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet7.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-19), '=W137518119128!'+j+str(i))
        if i == 31:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet8.write_string(26,5,'Reaction')
worksheet8.write_string(27,5,'Highest')
worksheet8.write_string(28,5,'Lowest')
worksheet8.write_string(29,5,'Difference')
worksheet8.write_string(30,5,'4')
worksheet8.write_string(31,5,'3')
worksheet8.write_string(32,5,'2')
worksheet8.write_string(33,5,'1')
worksheet8.write_string(34,5,'0.5')
worksheet8.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet8.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-18), '=W141619730500!'+j+str(i))
        if i == 31:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet9.write_string(26,5,'Reaction')
worksheet9.write_string(27,5,'Highest')
worksheet9.write_string(28,5,'Lowest')
worksheet9.write_string(29,5,'Difference')
worksheet9.write_string(30,5,'4')
worksheet9.write_string(31,5,'3')
worksheet9.write_string(32,5,'2')
worksheet9.write_string(33,5,'1')
worksheet9.write_string(34,5,'0.5')
worksheet9.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet9.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-17), '=W141618534656!'+j+str(i))
        if i == 31:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(26,5,'Reaction')
worksheet10.write_string(27,5,'Highest')
worksheet10.write_string(28,5,'Lowest')
worksheet10.write_string(29,5,'Difference')
worksheet10.write_string(30,5,'4')
worksheet10.write_string(31,5,'3')
worksheet10.write_string(32,5,'2')
worksheet10.write_string(33,5,'1')
worksheet10.write_string(34,5,'0.5')
worksheet10.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet10.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-16), '=W141618675293!'+j+str(i))
        if i == 31:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)
            
worksheet11.write_string('G3','Lot 1: 6070341')
worksheet11.write_string('I3','Lot 2: 6070342')
worksheet11.write_string('K3','Lot 3: 6070361')

worksheet11.write_string('F5','W141617235292')
worksheet11.write_string('F6','W141617555914')
worksheet11.write_string('F7','W141618750091')
worksheet11.write_string('F8','W141618591839')
worksheet11.write_string('F9','W141618133512')
worksheet11.write_string('F10','W141618154440')
worksheet11.write_string('F11','W137518119128')
worksheet11.write_string('F12','W141619730500')
worksheet11.write_string('F13','W141618534656')
worksheet11.write_string('F14','W141618675293')




worksheet11.write_string('F16','Difference High - Low')
worksheet11.write_string('G16', 'Frequency')
worksheet11.write_string('F17','4')
worksheet11.write_string('F18','3')
worksheet11.write_string('F19','2')
worksheet11.write_string('F20','1')
worksheet11.write_string('F21','0.5')
worksheet11.write_string('F22','0')
worksheet11.write('G17','=IF((COUNTIF(G5:G14,"=4")/10)=0,"",COUNTIF(G5:G14,"=4")/10)',percent_format)
worksheet11.write('G18','=IF((COUNTIF(G5:G14,"=3")/10)=0,"",COUNTIF(G5:G14,"=3")/10)',percent_format)
worksheet11.write('G19','=IF((COUNTIF(G5:G14,"=2")/10)=0,"",COUNTIF(G5:G14,"=2")/10)',percent_format)
worksheet11.write('G20','=IF((COUNTIF(G5:G14,"=1")/10)=0,"",COUNTIF(G5:G14,"=1")/10)',percent_format)
worksheet11.write('G21','=IF((COUNTIF(G5:G14,"=0.5")/10)=0,"",COUNTIF(G5:G14,"=0.5")/10)',percent_format)
worksheet11.write('G22','=IF((COUNTIF(G5:G14,"=0")/10)=0,"",COUNTIF(G5:G14,"=0")/10)',percent_format)


worksheet11.write_string('I16', 'Frequency')
worksheet11.write_string('I17','4')
worksheet11.write_string('I18','3')
worksheet11.write_string('I19','2')
worksheet11.write_string('I20','1')
worksheet11.write_string('I21','0.5')
worksheet11.write_string('I22','0')
worksheet11.write('I17','=IF((COUNTIF(I5:I14,"=4")/10)=0,"",COUNTIF(I5:I14,"=4")/10)',percent_format)
worksheet11.write('I18','=IF((COUNTIF(I5:I14,"=3")/10)=0,"",COUNTIF(I5:I14,"=3")/10)',percent_format)
worksheet11.write('I19','=IF((COUNTIF(I5:I14,"=2")/10)=0,"",COUNTIF(I5:I14,"=2")/10)',percent_format)
worksheet11.write('I20','=IF((COUNTIF(I5:I14,"=1")/10)=0,"",COUNTIF(I5:I14,"=1")/10)',percent_format)
worksheet11.write('I21','=IF((COUNTIF(I5:I14,"=0.5")/10)=0,"",COUNTIF(I5:I14,"=0.5")/10)',percent_format)
worksheet11.write('I22','=IF((COUNTIF(I5:I14,"=0")/10)=0,"",COUNTIF(I5:I14,"=0")/10)',percent_format)


worksheet11.write_string('K16', 'Frequency')
worksheet11.write_string('K17','4')
worksheet11.write_string('K18','3')
worksheet11.write_string('K19','2')
worksheet11.write_string('K20','1')
worksheet11.write_string('K21','0.5')
worksheet11.write_string('K22','0')
worksheet11.write('K17','=IF((COUNTIF(K5:K14,"=4")/10)=0,"",COUNTIF(K5:K14,"=4")/10)',percent_format)
worksheet11.write('K18','=IF((COUNTIF(K5:K14,"=3")/10)=0,"",COUNTIF(K5:K14,"=3")/10)',percent_format)
worksheet11.write('K19','=IF((COUNTIF(K5:K14,"=2")/10)=0,"",COUNTIF(K5:K14,"=2")/10)',percent_format)
worksheet11.write('K20','=IF((COUNTIF(K5:K14,"=1")/10)=0,"",COUNTIF(K5:K14,"=1")/10)',percent_format)
worksheet11.write('K21','=IF((COUNTIF(K5:K14,"=0.5")/10)=0,"",COUNTIF(K5:K14,"=0.5")/10)',percent_format)
worksheet11.write('K22','=IF((COUNTIF(K5:K14,"=0")/10)=0,"",COUNTIF(K5:K14,"=0")/10)',percent_format)
   
worksheet1.set_column('A:L', 15)
worksheet2.set_column('A:L', 15)
worksheet3.set_column('A:L', 15)
worksheet4.set_column('A:L', 15)
worksheet5.set_column('A:L', 15)
worksheet6.set_column('A:L', 15)
worksheet7.set_column('A:L', 15)
worksheet8.set_column('A:L', 15)
worksheet9.set_column('A:L', 15)
worksheet10.set_column('A:L', 15)
worksheet11.set_column('E:F', 20)

writer.save()

output_manual_pool =  r"G:/USRED_CLINICAL DATA/CLINTRIALS_IN_PROGRESS/IHD/Bio-Rad Canada Reproducibility Study/DM only/Output files from python/Canada_IgG_Pool_split_v5.xlsx"

writer = pd.ExcelWriter(output_manual_pool, engine = 'xlsxwriter')
Sam1_iggpool.to_excel(writer, startcol =0, startrow =5, sheet_name='W141617235292')
Sam2_iggpool.to_excel(writer,startcol = 0, startrow =5, sheet_name='W141617555914')
Sam3_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618750091')
Sam4_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618591839')
Sam5_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618133512')
Sam6_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618154440')
Sam7_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W137518119128')
Sam8_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141619730500')
Sam9_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618534656')
Sam10_iggpool.to_excel(writer, startcol =0, startrow =5,sheet_name='W141618675293')


workbook  = writer.book
worksheet1 = writer.sheets['W141617235292']
worksheet1.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet2 = writer.sheets['W141617555914']
worksheet2.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet3 = writer.sheets['W141618750091']
worksheet3.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet4 = writer.sheets['W141618591839']
worksheet4.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet5 = writer.sheets['W141618133512']
worksheet5.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet6 = writer.sheets['W141618154440']
worksheet6.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet7 = writer.sheets['W137518119128']
worksheet7.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet8 = writer.sheets['W141619730500']
worksheet8.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet9 = writer.sheets['W141618534656']
worksheet9.write_string(4, 4, 'IAT - Coombs IgG + Pool') 
worksheet10 = writer.sheets['W141618675293']
worksheet10.write_string(4, 4, 'IAT - Coombs IgG + Pool') 


worksheet1.write_string(26,5,'Reaction')
worksheet1.write_string(27,5,'Highest')
worksheet1.write_string(28,5,'Lowest')
worksheet1.write_string(29,5,'Difference')
worksheet1.write_string(30,5,'4')
worksheet1.write_string(31,5,'3')
worksheet1.write_string(32,5,'2')
worksheet1.write_string(33,5,'1')
worksheet1.write_string(34,5,'0.5')
worksheet1.write_string(35,5,'0')





worksheet11 = workbook.add_worksheet('Synopsis strength diff')
percent_format = workbook.add_format({'num_format': '0%'})
alpha = ['G','I','K']
# for y in range(1,11):

for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet1.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet1.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet1.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-25), '=W141617235292!'+j+str(i))
        if i == 31:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet1.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


            
            


worksheet2.write_string(26,5,'Reaction')
worksheet2.write_string(27,5,'Highest')
worksheet2.write_string(28,5,'Lowest')
worksheet2.write_string(29,5,'Difference')
worksheet2.write_string(30,5,'4')
worksheet2.write_string(31,5,'3')
worksheet2.write_string(32,5,'2')
worksheet2.write_string(33,5,'1')
worksheet2.write_string(34,5,'0.5')
worksheet2.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet2.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet2.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet2.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-24), '=W141617555914!'+j+str(i))
        if i == 31:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet2.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet3.write_string(26,5,'Reaction')
worksheet3.write_string(27,5,'Highest')
worksheet3.write_string(28,5,'Lowest')
worksheet3.write_string(29,5,'Difference')
worksheet3.write_string(30,5,'4')
worksheet3.write_string(31,5,'3')
worksheet3.write_string(32,5,'2')
worksheet3.write_string(33,5,'1')
worksheet3.write_string(34,5,'0.5')
worksheet3.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet3.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet3.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet3.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-23), '=W141618750091!'+j+str(i))
        if i == 31:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet3.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet4.write_string(26,5,'Reaction')
worksheet4.write_string(27,5,'Highest')
worksheet4.write_string(28,5,'Lowest')
worksheet4.write_string(29,5,'Difference')
worksheet4.write_string(30,5,'4')
worksheet4.write_string(31,5,'3')
worksheet4.write_string(32,5,'2')
worksheet4.write_string(33,5,'1')
worksheet4.write_string(34,5,'0.5')
worksheet4.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet4.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet4.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet4.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-22), '=W141618591839!'+j+str(i))
        if i == 31:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet4.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet5.write_string(26,5,'Reaction')
worksheet5.write_string(27,5,'Highest')
worksheet5.write_string(28,5,'Lowest')
worksheet5.write_string(29,5,'Difference')
worksheet5.write_string(30,5,'4')
worksheet5.write_string(31,5,'3')
worksheet5.write_string(32,5,'2')
worksheet5.write_string(33,5,'1')
worksheet5.write_string(34,5,'0.5')
worksheet5.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet5.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet5.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet5.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-21), '=W141618133512!'+j+str(i))
        if i == 31:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet5.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet6.write_string(26,5,'Reaction')
worksheet6.write_string(27,5,'Highest')
worksheet6.write_string(28,5,'Lowest')
worksheet6.write_string(29,5,'Difference')
worksheet6.write_string(30,5,'4')
worksheet6.write_string(31,5,'3')
worksheet6.write_string(32,5,'2')
worksheet6.write_string(33,5,'1')
worksheet6.write_string(34,5,'0.5')
worksheet6.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet6.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet6.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet6.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-20), '=W141618154440!'+j+str(i))
        if i == 31:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet6.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)





worksheet7.write_string(26,5,'Reaction')
worksheet7.write_string(27,5,'Highest')
worksheet7.write_string(28,5,'Lowest')
worksheet7.write_string(29,5,'Difference')
worksheet7.write_string(30,5,'4')
worksheet7.write_string(31,5,'3')
worksheet7.write_string(32,5,'2')
worksheet7.write_string(33,5,'1')
worksheet7.write_string(34,5,'0.5')
worksheet7.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet7.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet7.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet7.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-19), '=W137518119128!'+j+str(i))
        if i == 31:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet7.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)




worksheet8.write_string(26,5,'Reaction')
worksheet8.write_string(27,5,'Highest')
worksheet8.write_string(28,5,'Lowest')
worksheet8.write_string(29,5,'Difference')
worksheet8.write_string(30,5,'4')
worksheet8.write_string(31,5,'3')
worksheet8.write_string(32,5,'2')
worksheet8.write_string(33,5,'1')
worksheet8.write_string(34,5,'0.5')
worksheet8.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet8.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet8.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet8.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-18), '=W141619730500!'+j+str(i))
        if i == 31:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet8.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)



worksheet9.write_string(26,5,'Reaction')
worksheet9.write_string(27,5,'Highest')
worksheet9.write_string(28,5,'Lowest')
worksheet9.write_string(29,5,'Difference')
worksheet9.write_string(30,5,'4')
worksheet9.write_string(31,5,'3')
worksheet9.write_string(32,5,'2')
worksheet9.write_string(33,5,'1')
worksheet9.write_string(34,5,'0.5')
worksheet9.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet9.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet9.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet9.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-17), '=W141618534656!'+j+str(i))
        if i == 31:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet9.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)


worksheet10.write_string(26,5,'Reaction')
worksheet10.write_string(27,5,'Highest')
worksheet10.write_string(28,5,'Lowest')
worksheet10.write_string(29,5,'Difference')
worksheet10.write_string(30,5,'4')
worksheet10.write_string(31,5,'3')
worksheet10.write_string(32,5,'2')
worksheet10.write_string(33,5,'1')
worksheet10.write_string(34,5,'0.5')
worksheet10.write_string(35,5,'0')
for i in range(28,37):
    for j in alpha:
        print(j +str(i))
        if i ==28:
            worksheet10.write(j+str(i),'=MAX('+j + str(7)+':'+j +str(24)+')')
        if i ==29:
            worksheet10.write(j+str(i),'=MIN('+j + str(7)+':'+j +str(24)+')')
        if i ==30:
            worksheet10.write(j+str(i), '='+j+str(28) +'-'+j+str(29))
            worksheet11.write(j+str(i-16), '=W141618675293!'+j+str(i))
        if i == 31:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=4"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=4"'+')'+'/'+str(18)+')',percent_format)
        if i == 32:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=3"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=3"'+')'+'/'+str(18)+')',percent_format)
        if i == 33:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=2"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=2"'+')'+'/'+str(18)+')',percent_format)
        if i == 34:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=1"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=1"'+')'+'/'+str(18)+')',percent_format)
        if i == 35:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0.5"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0.5"'+')'+'/'+str(18)+')',percent_format)
        if i == 36:
            worksheet10.write_formula(j+str(i),'=IF((COUNTIF('+j+str(7)+':'+j +str(24) + ','+'"=0"'+')' + '/'+ str(18) +')=0'+','+'""'+',COUNTIF('+j+str(7)+':'+ j+str(24)+','+'"=0"'+')'+'/'+str(18)+')',percent_format)
            
worksheet11.write_string('G3','Lot 1: 6070341')
worksheet11.write_string('I3','Lot 2: 6070342')
worksheet11.write_string('K3','Lot 3: 6070361')

worksheet11.write_string('F5','W141617235292')
worksheet11.write_string('F6','W141617555914')
worksheet11.write_string('F7','W141618750091')
worksheet11.write_string('F8','W141618591839')
worksheet11.write_string('F9','W141618133512')
worksheet11.write_string('F10','W141618154440')
worksheet11.write_string('F11','W137518119128')
worksheet11.write_string('F12','W141619730500')
worksheet11.write_string('F13','W141618534656')
worksheet11.write_string('F14','W141618675293')




worksheet11.write_string('F16','Difference High - Low')
worksheet11.write_string('G16', 'Frequency')
worksheet11.write_string('F17','4')
worksheet11.write_string('F18','3')
worksheet11.write_string('F19','2')
worksheet11.write_string('F20','1')
worksheet11.write_string('F21','0.5')
worksheet11.write_string('F22','0')
worksheet11.write('G17','=IF((COUNTIF(G5:G14,"=4")/10)=0,"",COUNTIF(G5:G14,"=4")/10)',percent_format)
worksheet11.write('G18','=IF((COUNTIF(G5:G14,"=3")/10)=0,"",COUNTIF(G5:G14,"=3")/10)',percent_format)
worksheet11.write('G19','=IF((COUNTIF(G5:G14,"=2")/10)=0,"",COUNTIF(G5:G14,"=2")/10)',percent_format)
worksheet11.write('G20','=IF((COUNTIF(G5:G14,"=1")/10)=0,"",COUNTIF(G5:G14,"=1")/10)',percent_format)
worksheet11.write('G21','=IF((COUNTIF(G5:G14,"=0.5")/10)=0,"",COUNTIF(G5:G14,"=0.5")/10)',percent_format)
worksheet11.write('G22','=IF((COUNTIF(G5:G14,"=0")/10)=0,"",COUNTIF(G5:G14,"=0")/10)',percent_format)


worksheet11.write_string('I16', 'Frequency')
worksheet11.write_string('I17','4')
worksheet11.write_string('I18','3')
worksheet11.write_string('I19','2')
worksheet11.write_string('I20','1')
worksheet11.write_string('I21','0.5')
worksheet11.write_string('I22','0')
worksheet11.write('I17','=IF((COUNTIF(I5:I14,"=4")/10)=0,"",COUNTIF(I5:I14,"=4")/10)',percent_format)
worksheet11.write('I18','=IF((COUNTIF(I5:I14,"=3")/10)=0,"",COUNTIF(I5:I14,"=3")/10)',percent_format)
worksheet11.write('I19','=IF((COUNTIF(I5:I14,"=2")/10)=0,"",COUNTIF(I5:I14,"=2")/10)',percent_format)
worksheet11.write('I20','=IF((COUNTIF(I5:I14,"=1")/10)=0,"",COUNTIF(I5:I14,"=1")/10)',percent_format)
worksheet11.write('I21','=IF((COUNTIF(I5:I14,"=0.5")/10)=0,"",COUNTIF(I5:I14,"=0.5")/10)',percent_format)
worksheet11.write('I22','=IF((COUNTIF(I5:I14,"=0")/10)=0,"",COUNTIF(I5:I14,"=0")/10)',percent_format)


worksheet11.write_string('K16', 'Frequency')
worksheet11.write_string('K17','4')
worksheet11.write_string('K18','3')
worksheet11.write_string('K19','2')
worksheet11.write_string('K20','1')
worksheet11.write_string('K21','0.5')
worksheet11.write_string('K22','0')
worksheet11.write('K17','=IF((COUNTIF(K5:K14,"=4")/10)=0,"",COUNTIF(K5:K14,"=4")/10)',percent_format)
worksheet11.write('K18','=IF((COUNTIF(K5:K14,"=3")/10)=0,"",COUNTIF(K5:K14,"=3")/10)',percent_format)
worksheet11.write('K19','=IF((COUNTIF(K5:K14,"=2")/10)=0,"",COUNTIF(K5:K14,"=2")/10)',percent_format)
worksheet11.write('K20','=IF((COUNTIF(K5:K14,"=1")/10)=0,"",COUNTIF(K5:K14,"=1")/10)',percent_format)
worksheet11.write('K21','=IF((COUNTIF(K5:K14,"=0.5")/10)=0,"",COUNTIF(K5:K14,"=0.5")/10)',percent_format)
worksheet11.write('K22','=IF((COUNTIF(K5:K14,"=0")/10)=0,"",COUNTIF(K5:K14,"=0")/10)',percent_format)
   
worksheet1.set_column('A:L', 15)
worksheet2.set_column('A:L', 15)
worksheet3.set_column('A:L', 15)
worksheet4.set_column('A:L', 15)
worksheet5.set_column('A:L', 15)
worksheet6.set_column('A:L', 15)
worksheet7.set_column('A:L', 15)
worksheet8.set_column('A:L', 15)
worksheet9.set_column('A:L', 15)
worksheet10.set_column('A:L', 15)
worksheet11.set_column('E:F', 20)

writer.save()


# In[2]:


print('Sheet')


# In[ ]:




